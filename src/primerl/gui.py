"""qPCR-focused GUI for primerl."""

from __future__ import annotations

import gzip
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import webbrowser
import warnings
from bisect import bisect_left, bisect_right
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import tkinter as tk
import tkinter.font as tkfont
from tkinter import messagebox, ttk
from urllib.parse import urlencode
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import ttkbootstrap as tb
from ttkbootstrap.constants import *

from .ensembl_adapter import (
    EnsemblError,
    EnsemblNoGeneFound,
    build_lookup_symbol_url,
    build_sequence_id_url,
    choose_preferred_transcript,
    extract_transcript_choices,
    fetch_json_with_transport,
    map_ensembl_seq_type,
)
from .io_fasta import read_first_fasta_sequence
from .mfeprimer_spec import (
    DEFAULT_SPEC_PARAMS_RAW,
    SPEC_PRESET_SOFT,
    SPEC_PRESET_STRICT,
    build_mfeprimer_spec_cmd,
    preset_from_spec_param_raw,
    resolve_spec_param_tokens,
    spec_param_raw_for_preset,
)
from .primer3_qpcr import (
    Primer3RunSettings,
    QpcrFilterSettings,
    clean_sequence,
    collect_qpcr_pairs_from_primer3,
    parse_primer3_kv_output,
    run_primer3_qpcr_output,
    sort_qpcr_pairs,
)
from .spidey_adapter import (
    analyze_spidey_output,
    build_spidey_args,
    extract_intron_exon_bounds,
    run_spidey_with_transport,
)
from .platform_compat import (
    BIN_EXT,
    IS_APPLE_SILICON,
    IS_WINDOWS,
    candidate_exec_names,
    open_file,
    subprocess_run,
)


COLUMNS = [
    ("sel", "Sel", 44),
    ("spec", "Spec Check", 88),
    ("f_seq", "Forward Primer", 220),
    ("f_pos", "Pos", 60),
    ("f_len", "Len", 50),
    ("f_tm", "Tm", 70),
    ("r_seq", "Reverse Primer", 220),
    ("r_pos", "Pos", 60),
    ("r_len", "Len", 50),
    ("r_tm", "Tm", 70),
    ("amp", "Amp", 70),
    ("pd", "Ext. dimer dG", 110),
    ("pd_full", "Full dimer dG", 110),
]

ENSEMBL_SPECIES = [
    "Homo_sapiens",
    "Mus_musculus",
    "Danio_rerio",
    "Rattus_norvegicus",
    "Drosophila_melanogaster",
    "Caenorhabditis_elegans",
]

ENSEMBL_TYPES = ["genomic", "cdna", "coding", "utr5", "utr3"]

BLAST_SPECIES_QUERIES: dict[str, str] = {
    "Any": "",
    "Human": "Homo sapiens[Organism]",
    "Mouse": "Mus musculus[Organism]",
    "Zebrafish": "Danio rerio[Organism]",
    "Drosophila": "Drosophila melanogaster[Organism]",
}

ENSEMBL_DB_SPECIES_CHOICES: list[tuple[str, str]] = [
    ("Zebrafish (Danio rerio)", "danio_rerio"),
    ("Mouse (Mus musculus)", "mus_musculus"),
    ("Rat (Rattus norvegicus)", "rattus_norvegicus"),
    ("Human (Homo sapiens)", "homo_sapiens"),
    ("Drosophila (D. melanogaster)", "drosophila_melanogaster"),
]

# Allow broader off-target amplicon detection in MFEprimer spec mode.
# Design constraints can stay narrow (for example 100-300), while spec checks
# still consider larger potential off-target products.
SPEC_OFFTARGET_MIN_AMP_SIZE_BP = 60
SPEC_OFFTARGET_MAX_AMP_SIZE_BP = 600

if IS_WINDOWS:
    BINARY_PROFILE_CHOICES = (
        "Upstream original src",
        "Clang znver2",
        "Clang znver4",
        "Clang x86-64-v3",
    )
elif IS_APPLE_SILICON:
    BINARY_PROFILE_CHOICES = (
        "Upstream original src",
        "Apple Silicon native",
    )
else:
    BINARY_PROFILE_CHOICES = ("Upstream original src",)

PAD_S = 8
PAD_M = 16
PAD_L = 24
PAD_XS = 4
PAD_XXS = 2
LIGHT_THEME = "flatly"
DARK_THEME = "darkly"

try:
    from ttkbootstrap.tooltip import ToolTip as _BootstrapToolTip
except Exception:  # pragma: no cover - tooltip support is optional
    _BootstrapToolTip = None


def _enable_windows_dpi_awareness() -> None:
    if os.name != "nt":
        return
    try:
        import ctypes

        user32 = ctypes.windll.user32
        shcore = getattr(ctypes.windll, "shcore", None)

        # Prefer per-monitor v2 when available (Windows 10+), then fall back.
        if hasattr(user32, "SetProcessDpiAwarenessContext"):
            if user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4)):
                return
            if user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-3)):
                return

        if shcore is not None and hasattr(shcore, "SetProcessDpiAwareness"):
            # 2 = PROCESS_PER_MONITOR_DPI_AWARE, 1 = SYSTEM_DPI_AWARE
            if shcore.SetProcessDpiAwareness(2) == 0:
                return
            if shcore.SetProcessDpiAwareness(1) == 0:
                return

        if hasattr(user32, "SetProcessDPIAware"):
            user32.SetProcessDPIAware()
    except Exception:
        # Best-effort; GUI should still start even if DPI APIs are unavailable.
        pass


def _existing_default(*candidates: str) -> str:
    for c in candidates:
        if c and Path(c).exists():
            return c
    return ""


def _primerl_root() -> Path:
    """Resolve primerl runtime root for both dev and installed layouts.

    Dev layout:
      <repo>/src/primerl/gui.py and assets in a sibling runtime-assets folder
    Installed layout:
      <install>/app/src/primerl/gui.py and assets in <install>/
    """
    app_root = Path(__file__).resolve().parents[2]
    candidates = [app_root.parent, app_root]
    try:
        # Dev layout: accept any direct child folder that looks like a runtime root.
        child_dirs = [p for p in app_root.iterdir() if p.is_dir()]
        candidates.extend(child_dirs)
        # PyInstaller one-dir often nests app resources under _internal/<AppName>/.
        for child in child_dirs:
            try:
                candidates.extend(p for p in child.iterdir() if p.is_dir())
            except OSError:
                continue
    except OSError:
        pass
    def _layout_score(root: Path) -> int:
        tool_bin = root / "tools" / "bin"
        runtime = root / "runtime"
        if not (tool_bin.is_dir() and runtime.is_dir()):
            return -1
        score = 0
        # Prefer populated runtime roots over empty scaffolding directories.
        try:
            score += min(20, sum(1 for _ in tool_bin.iterdir()))
        except OSError:
            pass
        if any((tool_bin / n).exists() for n in candidate_exec_names("primer3_core")):
            score += 50
        if any((tool_bin / n).exists() for n in candidate_exec_names("ntthal")):
            score += 20
        if any((tool_bin / n).exists() for n in candidate_exec_names("spidey")):
            score += 15
        if any((tool_bin / n).exists() for n in candidate_exec_names("mfeprimer")):
            score += 10
        if (runtime / "gui_settings.example.json").exists():
            score += 5
        # DBs may be intentionally absent in lightweight app bundles.
        if (root / "databases").is_dir():
            score += 5
        return score

    best_root = app_root
    best_score = -1
    seen: set[Path] = set()
    for c in candidates:
        if c in seen:
            continue
        seen.add(c)
        score = _layout_score(c)
        if score > best_score:
            best_score = score
            best_root = c
    if best_score >= 0:
        return best_root
    return app_root


def _default_user_data_root() -> Path:
    if IS_WINDOWS:
        local_appdata = str(os.environ.get("LOCALAPPDATA") or "").strip()
        if local_appdata:
            return Path(local_appdata) / "PrimeRL"
    return Path.home() / ".primerl"


def _primerl_data_root() -> Path:
    env_override = str(os.environ.get("PRIMERL_DATA_DIR") or "").strip()
    if env_override:
        return Path(env_override)
    resource_root = _primerl_root()
    # Portable mode keeps mutable data next to the application bundle.
    if (resource_root / ".primerl_portable").exists():
        return resource_root
    if getattr(sys, "frozen", False):
        return _default_user_data_root()
    return resource_root


def _primerl_resource_path(*parts: str) -> str:
    return str(_primerl_root() / Path(*parts))


def _primerl_data_path(*parts: str) -> str:
    return str(_primerl_data_root() / Path(*parts))


def _primerl_path(*parts: str) -> str:
    # Backward-compatible alias for read-only app resources.
    return _primerl_resource_path(*parts)


def _is_processor_feature_present(flag_id: int) -> bool:
    if not IS_WINDOWS:
        return False
    try:
        import ctypes

        return bool(ctypes.windll.kernel32.IsProcessorFeaturePresent(int(flag_id)))
    except Exception:
        return False


def _detect_best_binary_profile_for_cpu() -> tuple[str, str]:
    if IS_APPLE_SILICON:
        return "Apple Silicon native", "Apple Silicon (arm64) detected"
    if not IS_WINDOWS:
        return "Upstream original src", "non-Windows platform detected"
    # Windows PF constants: 40 = AVX2, 41 = AVX512F
    has_avx2 = _is_processor_feature_present(40)
    has_avx512f = _is_processor_feature_present(41)
    proc_id = str(os.environ.get("PROCESSOR_IDENTIFIER") or "").lower()
    is_amd = ("authenticamd" in proc_id) or ("amd" in proc_id)
    if is_amd and has_avx512f:
        return "Clang znver4", "AMD CPU with AVX-512 support detected"
    if is_amd and has_avx2:
        return "Clang znver2", "AMD CPU with AVX2 support detected"
    return "Clang x86-64-v3", "generic x86-64-v3 fallback"


def _infer_binary_profile_from_paths(primer3_path: str, ntthal_path: str) -> str:
    p3 = str(primer3_path or "").replace("\\", "/").lower()
    nt = str(ntthal_path or "").replace("\\", "/").lower()
    both = f"{p3} {nt}"
    if "/clang_profiles/apple_silicon/" in both:
        return "Apple Silicon native"
    if "/clang_profiles/znver4/" in both:
        return "Clang znver4"
    if "/clang_profiles/znver2/" in both:
        return "Clang znver2"
    if "/clang_profiles/x86_64_v3/" in both:
        return "Clang x86-64-v3"
    if "primer3_clang" in both:
        return "Clang x86-64-v3"
    return "Upstream original src"


def _tool_bin_roots() -> list[Path]:
    roots: list[Path] = []
    seen: set[Path] = set()

    def _add(path: Path) -> None:
        p = path.resolve()
        if p in seen:
            return
        seen.add(p)
        if p.is_dir():
            roots.append(p)

    resource_root = _primerl_root()
    _add(resource_root / "tools" / "bin")
    _add(resource_root.parent / "tools" / "bin")

    module_root = Path(__file__).resolve().parents[2]
    _add(module_root / "tools" / "bin")
    _add(module_root.parent / "tools" / "bin")

    if getattr(sys, "frozen", False):
        exe_contents = Path(sys.executable).resolve().parent.parent
        _add(exe_contents / "Resources" / "PrimeRL" / "tools" / "bin")
        _add(exe_contents / "Frameworks" / "PrimeRL" / "tools" / "bin")
        _add(exe_contents / "Resources" / "tools" / "bin")
        _add(exe_contents / "Frameworks" / "tools" / "bin")

    if roots:
        return roots

    return [Path(_primerl_path("tools", "bin"))]


def _tool_bin_exec_default(stem: str) -> str:
    for base in _tool_bin_roots():
        for name in candidate_exec_names(stem):
            p = base / name
            if p.exists() and p.is_file():
                return str(p)
    return ""


def _tool_profile_exec_default(profile_rel: str, stem: str) -> str:
    for bin_root in _tool_bin_roots():
        base = bin_root / "clang_profiles" / profile_rel
        for name in candidate_exec_names(stem):
            p = base / name
            if p.exists() and p.is_file():
                return str(p)
    return ""


def _path_exec_default(path: str) -> str:
    for name in candidate_exec_names(path):
        p = shutil.which(name)
        if p:
            return p
    return ""


def _pd_stub(_s1: str, _s2: str, full: bool) -> float:
    return -2.0 if not full else -5.0


def _seq_from_text(widget_text: str) -> str:
    lines = [ln.strip() for ln in widget_text.splitlines() if ln.strip() and not ln.startswith(">")]
    return clean_sequence("".join(lines))


def _sanitize_oligo_name_token(raw: str) -> str:
    token = (raw or "").strip()
    token = re.sub(r"\s+", "_", token)
    token = re.sub(r"[^A-Za-z0-9_-]", "", token)
    return token or "gene"


def _resolve_microsynth_template_path() -> Path:
    rel = Path("third_party") / "Order sheets" / "MicrosynthUploadFormDNA.xlsx"
    exe_contents = Path(sys.executable).resolve().parent.parent if getattr(sys, "frozen", False) else Path()
    candidates = [
        Path.cwd() / rel,
        Path(__file__).resolve().parents[2] / rel,
        exe_contents / "Resources" / rel,
        _primerl_root().parent / rel,
    ]
    for p in candidates:
        if p.exists():
            return p
    return candidates[1]


def _load_openpyxl_workbook_loader():
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]

        return load_workbook
    except Exception:
        pass

    py_home_candidates = [
        _primerl_root().parent / "python",
        Path(sys.executable).resolve().parent,
        Path.cwd() / "python",
    ]
    site_pkgs: list[Path] = []
    for py_home in py_home_candidates:
        site_pkgs.extend(
            [
                py_home / "Lib" / "site-packages",
                py_home / "lib" / "site-packages",
            ]
        )
    for sp in site_pkgs:
        if sp.exists():
            sp_str = str(sp)
            if sp_str not in sys.path:
                sys.path.insert(0, sp_str)
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]

        return load_workbook
    except Exception:
        return None


def _center_dialog_on_parent(win: tk.Toplevel, parent: tk.Misc) -> None:
    """Place dialog centered over parent, keeping it on the same monitor."""
    try:
        parent.update_idletasks()
        win.update_idletasks()
        pw = max(1, int(parent.winfo_width()))
        ph = max(1, int(parent.winfo_height()))
        px = int(parent.winfo_rootx())
        py = int(parent.winfo_rooty())
        ww = max(1, int(win.winfo_reqwidth()))
        wh = max(1, int(win.winfo_reqheight()))
        x = px + (pw - ww) // 2
        y = py + (ph - wh) // 2
        win.geometry(f"+{x}+{y}")
    except Exception:
        pass


def _find_longest_orf(seq: str) -> tuple[int, int] | None:
    """Return (start, end_exclusive) for the longest ATG..STOP ORF."""
    s = (seq or "").upper()
    n = len(s)
    if n < 6:
        return None
    stops = {"TAA", "TAG", "TGA"}
    best: tuple[int, int] | None = None
    for frame in range(3):
        i = frame
        while i <= n - 3:
            codon = s[i : i + 3]
            if codon != "ATG":
                i += 3
                continue
            j = i + 3
            while j <= n - 3:
                stop = s[j : j + 3]
                if stop in stops:
                    cand = (i, j + 3)
                    if best is None or (cand[1] - cand[0]) > (best[1] - best[0]):
                        best = cand
                    break
                j += 3
            i += 3
    return best


def _run_spidey_alignment(
    *,
    spidey_path: str,
    genomic_seq: str,
    mrna_seq: str,
    print_alignment: int,
    large_intron: bool,
) -> tuple[bool, str, str]:
    spidey_exec = Path(spidey_path)
    spidey_lib_dir = spidey_exec.parent / "lib"

    with tempfile.TemporaryDirectory() as td:
        dna_tmp = Path(td) / "dna.tmp.fasta"
        mrna_tmp = Path(td) / "mrna.tmp.fasta"
        dna_tmp.write_text(f">dna\n{genomic_seq}\n", encoding="utf-8")
        mrna_tmp.write_text(f">mrna\n{mrna_seq}\n", encoding="utf-8")

        args = build_spidey_args(
            spidey_exec=spidey_path,
            dna_tmp_path=str(dna_tmp),
            mrna_tmp_path=str(mrna_tmp),
            print_alignment=print_alignment,
            large_intron=large_intron,
        )

        def _transport(cmd: list[str]) -> tuple[int, str]:
            env = dict(os.environ)
            if spidey_lib_dir.exists() and spidey_lib_dir.is_dir():
                old_ld = str(env.get("LD_LIBRARY_PATH") or "").strip()
                env["LD_LIBRARY_PATH"] = (
                    f"{spidey_lib_dir}:{old_ld}" if old_ld else str(spidey_lib_dir)
                )
            proc = subprocess_run(
                cmd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
            )
            out = (proc.stdout or "")
            if proc.stderr:
                out = out + ("\n" if out else "") + proc.stderr
            return proc.returncode, out

        res = run_spidey_with_transport(args, _transport)
        return res.ok, res.output, res.error


def _ntthal_dg_kcal_uncached(
    ntthal_path: Path,
    key: tuple[str, str, str],
) -> float | None:
    cmd = [
        str(ntthal_path),
        "-s1",
        key[0],
        "-s2",
        key[1],
        "-a",
        key[2],
        "-mv",
        "50",
        "-dv",
        "1.5",
        "-n",
        "0.2",
        "-d",
        "200",
        "-t",
        "37",
    ]
    cfg = ntthal_path.parent / "primer3_config"
    if cfg.exists():
        cmd.extend(["-path", str(cfg)])
    try:
        proc = subprocess_run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=10)
        txt = (proc.stdout or "") + "\n" + (proc.stderr or "")
        m = re.search(r"dG\s*=\s*(-?\d+(?:\.\d+)?)", txt)
        if not m:
            return None
        return float(m.group(1)) / 1000.0
    except Exception:
        return None


def _auto_threads(work_items: int, max_cap: int = 16) -> int:
    cpu = os.cpu_count() or 4
    if work_items <= 0:
        return 1
    if work_items < 200:
        return 1

    # Keep some headroom for UI and OS responsiveness.
    reserve = 2 if cpu >= 8 else 1
    usable = max(1, cpu - reserve)
    cap = min(max_cap, usable)

    if work_items < 600:
        return min(4, cap)
    if work_items < 1500:
        return min(8, cap)
    if work_items < 4000:
        return min(12, cap)
    return cap


def _available_cpu_threads(reserve: int = 1, max_cap: int = 32) -> int:
    cpu = os.cpu_count() or 1
    usable = cpu - max(0, int(reserve))
    if usable < 1:
        usable = 1
    return max(1, min(int(max_cap), usable))


def _resolve_ntthal_exe(primer3_path: str, ntthal_path: str = "") -> Path | None:
    explicit = Path((ntthal_path or "").strip())
    if explicit.exists() and explicit.is_file():
        return explicit
    p3 = Path(primer3_path or "")
    candidates: list[Path] = []
    if p3.name:
        for name in ("ntthal", "ntthal_v2.6.1_AVX2_FMA3"):
            for exec_name in candidate_exec_names(name):
                candidates.append(p3.with_name(exec_name))
    for bin_root in _tool_bin_roots():
        for name in ("ntthal", "ntthal_v2.6.1_AVX2_FMA3"):
            for exec_name in candidate_exec_names(name):
                candidates.append(bin_root / exec_name)
    for ntthal in candidates:
        if ntthal.exists() and ntthal.is_file():
            return ntthal
    return None


def _refine_rows_with_ntthal(rows: list[list[object]], primer3_path: str, ntthal_path: str = "") -> list[list[object]]:
    ntthal = _resolve_ntthal_exe(primer3_path, ntthal_path)
    if ntthal is None:
        return rows

    cache: dict[tuple[str, str, str], float | None] = {}
    keys: set[tuple[str, str, str]] = set()
    for row in rows:
        sf = str(row[0]).upper()
        sr = str(row[4]).upper()
        for a, b in ((sf, sf), (sf, sr), (sr, sr)):
            keys.add((a, b, "END1"))
            keys.add((a, b, "END2"))
            keys.add((a, b, "ANY"))

    workers = _auto_threads(len(keys), max_cap=16)
    if workers == 1 or len(keys) < 20:
        for key in keys:
            cache[key] = _ntthal_dg_kcal_uncached(ntthal, key)
    else:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = {ex.submit(_ntthal_dg_kcal_uncached, ntthal, key): key for key in keys}
            for fut in as_completed(futs):
                key = futs[fut]
                try:
                    cache[key] = fut.result()
                except Exception:
                    cache[key] = None

    out: list[list[object]] = []
    for row in rows:
        rr = list(row)
        sf = str(rr[0]).upper()
        sr = str(rr[4]).upper()

        # Extensible-like score: worst (most negative) of END1/END2 across self/cross/self.
        ext_scores: list[float] = []
        for a, b in ((sf, sf), (sf, sr), (sr, sr)):
            d1 = cache.get((a, b, "END1"))
            d2 = cache.get((a, b, "END2"))
            vals = [v for v in (d1, d2) if v is not None]
            if vals:
                ext_scores.append(min(vals))

        full_scores: list[float] = []
        for a, b in ((sf, sf), (sf, sr), (sr, sr)):
            d = cache.get((a, b, "ANY"))
            if d is not None:
                full_scores.append(d)

        if ext_scores:
            rr[10] = f"{min(ext_scores):.2f}"
        if full_scores:
            rr[13] = f"{min(full_scores):.2f}"
        out.append(rr)
    return out


def _filter_rows_with_mfeprimer(
    rows: list[list[object]],
    mfeprimer_path: str,
    dg_cutoff: float = -3.0,
    batch_size: int = 1000,
    timeout_sec: int = 120,
) -> tuple[list[list[object]], str]:
    raw = (mfeprimer_path or "").strip()
    if not raw:
        return rows, "MFEprimer path not set; skipped."
    exe = Path(raw)
    if not exe.exists() or exe.is_dir():
        return rows, "MFEprimer executable not found; skipped."

    total = len(rows)
    if total == 0:
        return rows, "MFEprimer skipped (no pairs)."
    batch = max(1, int(batch_size))

    with tempfile.TemporaryDirectory() as td:
        flagged: set[int] = set()
        timed_out = 0
        failed = 0
        processed_batches = 0
        cross_pair_hits = 0
        detected_threads = _available_cpu_threads(reserve=1, max_cap=32)

        # Parse text report:
        # Dimer N: p12_f x p34_r
        #   Score: ..., Delta G = -8.12 kcal/mol
        pat = re.compile(
            r"Dimer\s+\d+:\s+p(\d+)_([fr])\s+x\s+p(\d+)_([fr])\s+.*?Delta G\s*=\s*(-?\d+(?:\.\d+)?)",
            re.IGNORECASE | re.DOTALL,
        )

        for start in range(0, total, batch):
            end = min(total, start + batch)
            inp = Path(td) / f"pairs_{start}_{end}.fasta"
            out = Path(td) / f"mfeprimer_dimer_{start}_{end}.txt"
            lines = []
            for i in range(start, end):
                row = rows[i]
                lines.append(f">p{i}_f")
                lines.append(str(row[0]))
                lines.append(f">p{i}_r")
                lines.append(str(row[4]))
            inp.write_text("\n".join(lines) + "\n", encoding="utf-8")

            cmd = [
                str(exe),
                "dimer",
                "-i",
                str(inp),
                "-o",
                str(out),
                "-d",
                str(dg_cutoff),
                "-s",
                "5",
                "-m",
                "1",
                "-p",
                "-c",
                str(detected_threads),
            ]
            try:
                proc = subprocess_run(
                    cmd,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=timeout_sec,
                )
            except subprocess.TimeoutExpired:
                timed_out += 1
                continue
            except Exception:
                failed += 1
                continue

            processed_batches += 1
            if proc.returncode != 0:
                failed += 1
                continue

            payload_txt = ""
            if out.exists():
                payload_txt = out.read_text(encoding="utf-8", errors="replace")
            if not payload_txt.strip():
                payload_txt = proc.stdout or ""
            if not payload_txt.strip():
                continue

            for m in pat.finditer(payload_txt):
                i1 = int(m.group(1))
                i2 = int(m.group(3))
                try:
                    dg = float(m.group(5))
                except ValueError:
                    continue
                if dg > dg_cutoff:
                    continue
                # Single-pair design QC: only penalize dimers formed within the same pair.
                if i1 == i2:
                    flagged.add(i1)
                else:
                    cross_pair_hits += 1

        if not flagged:
            if timed_out or failed:
                return rows, (
                    f"MFEprimer completed with issues; no pairs filtered "
                    f"(timeouts={timed_out}, failed={failed}, threads={detected_threads})."
                )
            return rows, f"MFEprimer ran; no flagged dimer hits. (threads={detected_threads})"

        filtered = [row for i, row in enumerate(rows) if i not in flagged]
        note = f"MFEprimer filtered {len(rows) - len(filtered)} pair(s)."
        if total > batch:
            note += f" batches={((total - 1) // batch) + 1}"
        if cross_pair_hits:
            note += f" (ignored cross-pair hits={cross_pair_hits})"
        if timed_out or failed:
            note += f" (timeouts={timed_out}, failed={failed})"
        if processed_batches == 0:
            return rows, f"MFEprimer could not complete; no pairs filtered. (threads={detected_threads})"
        note += f" (threads={detected_threads})"
        return filtered, note


def _normalize_ensembl_gene_id(raw: str) -> str:
    txt = (raw or "").strip()
    if not txt:
        return ""
    # Ensembl gene IDs are usually ENS*G..., but some species (for example
    # Drosophila) expose stable IDs like FBgn.... Keep matching strict to avoid
    # accidentally treating symbols/descriptions as stable gene IDs.
    patterns = [
        r"(ENS[A-Z]*G\d+)(?:\.\d+)?",
        r"(FBGN\d+)(?:\.\d+)?",
    ]
    up = txt.upper()
    for pat in patterns:
        m = re.search(pat, up)
        if m:
            return m.group(1)
    return ""


def _filter_rows_with_mfeprimer_spec(
    rows: list[list[object]],
    mfeprimer_path: str,
    db_fasta_path: str,
    target_gene_id: str = "",
    target_gene_symbol: str = "",
    snp_check_enabled: bool = True,
    snp_bed_path: str = "",
    snp_non3p_policy: str = "soft",
    snp_3p_window: int = 5,
    max_amplicons: int = 1,
    min_amp_size: int = 80,
    max_amp_size: int = 300,
    batch_size: int = 400,
    timeout_sec: int = 240,
    spec_selection_mode: str = "strict_pass",
    spec_remove_pct: float = 10.0,
    metrics_out: dict[str, int] | None = None,
    spec_extra_args: list[str] | None = None,
) -> tuple[list[list[object]], str]:
    if isinstance(metrics_out, dict):
        # Preserve removals already counted upstream (for example pre-spec SNP filtering).
        metrics_out.setdefault("target_removed", 0)
        metrics_out.setdefault("snp_removed", 0)

    def _target_gene_ids(v: str) -> set[str]:
        ids: set[str] = set()
        for part in re.split(r"[,\s;]+", (v or "").strip()):
            gid = _normalize_ensembl_gene_id(part)
            if gid:
                ids.add(gid)
        return ids

    def _extract_hit_gene_ids(hit_desc: str) -> set[str]:
        ids: set[str] = set()
        txt = hit_desc or ""
        for m in re.finditer(r"\bgene:([^\s]+)", txt, re.IGNORECASE):
            gid = _normalize_ensembl_gene_id(m.group(1))
            if gid:
                ids.add(gid)
        return ids

    def _norm_symbol(raw: str) -> str:
        s = (raw or "").strip().lower()
        return re.sub(r"[^a-z0-9_.-]", "", s)

    def _target_symbols(v: str) -> set[str]:
        out: set[str] = set()
        for part in re.split(r"[,\s;]+", (v or "").strip()):
            n = _norm_symbol(part)
            if n:
                out.add(n)
        return out

    def _extract_hit_symbols(hit_desc: str) -> set[str]:
        out: set[str] = set()
        txt = hit_desc or ""
        # RefSeq headers commonly include official symbols in parentheses, e.g. "(meis1b)".
        for m in re.finditer(r"\(([A-Za-z0-9_.-]{2,})\)", txt):
            n = _norm_symbol(m.group(1))
            if n:
                out.add(n)
        # Also accept simple key-value style when present.
        for m in re.finditer(r"\bgene(?:_symbol)?[:=]\s*([A-Za-z0-9_.-]+)", txt, re.IGNORECASE):
            n = _norm_symbol(m.group(1))
            if n:
                out.add(n)
        return out

    def _norm_tx_id(raw: str) -> str:
        return (raw or "").strip().split(".", 1)[0]

    def _load_snp_bed(path_txt: str) -> tuple[dict[str, list[int]], int]:
        by_tx: dict[str, list[int]] = {}
        loaded = 0
        ptxt = (path_txt or "").strip()
        if not ptxt:
            return by_tx, loaded
        p = Path(ptxt)
        if not p.exists() or p.is_dir():
            return by_tx, loaded
        try:
            with p.open("r", encoding="utf-8", errors="replace") as fh:
                for raw_ln in fh:
                    ln = raw_ln.strip()
                    if not ln or ln.startswith("#"):
                        continue
                    cols = ln.split("\t")
                    if len(cols) < 3:
                        continue
                    tx = cols[0].strip()
                    if not tx:
                        continue
                    try:
                        # BED is 0-based half-open.
                        b0 = int(cols[1])
                        b1 = int(cols[2])
                    except ValueError:
                        continue
                    if b1 <= b0:
                        continue
                    key = _norm_tx_id(tx)
                    arr = by_tx.setdefault(key, [])
                    arr.extend(range(b0 + 1, b1 + 1))
                    loaded += 1
        except Exception:
            return {}, 0
        for k in list(by_tx.keys()):
            by_tx[k] = sorted(set(by_tx[k]))
        return by_tx, loaded

    def _count_snps_in_range(snp_by_tx: dict[str, list[int]], tx_id: str, s: int, e: int) -> int:
        key = _norm_tx_id(tx_id)
        arr = snp_by_tx.get(key)
        if not arr:
            return 0
        lo = min(s, e)
        hi = max(s, e)
        c = 0
        for pos in arr:
            if pos < lo:
                continue
            if pos > hi:
                break
            c += 1
        return c

    def _summarize_spec_payload(payload_txt: str) -> str:
        txt = payload_txt or ""
        parts: list[str] = []
        m_desc = re.search(r"Descriptions of \[\s*(\d+)\s*\] potential amplicons", txt, re.IGNORECASE)
        if m_desc:
            parts.append(f"potential_amplicons={m_desc.group(1)}")
        m_amp = re.search(r"Amp\s+\d+:\s+.+", txt, re.IGNORECASE)
        if m_amp:
            parts.append(re.sub(r"\s+", " ", m_amp.group(0)).strip()[:140])
        m_params = re.search(r"Kvalue:\s*(\d+).*?MisMatch:\s*(\d+)", txt, re.IGNORECASE | re.DOTALL)
        if m_params:
            parts.append(f"k={m_params.group(1)} misMatch={m_params.group(2)}")
        return "; ".join(parts)

    raw = (mfeprimer_path or "").strip()
    if not raw:
        return rows, "MFEprimer spec path not set; skipped."
    exe = Path(raw)
    if not exe.exists() or exe.is_dir():
        return rows, "MFEprimer executable not found; spec skipped."

    db_raw = (db_fasta_path or "").strip()
    if not db_raw:
        return rows, "MFEprimer spec DB not set; skipped."
    db = Path(db_raw)
    if not db.exists() or db.is_dir():
        return rows, "MFEprimer spec DB not found; skipped."

    total = len(rows)
    if total == 0:
        return rows, "MFEprimer spec skipped (no pairs)."
    batch = max(1, int(batch_size))
    snp_enabled = bool(snp_check_enabled)
    snp_map, snp_records_loaded = _load_snp_bed(snp_bed_path) if snp_enabled else ({}, 0)
    non3p_hard = str(snp_non3p_policy or "soft").strip().lower() == "hard"
    window3 = max(1, int(snp_3p_window))
    selection_mode = str(spec_selection_mode or "strict_pass").strip().lower()
    if selection_mode not in {"strict_pass", "score_top_pct"}:
        selection_mode = "strict_pass"
    remove_pct = max(0.0, min(100.0, float(spec_remove_pct)))

    # Require pre-indexed DB for safety: indexing whole genomes can be very large/slow.
    idx = Path(str(db) + ".primerqc")
    if not idx.exists():
        return rows, "MFEprimer spec DB is not indexed (.primerqc missing); skipped."

    with tempfile.TemporaryDirectory() as td:
        kept_idxs: set[int] = set()
        timed_out = 0
        failed = 0
        amp_hits: dict[int, int] = {}
        on_target_hits: dict[int, int] = {}
        off_target_hits: dict[int, int] = {}
        unknown_target_hits: dict[int, int] = {}
        snp_3p_hits: dict[int, int] = {}
        snp_non3p_hits: dict[int, int] = {}
        snp_pairs_evaluated = 0
        target_ids = _target_gene_ids(target_gene_id)
        target_syms = _target_symbols(target_gene_symbol)
        first_payload_summary = ""
        amp_pat = re.compile(
            r"Amp\s+\d+:\s+p(\d+)_fp\s+\+\s+p(\d+)_rp\s+==>\s*(.+)",
            re.IGNORECASE,
        )
        detected_threads = _available_cpu_threads(reserve=1, max_cap=32)
        batches = [(start, min(total, start + batch)) for start in range(0, total, batch)]
        max_parallel_jobs = max(1, min(3, len(batches), detected_threads))
        threads_per_job = max(1, detected_threads // max_parallel_jobs)

        def _run_spec_batch(start: int, end: int) -> tuple[int, int, str, str, str]:
            try:
                inp = Path(td) / f"spec_pairs_{start}_{end}.tsv"
                out = Path(td) / f"mfeprimer_spec_{start}_{end}.txt"
                lines = ["name\tFP\tRP"]
                for i in range(start, end):
                    row = rows[i]
                    lines.append(f"p{i}\t{row[0]}\t{row[4]}")
                inp.write_text("\n".join(lines) + "\n", encoding="utf-8")

                cmd = build_mfeprimer_spec_cmd(
                    exe=exe,
                    inp=inp,
                    db=db,
                    out=out,
                    min_amp_size=min_amp_size,
                    max_amp_size=max_amp_size,
                    threads_per_job=threads_per_job,
                    spec_extra_args=spec_extra_args,
                    snp_bed_path=(snp_bed_path if snp_enabled else ""),
                    snp_records_loaded=(snp_records_loaded if snp_enabled else 0),
                )
                proc = subprocess_run(
                    cmd,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=timeout_sec,
                )
                if proc.returncode != 0:
                    err_txt = (proc.stderr or proc.stdout or "").strip()
                    return start, end, "", "failed", err_txt[:400]
                payload_txt = ""
                if out.exists():
                    payload_txt = out.read_text(encoding="utf-8", errors="replace")
                if not payload_txt.strip():
                    payload_txt = proc.stdout or ""
                return start, end, payload_txt, "ok", ""
            except subprocess.TimeoutExpired:
                return start, end, "", "timeout", f"timed out after {timeout_sec}s"
            except Exception as exc:
                return start, end, "", "failed", str(exc)[:400]

        with ThreadPoolExecutor(max_workers=max_parallel_jobs) as ex_pool:
            future_map = {
                ex_pool.submit(_run_spec_batch, start, end): (start, end)
                for start, end in batches
            }
            first_issue: str = ""
            for fut in as_completed(future_map):
                start, end = future_map[fut]
                try:
                    _, _, payload_txt, status, err_snippet = fut.result()
                except Exception as exc:
                    failed += 1
                    if not first_issue:
                        first_issue = f"worker exception: {exc}"
                    continue

                if status == "timeout":
                    timed_out += 1
                    if not first_issue and err_snippet:
                        first_issue = err_snippet
                    continue
                if status != "ok":
                    failed += 1
                    if not first_issue and err_snippet:
                        first_issue = err_snippet
                    continue
                if not payload_txt.strip():
                    continue
                if not first_payload_summary:
                    first_payload_summary = _summarize_spec_payload(payload_txt)

                for m in amp_pat.finditer(payload_txt):
                    i1 = int(m.group(1))
                    i2 = int(m.group(2))
                    if i1 != i2:
                        continue
                    amp_hits[i1] = amp_hits.get(i1, 0) + 1
                    if target_ids:
                        hit_ids = _extract_hit_gene_ids(m.group(3))
                        if hit_ids:
                            if hit_ids.intersection(target_ids):
                                on_target_hits[i1] = on_target_hits.get(i1, 0) + 1
                            else:
                                off_target_hits[i1] = off_target_hits.get(i1, 0) + 1
                        else:
                            hit_syms = _extract_hit_symbols(m.group(3))
                            if hit_syms and target_syms:
                                if hit_syms.intersection(target_syms):
                                    on_target_hits[i1] = on_target_hits.get(i1, 0) + 1
                                else:
                                    off_target_hits[i1] = off_target_hits.get(i1, 0) + 1
                            else:
                                unknown_target_hits[i1] = unknown_target_hits.get(i1, 0) + 1

                amp_detail_pat = re.compile(
                    r"Amp\s+\d+:\s+p(\d+)_fp\s+\+\s+p(\d+)_rp\s+==>\s*([^\s:]+):\d+-\d+.*?"
                    r"F:.*?Start\s*=\s*(\d+),\s*End\s*=\s*(\d+).*?"
                    r"R:.*?Start\s*=\s*(\d+),\s*End\s*=\s*(\d+)",
                    re.IGNORECASE | re.DOTALL,
                )
                for m in amp_detail_pat.finditer(payload_txt):
                    i1 = int(m.group(1))
                    i2 = int(m.group(2))
                    if i1 != i2:
                        continue
                    tx_id = m.group(3)
                    f_start = int(m.group(4))
                    f_end = int(m.group(5))
                    r_start = int(m.group(6))
                    r_end = int(m.group(7))
                    if not snp_map:
                        continue
                    snp_pairs_evaluated += 1
                    f_all = _count_snps_in_range(snp_map, tx_id, f_start, f_end)
                    r_all = _count_snps_in_range(snp_map, tx_id, r_start, r_end)
                    f3_s = max(min(f_start, f_end), max(f_start, f_end) - window3 + 1)
                    f3_e = max(f_start, f_end)
                    r3_s = min(r_start, r_end)
                    r3_e = min(max(r_start, r_end), min(r_start, r_end) + window3 - 1)
                    f_3p = _count_snps_in_range(snp_map, tx_id, f3_s, f3_e)
                    r_3p = _count_snps_in_range(snp_map, tx_id, r3_s, r3_e)
                    if (f_3p + r_3p) > 0:
                        snp_3p_hits[i1] = snp_3p_hits.get(i1, 0) + (f_3p + r_3p)
                    non3p = max(0, (f_all + r_all) - (f_3p + r_3p))
                    if non3p > 0:
                        snp_non3p_hits[i1] = snp_non3p_hits.get(i1, 0) + non3p

                for i in range(start, end):
                    c = amp_hits.get(i, 0)
                    has_3p = snp_3p_hits.get(i, 0) > 0
                    has_non3p = snp_non3p_hits.get(i, 0) > 0
                    snp_fail = has_3p or (non3p_hard and has_non3p)
                    if target_ids:
                        on_t = on_target_hits.get(i, 0)
                        off_t = off_target_hits.get(i, 0)
                        unk_t = unknown_target_hits.get(i, 0)
                        if c >= 1 and on_t >= 1 and off_t == 0 and unk_t == 0 and not snp_fail:
                            kept_idxs.add(i)
                    elif c <= max_amplicons and c >= 1 and not snp_fail:
                        kept_idxs.add(i)

        score_mode_note = ""
        if target_ids and selection_mode == "score_top_pct":
            mandatory_remove: set[int] = set()
            scored: list[tuple[int, int, int, int, int]] = []
            for i in range(total):
                c = amp_hits.get(i, 0)
                on_t = on_target_hits.get(i, 0)
                off_t = off_target_hits.get(i, 0)
                unk_t = unknown_target_hits.get(i, 0)
                has_3p = snp_3p_hits.get(i, 0) > 0
                has_non3p = snp_non3p_hits.get(i, 0) > 0
                snp_fail = has_3p or (non3p_hard and has_non3p)
                if c < 1 or on_t < 1 or snp_fail:
                    mandatory_remove.add(i)
                    continue
                # Quantile scoring mode prioritizes explicit off-target evidence.
                score = (3 * off_t) + (1 * unk_t)
                if score > 0:
                    scored.append((score, off_t, unk_t, c, i))

            target_remove = int(math.ceil((total * remove_pct) / 100.0))
            remove_idxs: set[int] = set(mandatory_remove)
            scored.sort(key=lambda t: (t[0], t[1], t[2], t[3], -t[4]), reverse=True)
            remaining_quota = max(0, target_remove - len(remove_idxs))
            for tup in scored[:remaining_quota]:
                remove_idxs.add(tup[4])
            kept_idxs = {i for i in range(total) if i not in remove_idxs}

            score_removed = max(0, len(remove_idxs) - len(mandatory_remove))
            score_mode_note = (
                f" score_mode=top_pct({remove_pct:.1f}%), target_remove={target_remove}, "
                f"mandatory_remove={len(mandatory_remove)}, score_removed={score_removed}, "
                f"scored_candidates={len(scored)}"
            )

        if not kept_idxs and not (timed_out or failed):
            zero = sum(1 for i in range(total) if amp_hits.get(i, 0) == 0)
            if target_ids:
                off_pairs = sum(1 for i in range(total) if off_target_hits.get(i, 0) > 0)
                unknown_pairs = sum(1 for i in range(total) if unknown_target_hits.get(i, 0) > 0)
                on_only = sum(
                    1
                    for i in range(total)
                    if amp_hits.get(i, 0) >= 1
                    and on_target_hits.get(i, 0) >= 1
                    and off_target_hits.get(i, 0) == 0
                    and unknown_target_hits.get(i, 0) == 0
                )
                if selection_mode == "score_top_pct":
                    note = (
                        f"MFEprimer spec scored and removed all pairs "
                        f"(target_gene_id={target_gene_id.strip()}, zero_amp={zero}, on_only={on_only}, "
                        f"off_target={off_pairs}, unknown_target={unknown_pairs}, "
                        f"threads={detected_threads}, procs={max_parallel_jobs}, per_proc={threads_per_job})"
                    )
                else:
                    note = (
                        f"MFEprimer spec ran; no pairs passed target-gene-ID-only rule. "
                        f"(spec_mode=target_gene_id_only, target_gene_id={target_gene_id.strip()}, "
                        f"zero_amp={zero}, on_only={on_only}, off_target={off_pairs}, "
                        f"unknown_target={unknown_pairs}, threads={detected_threads}, "
                        f"procs={max_parallel_jobs}, per_proc={threads_per_job})"
                    )
                if snp_records_loaded:
                    note += (
                        f" SNP(policy_3p=hard, policy_non3p={'hard' if non3p_hard else 'soft'}, "
                        f"window3p={window3}, pairs_with_3p_snp="
                        f"{sum(1 for i in range(total) if snp_3p_hits.get(i, 0) > 0)}, "
                        f"pairs_with_non3p_snp={sum(1 for i in range(total) if snp_non3p_hits.get(i, 0) > 0)}, "
                        f"bed_records={snp_records_loaded}, amp_evaluated={snp_pairs_evaluated})"
                    )
                if zero == total and first_payload_summary:
                    note += f" Sample spec payload: {first_payload_summary}"
                if score_mode_note:
                    note += score_mode_note
                if isinstance(metrics_out, dict):
                    removed_idxs = set(range(total))
                    snp_fail_idxs = set(
                        j
                        for j in range(total)
                        if (snp_3p_hits.get(j, 0) > 0) or (non3p_hard and snp_non3p_hits.get(j, 0) > 0)
                    )
                    snp_removed = len(removed_idxs.intersection(snp_fail_idxs))
                    target_removed = max(0, len(removed_idxs) - snp_removed)
                    prev_target = int(metrics_out.get("target_removed", 0))
                    prev_snp = int(metrics_out.get("snp_removed", 0))
                    metrics_out["target_removed"] = prev_target + int(target_removed)
                    metrics_out["snp_removed"] = prev_snp + int(snp_removed)
                return [], note
            multi = sum(1 for i in range(total) if amp_hits.get(i, 0) > max_amplicons)
            one_to_max = sum(1 for i in range(total) if 1 <= amp_hits.get(i, 0) <= max_amplicons)
            note = (
                f"MFEprimer spec ran; no pairs passed amplicon-specificity rule. "
                f"(spec_mode=count_only, zero_amp={zero}, one_to_max={one_to_max}, "
                f"multi_amp={multi}, max_amp={max_amplicons}, threads={detected_threads}, "
                f"procs={max_parallel_jobs}, per_proc={threads_per_job})"
            )
            if snp_records_loaded:
                note += (
                    f" SNP(policy_3p=hard, policy_non3p={'hard' if non3p_hard else 'soft'}, "
                    f"window3p={window3}, pairs_with_3p_snp="
                    f"{sum(1 for i in range(total) if snp_3p_hits.get(i, 0) > 0)}, "
                    f"pairs_with_non3p_snp={sum(1 for i in range(total) if snp_non3p_hits.get(i, 0) > 0)}, "
                    f"bed_records={snp_records_loaded}, amp_evaluated={snp_pairs_evaluated})"
                )
            if zero == total and first_payload_summary:
                note += f" Sample spec payload: {first_payload_summary}"
            if score_mode_note:
                note += score_mode_note
            return [], note

        if not kept_idxs and (timed_out or failed):
            note = (
                f"MFEprimer spec completed with issues; no filtering applied "
                f"(timeouts={timed_out}, failed={failed}, threads={detected_threads}, "
                f"procs={max_parallel_jobs}, per_proc={threads_per_job})."
            )
            if first_issue:
                note += f" First error: {first_issue}"
            return rows, note

        filtered = [row for i, row in enumerate(rows) if i in kept_idxs]
        removed = len(rows) - len(filtered)
        removed_idxs = set(i for i in range(total) if i not in kept_idxs)
        snp_fail_idxs = set(
            i
            for i in range(total)
            if (snp_3p_hits.get(i, 0) > 0) or (non3p_hard and snp_non3p_hits.get(i, 0) > 0)
        )
        snp_removed = len(removed_idxs.intersection(snp_fail_idxs))
        target_removed = max(0, len(removed_idxs) - snp_removed)
        if isinstance(metrics_out, dict):
            prev_target = int(metrics_out.get("target_removed", 0))
            prev_snp = int(metrics_out.get("snp_removed", 0))
            metrics_out["target_removed"] = prev_target + int(target_removed)
            metrics_out["snp_removed"] = prev_snp + int(snp_removed)
        zero = sum(1 for i in range(total) if amp_hits.get(i, 0) == 0)
        if target_ids:
            off_pairs = sum(1 for i in range(total) if off_target_hits.get(i, 0) > 0)
            unknown_pairs = sum(1 for i in range(total) if unknown_target_hits.get(i, 0) > 0)
            on_only = sum(
                1
                for i in range(total)
                if amp_hits.get(i, 0) >= 1
                and on_target_hits.get(i, 0) >= 1
                and off_target_hits.get(i, 0) == 0
                and unknown_target_hits.get(i, 0) == 0
            )
            if selection_mode == "score_top_pct":
                note = (
                    f"MFEprimer spec scored and removed {removed} pair(s) "
                    f"(target_gene_id={target_gene_id.strip()}, zero_amp={zero}, on_only={on_only}, "
                    f"off_target={off_pairs}, unknown_target={unknown_pairs})."
                )
            else:
                note = (
                    f"MFEprimer spec filtered {removed} pair(s); kept target-gene-ID-only amplicons "
                    f"(target_gene_id={target_gene_id.strip()}, zero_amp={zero}, on_only={on_only}, "
                    f"off_target={off_pairs}, unknown_target={unknown_pairs})."
                )
        else:
            multi = sum(1 for i in range(total) if amp_hits.get(i, 0) > max_amplicons)
            note = (
                f"MFEprimer spec filtered {removed} pair(s); kept exact/specific amplicons "
                f"(zero_amp={zero}, multi_amp={multi}, max_amp={max_amplicons})."
            )
        if snp_records_loaded:
            note += (
                f" SNP(policy_3p=hard, policy_non3p={'hard' if non3p_hard else 'soft'}, "
                f"window3p={window3}, pairs_with_3p_snp="
                f"{sum(1 for i in range(total) if snp_3p_hits.get(i, 0) > 0)}, "
                f"pairs_with_non3p_snp={sum(1 for i in range(total) if snp_non3p_hits.get(i, 0) > 0)}, "
                f"bed_records={snp_records_loaded}, amp_evaluated={snp_pairs_evaluated})"
            )
        if total > batch:
            note += f" batches={((total - 1) // batch) + 1}"
        if score_mode_note:
            note += score_mode_note
        if timed_out or failed:
            note += f" (timeouts={timed_out}, failed={failed})"
            if first_issue:
                note += f" First error: {first_issue}"
        note += f" (threads={detected_threads}, procs={max_parallel_jobs}, per_proc={threads_per_job})"
        return filtered, note


def _sort_rows_by_ext_dimer_desc(rows: list[list[object]]) -> list[list[object]]:
    def _ext(row: list[object]) -> float:
        try:
            return float(row[10])
        except (TypeError, ValueError, IndexError):
            return -9999.0

    return sorted(rows, key=_ext, reverse=True)


def _selected_exon_window(
    *,
    bounds: list[int],
    template_len: int,
    ex5_txt: str,
    ex3_txt: str,
) -> tuple[int, int, int, int, int]:
    exons = max(1, len(bounds) + 1)
    ex5 = int(ex5_txt) if ex5_txt.strip() else 1
    ex3 = int(ex3_txt) if ex3_txt.strip() else exons
    ex5 = max(1, min(exons, ex5))
    ex3 = max(1, min(exons, ex3))
    if ex5 > ex3:
        ex5, ex3 = ex3, ex5

    lo = 0 if ex5 <= 1 else int(bounds[ex5 - 2])
    hi = int(template_len) if ex3 >= exons else int(bounds[ex3 - 1])
    lo = max(0, min(int(template_len), lo))
    hi = max(lo, min(int(template_len), hi))
    return lo, hi, ex5, ex3, exons


def _filter_rows_to_cdna_window(
    rows: list[list[object]],
    *,
    lo: int,
    hi: int,
) -> tuple[list[list[object]], int]:
    kept: list[list[object]] = []
    removed = 0
    for row in rows:
        try:
            f_start = int(row[1])
            f_end = f_start + int(row[2])
            r_end = int(row[8])
            r_start = r_end - int(row[6])
        except Exception:
            kept.append(row)
            continue
        if f_start >= lo and f_end <= hi and r_start >= lo and r_end <= hi:
            kept.append(row)
        else:
            removed += 1
    return kept, removed


def _filter_rows_by_ntthal_ext_cutoff(
    rows: list[list[object]],
    ext_dg_cutoff: float,
) -> tuple[list[list[object]], str]:
    flagged = 0
    kept: list[list[object]] = []
    for row in rows:
        try:
            ext = float(row[10])
        except (TypeError, ValueError, IndexError):
            # Keep malformed rows; they can still be reviewed.
            kept.append(row)
            continue
        if ext <= ext_dg_cutoff:
            flagged += 1
            continue
        kept.append(row)
    return kept, f"ntthal cutoff filtered {flagged} pair(s) at <= {ext_dg_cutoff:.2f}."


def _filter_rows_by_target_transcript_snps(
    rows: list[list[object]],
    *,
    snp_cdna_positions: list[int],
    window3: int,
    non3p_hard: bool,
) -> tuple[list[list[object]], int, int, int]:
    snps = sorted(set(int(p) for p in snp_cdna_positions if int(p) > 0))
    if not snps:
        return list(rows), 0, 0, 0

    def _count(lo: int, hi: int) -> int:
        s = min(lo, hi)
        e = max(lo, hi)
        i0 = bisect_left(snps, s)
        i1 = bisect_right(snps, e)
        return max(0, i1 - i0)

    w3 = max(1, int(window3))
    kept: list[list[object]] = []
    removed = 0
    pairs_with_3p = 0
    pairs_with_non3p = 0
    for row in rows:
        try:
            f_start = int(row[1])
            f_len = int(row[2])
            r_end = int(row[8])
            r_len = int(row[6])
        except Exception:
            kept.append(row)
            continue

        f_end = f_start + f_len
        r_start = r_end - r_len

        f3_s = max(min(f_start, f_end), max(f_start, f_end) - w3 + 1)
        f3_e = max(f_start, f_end)
        r3_s = min(r_start, r_end)
        r3_e = min(max(r_start, r_end), min(r_start, r_end) + w3 - 1)

        f_3p = _count(f3_s, f3_e)
        r_3p = _count(r3_s, r3_e)
        hits_3p = f_3p + r_3p
        hits_all = _count(f_start, f_end) + _count(r_start, r_end)
        hits_non3p = max(0, hits_all - hits_3p)

        snp_fail = (hits_3p > 0) or (non3p_hard and hits_non3p > 0)
        if snp_fail:
            removed += 1
            if hits_3p > 0:
                pairs_with_3p += 1
            if hits_non3p > 0:
                pairs_with_non3p += 1
            continue
        kept.append(row)

    return kept, removed, pairs_with_3p, pairs_with_non3p


def _http_transport(url: str) -> tuple[int, str]:
    def _curl_json(max_time_sec: int) -> tuple[int, str]:
        curl = shutil.which("curl")
        if not curl:
            return 127, "curl executable not found"
        cmd = [
            curl,
            "-4",
            "--http1.1",
            "-sS",
            "-L",
            "--compressed",
            "--retry",
            "4",
            "--retry-delay",
            "1",
            "--retry-connrefused",
            "--connect-timeout",
            "15",
            "--max-time",
            str(max(20, int(max_time_sec))),
            "-H",
            "Accept: application/json",
            "-H",
            "Content-Type: application/json",
            url,
        ]
        proc = subprocess_run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        payload = (proc.stdout or "").strip()
        err = (proc.stderr or "").strip()
        if proc.returncode != 0:
            if payload and err:
                return proc.returncode, f"{err} {payload}".strip()
            return proc.returncode, (err or payload or f"curl exit {proc.returncode}")
        return 0, payload

    curl_err = ""
    if shutil.which("curl"):
        code, payload = _curl_json(max_time_sec=120)
        if code == 0 and payload:
            return 0, payload
        curl_err = payload

    req = Request(url, headers={"Accept": "application/json", "Content-Type": "application/json"})
    last_err = ""
    for timeout_sec in (30, 60, 90):
        try:
            with urlopen(req, timeout=timeout_sec) as r:
                txt = r.read().decode("utf-8", errors="replace")
            return 0, txt
        except HTTPError as e:
            body = ""
            try:
                body = e.read().decode("utf-8", errors="replace")
            except Exception:
                body = ""
            return 1, f"{e.code} {e.reason} {body}".strip()
        except URLError as e:
            last_err = str(e)
        except TimeoutError as e:
            last_err = str(e)
        except Exception as e:
            last_err = str(e)
        time.sleep(0.2)
    if last_err and curl_err:
        return 1, f"{last_err} (curl fallback: {curl_err})"
    return 1, (last_err or curl_err)


def _http_get_text(url: str, timeout_sec: int = 60) -> str:
    req = Request(
        url,
        headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "User-Agent": "primerl/1.0.0",
        },
    )
    last_err: Exception | None = None
    for t_sec in (max(10, int(timeout_sec)), max(20, int(timeout_sec) * 2)):
        try:
            with urlopen(req, timeout=t_sec) as r:
                return r.read().decode("utf-8", errors="replace")
        except Exception as exc:
            last_err = exc
            time.sleep(0.2)
    curl = shutil.which("curl")
    if curl:
        cmd = [
            curl,
            "-4",
            "--http1.1",
            "-sS",
            "-L",
            "--compressed",
            "--retry",
            "3",
            "--retry-delay",
            "1",
            "--retry-connrefused",
            "--connect-timeout",
            "15",
            "--max-time",
            str(max(30, int(timeout_sec) * 2)),
            url,
        ]
        proc = subprocess_run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if proc.returncode == 0 and (proc.stdout or "").strip():
            return proc.stdout
        if proc.stderr:
            last_err = RuntimeError(proc.stderr.strip())
    if last_err is not None:
        raise last_err
    raise RuntimeError("unexpected download failure")


def _resolve_ensembl_cdna_download(species_slug: str) -> tuple[str, str]:
    slug = (species_slug or "").strip().lower()
    if not slug:
        raise ValueError("Species slug is required.")

    base_url = f"https://ftp.ensembl.org/pub/current_fasta/{slug}/cdna/"
    listing = _http_get_text(base_url, timeout_sec=60)

    matches = re.findall(r'href=["\']([^"\']+\.cdna\.all\.fa\.gz)["\']', listing, flags=re.IGNORECASE)
    if not matches:
        # Fallback for plain index listings where links may not be quoted as expected.
        matches = re.findall(r'([A-Za-z0-9_.-]+\.cdna\.all\.fa\.gz)', listing, flags=re.IGNORECASE)
    candidates = sorted(set(Path(m).name for m in matches))
    if not candidates:
        raise RuntimeError(f"No Ensembl cDNA .all FASTA found for '{slug}'.")

    name = candidates[-1]
    return name, base_url + name


def _download_http_file(url: str, out_path: Path, timeout_sec: int = 120) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    req = Request(url, headers={"User-Agent": "primerl/1.0.0"})
    last_err: Exception | None = None
    curl = shutil.which("curl")
    if curl:
        cmd = [
            curl,
            "-4",
            "--http1.1",
            "-f",
            "-L",
            "--retry",
            "4",
            "--retry-delay",
            "1",
            "--retry-connrefused",
            "--connect-timeout",
            "20",
            "--max-time",
            str(max(60, int(timeout_sec) * 2)),
            "--output",
            str(out_path),
            url,
        ]
        proc = subprocess_run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if proc.returncode == 0 and out_path.exists() and out_path.stat().st_size > 0:
            return
        if proc.stderr:
            last_err = RuntimeError(proc.stderr.strip())

    for t_sec in (max(20, int(timeout_sec)), max(30, int(timeout_sec) * 2)):
        try:
            with urlopen(req, timeout=t_sec) as r, out_path.open("wb") as fh:
                shutil.copyfileobj(r, fh, length=1024 * 1024)
            return
        except Exception as exc:
            last_err = exc
            time.sleep(0.3)
    if last_err is not None:
        raise last_err
    raise RuntimeError("unexpected download failure")


def _gunzip_file(gz_path: Path, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(gz_path, "rb") as src, out_path.open("wb") as dst:
        shutil.copyfileobj(src, dst, length=1024 * 1024)


def _fetch_ensembl_variation_bed_temp(
    *,
    gene_id: str,
    transport,
) -> tuple[str, int, str]:
    """Fetch Ensembl variations for a gene ID and write one temporary BED file."""
    gid = _normalize_ensembl_gene_id(gene_id)
    if not gid:
        return "", 0, "invalid gene id"

    out_path = Path(tempfile.gettempdir()) / "primerl_current_snp.bed"
    url = f"https://rest.ensembl.org/overlap/id/{gid}?feature=variation"
    data = fetch_json_with_transport(url, transport)
    if isinstance(data, EnsemblError) or not isinstance(data, list):
        try:
            out_path.unlink(missing_ok=True)
        except Exception:
            pass
        if isinstance(data, EnsemblError):
            return "", 0, data.message
        return "", 0, "unexpected Ensembl variation payload"

    lines: list[str] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        seq_region = item.get("seq_region_name")
        start = item.get("start")
        end = item.get("end")
        if seq_region is None or start is None or end is None:
            continue
        try:
            s = int(start)
            e = int(end)
        except (TypeError, ValueError):
            continue
        if e < s:
            s, e = e, s
        bed_start = max(0, s - 1)
        name = str(item.get("id") or "var")
        lines.append(f"{seq_region}\t{bed_start}\t{e}\t{name}\t.\t.")

    uniq_lines = sorted(set(lines))
    if not uniq_lines:
        try:
            out_path.unlink(missing_ok=True)
        except Exception:
            pass
        return "", 0, "no variants returned by Ensembl"

    out_path.write_text("\n".join(uniq_lines) + "\n", encoding="ascii", errors="ignore")
    return str(out_path), len(uniq_lines), "ok"


def _cdna_bounds_and_snp_positions_from_transcript(
    *,
    lookup_payload: dict[str, object],
    transcript_id: str,
    snp_bed_path: str,
) -> tuple[list[int], list[int]]:
    tx_list = lookup_payload.get("Transcript")
    if not isinstance(tx_list, list):
        return [], []
    tx_obj: dict[str, object] | None = None
    for tx in tx_list:
        if isinstance(tx, dict) and str(tx.get("id") or "") == transcript_id:
            tx_obj = tx
            break
    if tx_obj is None:
        return [], []

    exon_list = tx_obj.get("Exon")
    if not isinstance(exon_list, list):
        return [], []
    exons: list[tuple[int, int]] = []
    for ex in exon_list:
        if not isinstance(ex, dict):
            continue
        try:
            s = int(ex.get("start"))  # type: ignore[arg-type]
            e = int(ex.get("end"))  # type: ignore[arg-type]
        except Exception:
            continue
        if e < s:
            s, e = e, s
        exons.append((s, e))
    if not exons:
        return [], []

    try:
        strand = int(tx_obj.get("strand") or lookup_payload.get("strand") or 1)
    except Exception:
        strand = 1
    exons_sorted = sorted(exons, key=lambda t: t[0])
    if strand < 0:
        exons_sorted = list(reversed(exons_sorted))

    bounds: list[int] = []
    cum = 0
    for i, (s, e) in enumerate(exons_sorted):
        cum += (e - s + 1)
        if i < len(exons_sorted) - 1:
            bounds.append(cum)

    raw_seq_region = str(tx_obj.get("seq_region_name") or lookup_payload.get("seq_region_name") or "").strip()
    seq_region_norm = raw_seq_region.lower().removeprefix("chr")
    bed_path = (snp_bed_path or "").strip()
    if not bed_path:
        return bounds, []
    p = Path(bed_path)
    if not p.exists() or p.is_dir():
        return bounds, []

    def _map_genomic_to_cdna(pos: int) -> int | None:
        offset = 0
        for s, e in exons_sorted:
            if s <= pos <= e:
                if strand < 0:
                    return offset + (e - pos) + 1
                return offset + (pos - s) + 1
            offset += (e - s + 1)
        return None

    snp_cdna: set[int] = set()
    try:
        with p.open("r", encoding="utf-8", errors="replace") as fh:
            for raw_ln in fh:
                ln = raw_ln.strip()
                if not ln or ln.startswith("#"):
                    continue
                cols = ln.split("\t")
                if len(cols) < 3:
                    continue
                chrom = cols[0].strip().lower().removeprefix("chr")
                if seq_region_norm and chrom != seq_region_norm:
                    continue
                try:
                    b0 = int(cols[1])
                    b1 = int(cols[2])
                except ValueError:
                    continue
                if b1 <= b0:
                    continue
                for pos in range(b0 + 1, b1 + 1):
                    cpos = _map_genomic_to_cdna(pos)
                    if cpos is not None:
                        snp_cdna.add(cpos)
    except Exception:
        return bounds, []

    return bounds, sorted(snp_cdna)



def _choose_transcript_dialog(root: tk.Tk, choices: list[object], preselected_label: str) -> str | None:
    labels = [str(getattr(c, "display_label", "")) for c in choices]
    if not labels:
        return None
    win = tk.Toplevel(root)
    win.title("Select transcript")
    win.transient(root)
    ttk.Label(win, text=f"Gene has {len(labels)} transcripts (longest preselected)").pack(anchor="w", padx=10, pady=(10, 4))
    selected = tk.StringVar(value=preselected_label if preselected_label in labels else labels[0])
    box = ttk.Combobox(win, textvariable=selected, values=labels, width=70, state="readonly")
    box.pack(fill="x", padx=10, pady=4)

    result: dict[str, str | None] = {"value": None}

    def _ok() -> None:
        result["value"] = selected.get().strip() or None
        win.destroy()

    def _cancel() -> None:
        result["value"] = None
        win.destroy()

    btns = ttk.Frame(win)
    btns.pack(fill="x", padx=10, pady=(6, 10))
    ttk.Button(btns, text="OK", command=_ok).pack(side="left")
    ttk.Button(btns, text="Cancel", command=_cancel).pack(side="left", padx=6)
    _center_dialog_on_parent(win, root)
    win.grab_set()
    win.focus_force()
    win.wait_window()
    return result["value"]


def _pairs_to_values(row: list[object], checked: bool = False, spec_pass: bool = False) -> tuple[str, ...]:
    return (
        "[x]" if checked else "[ ]",
        "PASS" if spec_pass else "",
        str(row[0]),
        str(row[1]),
        str(row[2]),
        str(row[3]),
        str(row[4]),
        str(row[5]),
        str(row[6]),
        str(row[7]),
        str(row[9]),
        str(row[10]),
        str(row[13]),
    )


def _row_spec_key(row: list[object]) -> str:
    return "|".join(str(v) for v in row)


def _attach_tooltip(widget: tk.Widget, text: str) -> None:
    if not text or _BootstrapToolTip is None:
        return
    try:
        _BootstrapToolTip(widget, text=text)
    except Exception:
        return


def launch_gui() -> int:
    _enable_windows_dpi_awareness()
    runtime_settings_path = Path(_primerl_data_path("runtime", "gui_settings.json"))
    default_max_genomic_view_bases = 10000

    def _sanitize_max_genomic_view_bases(raw: object) -> int:
        try:
            n = int(str(raw).strip())
        except Exception:
            n = default_max_genomic_view_bases
        return max(1000, min(2_000_000, n))

    def _load_runtime_settings() -> dict[str, object]:
        try:
            if runtime_settings_path.exists():
                data = json.loads(runtime_settings_path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
        except Exception:
            pass
        return {}

    runtime_settings = _load_runtime_settings()
    defaults = {
        "primer3": _existing_default(
            _tool_profile_exec_default("apple_silicon", "primer3_core"),
            _tool_bin_exec_default("primer3_core"),
        ),
        "spidey": _existing_default(
            _tool_bin_exec_default("spidey"),
            _existing_default(
                str(Path(_primerl_path("tools", "bin")) / "Spidey.exe"),
                str(Path(_primerl_path("tools", "bin")) / "spidey.exe"),
            ),
            _path_exec_default("spidey"),
        ),
        "mrna": _existing_default(
            _primerl_data_path("databases", "ensembl", "mRNA.fasta"),
            _primerl_resource_path("databases", "ensembl", "mRNA.fasta"),
        ),
        "genomic": _existing_default(
            _primerl_data_path("databases", "ensembl", "genomic.fasta"),
            _primerl_resource_path("databases", "ensembl", "genomic.fasta"),
        ),
        "mfeprimer": _existing_default(
            _tool_profile_exec_default("apple_silicon", "mfeprimer"),
            _tool_bin_exec_default("mfeprimer"),
            _path_exec_default("mfeprimer"),
        ),
        "mfeprimer_transcriptome_fasta": _existing_default(
            _primerl_data_path("databases", "ensembl", "Danio_rerio.GRCz11.cdna.all.fa"),
            _primerl_data_path("databases", "refseq", "Danio_rerio.RefSeq.rna.all.fa"),
            _primerl_resource_path("databases", "ensembl", "Danio_rerio.GRCz11.cdna.all.fa"),
            _primerl_resource_path("databases", "refseq", "Danio_rerio.RefSeq.rna.all.fa"),
        ),
        "ntthal": _existing_default(
            _tool_profile_exec_default("apple_silicon", "ntthal"),
            _tool_bin_exec_default("ntthal"),
            _tool_bin_exec_default("ntthal_v2.6.1_AVX2_FMA3"),
        ),
        "oligotm": _existing_default(
            _tool_profile_exec_default("apple_silicon", "oligotm"),
            _tool_bin_exec_default("oligotm"),
            _tool_bin_exec_default("oligotm_v2.6.1_AVX2_FMA3"),
        ),
    }

    root = tb.Window(themename=LIGHT_THEME)
    root.title("PrimeRL v1.2")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    try:
        screen_h = int(root.winfo_screenheight())
    except Exception:
        screen_h = 1080
    compact_layout = screen_h <= 1120
    seq_text_rows = 7 if compact_layout else 10
    map_canvas_height = 84 if compact_layout else 120
    result_tree_rows = 8 if compact_layout else 12

    root_container = ttk.Frame(root, padding=PAD_L)
    root_container.grid(row=0, column=0, sticky=NSEW)
    root_container.columnconfigure(0, weight=1)
    root_container.rowconfigure(4, weight=1)

    style = ttk.Style(root)
    base_ui_font = tkfont.nametofont("TkDefaultFont")
    font_family = str(base_ui_font.cget("family"))
    button_font = (font_family, 10, "bold")
    section_header_font = tkfont.Font(root=root, family=font_family, size=10, weight="normal")
    def _apply_ui_styles() -> None:
        style.configure("SectionHeader.TLabelframe", padding=PAD_M)
        style.configure("TButton", font=button_font)
        style.configure("Primary.TButton", font=button_font)
        style.configure("TLabelframe.Label", font=section_header_font)
        style.configure("SectionHeader.TLabelframe.Label", font=section_header_font)

    _apply_ui_styles()
    title_font = (font_family, 20, "bold")
    subtitle_font = section_header_font
    dark_mode_var = tk.IntVar(value=0)

    def _apply_theme() -> None:
        target_theme = DARK_THEME if bool(dark_mode_var.get()) else LIGHT_THEME
        try:
            root.style.theme_use(target_theme)
        except Exception:
            style.theme_use(target_theme)
        _apply_ui_styles()
        try:
            if bool(dark_mode_var.get()):
                map_canvas.configure(bg="#1f1f1f", highlightbackground="#4a4a4a")
            else:
                map_canvas.configure(bg="#f5f5f5", highlightbackground="#c0c0c0")
        except Exception:
            pass

    def _build_header(parent: ttk.Frame, mode_var: tk.IntVar, on_theme_toggle: object) -> None:
        header = ttk.Frame(parent, padding=(0, 0, 0, PAD_M))
        header.grid(row=0, column=0, sticky=EW)
        header.columnconfigure(0, weight=1)
        ttk.Label(header, text="PrimeRL", font=title_font).grid(row=0, column=0, sticky=W)
        ttk.Label(header, text="qPCR primer design workspace", font=subtitle_font).grid(
            row=1,
            column=0,
            sticky=W,
            pady=(2, 0),
        )
        ttk.Checkbutton(
            header,
            text="Dark mode",
            variable=mode_var,
            command=on_theme_toggle,
            bootstyle="round-toggle",
        ).grid(row=0, column=1, rowspan=2, sticky=E)

    _build_header(root_container, dark_mode_var, _apply_theme)

    def _build_top_boxes(parent: tk.Widget) -> tuple[ttk.Frame, ttk.LabelFrame, ttk.LabelFrame, ttk.LabelFrame]:
        top_frame = ttk.Frame(parent)
        top_frame.grid(row=1, column=0, sticky=EW, pady=(0, PAD_M))
        top_frame.grid_rowconfigure(0, weight=1)
        top_frame.grid_columnconfigure(0, weight=3, uniform="top")
        top_frame.grid_columnconfigure(1, weight=6, uniform="top")
        top_frame.grid_columnconfigure(2, weight=2, uniform="top")

        primer_frame = ttk.LabelFrame(
            top_frame,
            text="Primer Settings",
            style="SectionHeader.TLabelframe",
            padding=PAD_M,
            labelanchor="nw",
        )
        primer_frame.grid(row=0, column=0, sticky=NSEW, padx=(0, PAD_M), pady=0)
        filter_frame = ttk.LabelFrame(
            top_frame,
            text="Filter Options",
            style="SectionHeader.TLabelframe",
            padding=PAD_M,
            labelanchor="nw",
        )
        filter_frame.grid(row=0, column=1, sticky=NSEW, padx=(0, PAD_M), pady=0)
        controls_frame = ttk.LabelFrame(
            top_frame,
            text="Controls",
            style="SectionHeader.TLabelframe",
            padding=PAD_M,
            labelanchor="nw",
        )
        controls_frame.grid(row=0, column=2, sticky=NSEW, padx=0, pady=0)
        return top_frame, primer_frame, filter_frame, controls_frame

    top_boxes_frame, lf_primer, lf_opt, lf_ctrl = _build_top_boxes(root_container)
    lf_opt.columnconfigure(0, weight=1)
    lf_opt.columnconfigure(1, weight=1)
    spacer_row = 100
    action_row = spacer_row + 1
    lf_primer.rowconfigure(spacer_row, weight=1)
    lf_opt.rowconfigure(spacer_row, weight=1)

    primer_rules_group = ttk.LabelFrame(lf_opt, text="Primer rules")
    primer_rules_group.grid(row=0, column=0, sticky="nsew", padx=(PAD_S, PAD_S // 2), pady=(0, PAD_S // 2))
    specificity_group = ttk.LabelFrame(lf_opt, text="Specificity options")
    specificity_group.grid(row=0, column=1, sticky="nsew", padx=(PAD_S // 2, PAD_S), pady=(0, PAD_S // 2))
    primer_rules_group.columnconfigure(0, weight=1)
    specificity_group.columnconfigure(0, weight=1)
    for row_idx in range(3):
        primer_rules_group.rowconfigure(row_idx, weight=1)
    specificity_group.rowconfigure(0, weight=1)

    min_tm_var = tk.StringVar(value="58")
    max_tm_var = tk.StringVar(value="62")
    max_diff_var = tk.StringVar(value="2")

    min_len_var = tk.StringVar(value="20")
    max_len_var = tk.StringVar(value="24")

    min_amp_var = tk.StringVar(value="100")
    max_amp_var = tk.StringVar(value="300")
    ie_limit_var = tk.IntVar(value=0)
    ie_5p_var = tk.StringVar(value="")
    ie_3p_var = tk.StringVar(value="")
    def _build_primer_settings_group(parent: ttk.LabelFrame) -> None:
        tm_row = ttk.Frame(parent)
        tm_row.grid(row=0, column=0, columnspan=6, sticky="w", padx=PAD_S, pady=(PAD_XXS, 1))
        ttk.Label(tm_row, text="Tm").pack(side="left")
        ttk.Entry(tm_row, textvariable=min_tm_var, width=6).pack(side="left", padx=(PAD_XS, PAD_XXS))
        ttk.Label(tm_row, text="-").pack(side="left")
        ttk.Entry(tm_row, textvariable=max_tm_var, width=6).pack(side="left", padx=(PAD_XXS, PAD_XS))
        ttk.Label(tm_row, text="max \u0394Tm").pack(side="left", padx=(PAD_XXS, PAD_XS))
        ttk.Entry(tm_row, textvariable=max_diff_var, width=6).pack(side="left")

        len_row = ttk.Frame(parent)
        len_row.grid(row=1, column=0, columnspan=6, sticky="w", padx=PAD_S, pady=(1, 1))
        ttk.Label(len_row, text="Length").pack(side="left")
        ttk.Entry(len_row, textvariable=min_len_var, width=6).pack(side="left", padx=(PAD_XS, PAD_XXS))
        ttk.Label(len_row, text="-").pack(side="left")
        ttk.Entry(len_row, textvariable=max_len_var, width=6).pack(side="left", padx=(PAD_XXS, PAD_XS))
        ttk.Label(len_row, text="bases").pack(side="left", padx=(PAD_XXS, 0))

        amp_row = ttk.Frame(parent)
        amp_row.grid(row=2, column=0, columnspan=6, sticky="w", padx=PAD_S, pady=(1, 1))
        ttk.Label(amp_row, text="Amplicon").pack(side="left")
        ttk.Entry(amp_row, textvariable=min_amp_var, width=7).pack(side="left", padx=(PAD_XS, PAD_XXS))
        ttk.Label(amp_row, text="-").pack(side="left")
        ttk.Entry(amp_row, textvariable=max_amp_var, width=7).pack(side="left", padx=(PAD_XXS, PAD_XS))
        ttk.Label(amp_row, text="bases").pack(side="left", padx=(PAD_XXS, 0))

        exon_row = ttk.Frame(parent)
        exon_row.grid(row=3, column=0, columnspan=6, sticky="w", padx=PAD_S, pady=(1, PAD_XXS))
        ttk.Checkbutton(exon_row, text="Limit primers to exon(s)", variable=ie_limit_var).pack(side="left")
        ttk.Entry(exon_row, textvariable=ie_5p_var, width=5).pack(side="left", padx=(PAD_S, PAD_XXS))
        ttk.Label(exon_row, text="-").pack(side="left")
        ttk.Entry(exon_row, textvariable=ie_3p_var, width=5).pack(side="left", padx=(PAD_XXS, 0))
        ttk.Button(
            parent,
            text="Quick Launch Guide",
            command=lambda: _open_local_doc_page("How-To", "HOW_TO.html"),
        ).grid(row=action_row, column=0, columnspan=6, sticky="ew", padx=PAD_S, pady=(PAD_S, 0))

    _build_primer_settings_group(lf_primer)

    exclude_gc_var = tk.IntVar(value=1)
    gc_clamp_var = tk.IntVar(value=1)
    exclude_rr_var = tk.IntVar(value=1)
    ie_span_var = tk.IntVar(value=1)
    ie_overlap_var = tk.IntVar(value=1)
    exclude_ie_var = tk.StringVar(value="7")
    repeat_var = tk.StringVar(value="4")
    run_var = tk.StringVar(value="4")

    def _build_primer_rules_group(parent: ttk.LabelFrame) -> None:
        quality_row = ttk.Frame(parent)
        quality_row.grid(row=0, column=0, sticky="ew", padx=PAD_S, pady=(PAD_S, PAD_S // 2))
        ttk.Checkbutton(quality_row, text="GC Content 40-60%", variable=exclude_gc_var).pack(side="left")
        ttk.Checkbutton(quality_row, text="GC clamp", variable=gc_clamp_var).pack(side="left", padx=(10, 0))

        rr_row = ttk.Frame(parent)
        rr_row.grid(row=1, column=0, sticky="ew", padx=PAD_S, pady=(PAD_S // 2, PAD_S // 2))
        ttk.Checkbutton(rr_row, text="Exclude repeats / runs", variable=exclude_rr_var).pack(side="left")
        ttk.Label(rr_row, text="run").pack(side="left", padx=(8, 2))
        ttk.Entry(rr_row, textvariable=run_var, width=4).pack(side="left")
        ttk.Label(rr_row, text="repeat").pack(side="left", padx=(8, 2))
        ttk.Entry(rr_row, textvariable=repeat_var, width=4).pack(side="left")

        intron_row = ttk.Frame(parent)
        intron_row.grid(row=2, column=0, sticky="ew", padx=PAD_S, pady=(PAD_S // 2, PAD_S))
        ttk.Checkbutton(intron_row, text="Span intron/exon boundary", variable=ie_span_var).pack(side="left")
        ttk.Checkbutton(intron_row, text="Overlap by", variable=ie_overlap_var).pack(side="left", padx=(10, 0))
        ttk.Entry(intron_row, textvariable=exclude_ie_var, width=5).pack(side="left", padx=(4, 2))
        ttk.Label(intron_row, text="bases").pack(side="left")

    _build_primer_rules_group(primer_rules_group)

    def _build_sequence_boxes(parent: ttk.Frame) -> tuple[tk.Text, tk.Text]:
        seqs = ttk.Frame(parent)
        seqs.grid(row=2, column=0, sticky=EW, pady=(0, PAD_M))
        lf_gen = ttk.LabelFrame(seqs, text="Genomic Sequence", style="SectionHeader.TLabelframe", padding=PAD_S)
        lf_gen.grid(row=0, column=0, sticky=NSEW, padx=(0, PAD_S))
        lf_mrna = ttk.LabelFrame(seqs, text="cDNA Sequence", style="SectionHeader.TLabelframe", padding=PAD_S)
        lf_mrna.grid(row=0, column=1, sticky=NSEW, padx=(PAD_S, 0))
        seqs.columnconfigure(0, weight=1)
        seqs.columnconfigure(1, weight=1)

        gen_widget = tk.Text(lf_gen, height=seq_text_rows, wrap="word")
        gen_widget.pack(fill="both", expand=True, padx=PAD_S // 2, pady=PAD_S // 2)
        mrna_widget = tk.Text(lf_mrna, height=seq_text_rows, wrap="word")
        mrna_widget.pack(fill="both", expand=True, padx=PAD_S // 2, pady=PAD_S // 2)
        return gen_widget, mrna_widget

    gen_text, mrna_text = _build_sequence_boxes(root_container)

    tool = ttk.LabelFrame(root_container, text="Run Options", style="SectionHeader.TLabelframe", padding=PAD_M)
    tool.grid(row=3, column=0, sticky=EW, pady=(0, PAD_M))

    primer3_var = tk.StringVar(value=defaults["primer3"])
    spidey_var = tk.StringVar(value=defaults["spidey"])
    ntthal_var = tk.StringVar(value=defaults["ntthal"])
    oligotm_var = tk.StringVar(value=defaults["oligotm"])
    spidey_large_var = tk.IntVar(value=0)
    mfeprimer_var = tk.StringVar(value=defaults["mfeprimer"])
    mfeprimer_transcriptome_fasta_var = tk.StringVar(value=defaults["mfeprimer_transcriptome_fasta"])
    target_ensembl_gene_id_var = tk.StringVar(value="")
    snp_non3p_policy_var = tk.StringVar(value="hard")
    snp_3p_window_var = tk.StringVar(value="7")
    enable_snp_check_var = tk.IntVar(value=1)
    show_snp_map_var = tk.IntVar(value=1)
    run_mfeprimer_var = tk.IntVar(value=1 if defaults["mfeprimer"] else 0)
    mfe_dg_cutoff_var = tk.StringVar(value="-2")
    spec_top50_var = tk.IntVar(value=1)
    spec_selection_mode_var = tk.StringVar(value="strict_pass")
    spec_remove_pct_var = tk.StringVar(value="10")
    run_ntthal_cutoff_var = tk.IntVar(value=1)
    ntthal_ext_cutoff_var = tk.StringVar(value="-3.5")
    max_pairs_var = tk.StringVar(value="100")
    gene_name_var = tk.StringVar(value="")
    max_genomic_view_bases_var = tk.StringVar(
        value=str(_sanitize_max_genomic_view_bases(runtime_settings.get("max_genomic_view_bases", default_max_genomic_view_bases)))
    )
    binary_profile_var = tk.StringVar(
        value=str(runtime_settings.get("binary_profile") or _infer_binary_profile_from_paths(defaults["primer3"], defaults["ntthal"]))
    )
    mfeprimer_spec_params_var = tk.StringVar(
        value=str(runtime_settings.get("mfeprimer_spec_params") or DEFAULT_SPEC_PARAMS_RAW)
    )
    spec_sensitivity_preset_var = tk.StringVar(
        value=preset_from_spec_param_raw(str(mfeprimer_spec_params_var.get() or DEFAULT_SPEC_PARAMS_RAW))
    )

    def _ensure_tool_path(var: tk.StringVar, stem: str, required_label: str) -> str:
        current = str(var.get() or "").strip()
        if current:
            return current
        resolved = _existing_default(_tool_bin_exec_default(stem), _path_exec_default(stem))
        if resolved:
            var.set(resolved)
            return resolved
        raise ValueError(f"{required_label} path is required.")

    def _build_specificity_controls(parent: ttk.LabelFrame) -> tuple[ttk.Checkbutton, ttk.Combobox, ttk.Checkbutton, ttk.Combobox]:
        controls = ttk.Frame(parent)
        controls.grid(row=0, column=0, sticky="nsew", padx=PAD_S, pady=(PAD_S, PAD_S))
        controls.columnconfigure(0, weight=1)
        controls.columnconfigure(1, weight=1)
        controls.rowconfigure(0, weight=1)
        controls.rowconfigure(1, weight=1)

        spec_chk = ttk.Checkbutton(
            controls,
            text="Off-target exclusion",
            variable=spec_top50_var,
            command=lambda: _update_snp_check_controls(),
        )
        spec_chk.grid(row=0, column=0, sticky="w", pady=(0, PAD_S // 2))
        sensitivity_box = ttk.Combobox(
            controls,
            textvariable=spec_sensitivity_preset_var,
            values=(SPEC_PRESET_STRICT, SPEC_PRESET_SOFT),
            state="readonly",
        )
        sensitivity_box.grid(row=0, column=1, sticky="ew", padx=(PAD_S, 0))
        snp_chk = ttk.Checkbutton(
            controls,
            text="Avoid SNPs",
            variable=enable_snp_check_var,
            command=lambda: _update_snp_check_controls(),
        )
        snp_chk.grid(row=1, column=0, sticky="w", pady=(PAD_S // 2, 0))
        snp_box = ttk.Combobox(
            controls,
            textvariable=snp_non3p_policy_var,
            values=("soft", "hard"),
            state="readonly",
        )
        snp_box.grid(row=1, column=1, sticky="ew", padx=(PAD_S, 0), pady=(PAD_S // 2, 0))
        return spec_chk, sensitivity_box, snp_chk, snp_box

    spec_top50_chk, spec_sensitivity_box, snp_check_chk, snp_policy_box = _build_specificity_controls(specificity_group)
    def _build_run_options_row(parent: ttk.LabelFrame) -> tuple[ttk.Entry, ttk.Button, ttk.Button, ttk.Entry]:
        ttk.Label(parent, text="Max pairs").grid(row=1, column=0, sticky="w")
        max_pairs = ttk.Entry(parent, textvariable=max_pairs_var, width=8)
        max_pairs.grid(row=1, column=1, padx=(PAD_XS, 0), sticky="w")
        copy_btn = ttk.Button(parent, text="Copy selected", command=lambda: _copy_selected(), bootstyle="secondary")
        copy_btn.grid(row=1, column=2, padx=(PAD_S, 0), sticky="w")
        export_btn = ttk.Button(
            parent,
            text="Export DNA Order",
            command=lambda: _export_selected_to_microsynth(),
            bootstyle="secondary",
        )
        export_btn.grid(row=1, column=3, padx=(PAD_S, 0), sticky="w")
        parent.columnconfigure(4, weight=1)
        parent.columnconfigure(5, weight=0)
        meta = ttk.Frame(parent)
        meta.grid(row=1, column=5, sticky="e")
        meta.columnconfigure(1, weight=1, minsize=130)
        meta.columnconfigure(3, weight=1, minsize=170)
        ttk.Label(meta, text="Gene").grid(row=0, column=0, sticky="e", padx=(10, PAD_XS))
        ttk.Entry(meta, textvariable=gene_name_var, state="readonly").grid(row=0, column=1, sticky="ew")
        ttk.Label(meta, text="Target Ensembl gene ID").grid(row=0, column=2, sticky="e", padx=(10, PAD_XS))
        target_entry = ttk.Entry(meta, textvariable=target_ensembl_gene_id_var)
        target_entry.grid(row=0, column=3, sticky="ew")
        return max_pairs, copy_btn, export_btn, target_entry

    max_pairs_entry, copy_sel_btn, export_ms_btn, target_gene_id_entry = _build_run_options_row(tool)

    def _build_results_group(parent: ttk.Frame) -> tuple[ttk.Treeview, ttk.Checkbutton, tk.Canvas, set[str]]:
        results = ttk.LabelFrame(parent, text="Results", style="SectionHeader.TLabelframe", padding=PAD_S)
        results.grid(row=4, column=0, sticky=NSEW, pady=(0, PAD_M))

        result_tree = ttk.Treeview(
            results,
            columns=[c[0] for c in COLUMNS],
            show="headings",
            selectmode="extended",
            height=result_tree_rows,
        )
        sortable_cols = {"amp", "pd", "pd_full"}
        for cid, title, width in COLUMNS:
            if cid in sortable_cols:
                result_tree.heading(cid, text=title, command=lambda c=cid: _on_result_sort_header(c))
            else:
                result_tree.heading(cid, text=title)
            result_tree.column(cid, width=width, anchor="w")
        ys = ttk.Scrollbar(results, orient="vertical", command=result_tree.yview)
        xs = ttk.Scrollbar(results, orient="horizontal", command=result_tree.xview)
        result_tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        result_tree.tag_configure("spec_pass", foreground="#1f7a1f")
        result_tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        map_ctrl = ttk.Frame(results)
        map_ctrl.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(4, 0))
        map_ctrl.columnconfigure(0, weight=1)
        snp_view_chk = ttk.Checkbutton(
            map_ctrl,
            text="SNP view (graph)",
            variable=show_snp_map_var,
            command=lambda: _safe_draw_cdna_map(),
        )
        snp_view_chk.grid(row=0, column=1, sticky="e")
        map_widget = tk.Canvas(
            results,
            height=map_canvas_height,
            bg="#f5f5f5",
            highlightthickness=1,
            highlightbackground="#c0c0c0",
        )
        map_widget.grid(row=3, column=0, columnspan=2, sticky="ew")
        results.rowconfigure(0, weight=1)
        results.columnconfigure(0, weight=1)
        return result_tree, snp_view_chk, map_widget, sortable_cols

    tree, map_snp_view_chk, map_canvas, sortable_result_columns = _build_results_group(root_container)

    def _build_status_bar(parent: ttk.Frame) -> tuple[tk.StringVar, ttk.Frame, ttk.Label, ttk.Progressbar]:
        status_text = tk.StringVar(value="Bereit.")
        frame = ttk.Frame(parent)
        frame.grid(row=5, column=0, sticky=EW)
        label = ttk.Label(frame, textvariable=status_text, justify="left", anchor="w")
        label.grid(row=0, column=0, sticky="ew")
        progress = ttk.Progressbar(frame, mode="indeterminate", length=120, bootstyle="info-striped")
        progress.grid(row=0, column=1, padx=(PAD_S, PAD_S), sticky=E)
        ttk.Sizegrip(frame).grid(row=0, column=2, sticky=SE)
        frame.columnconfigure(0, weight=1)

        def _resize_status_wrap(_evt: object | None = None) -> None:
            try:
                w = max(220, int(frame.winfo_width()) - 140)
            except Exception:
                w = 220
            label.configure(wraplength=w)

        frame.bind("<Configure>", _resize_status_wrap)
        return status_text, frame, label, progress

    status_var, status_bar, status_lbl, busy_progress = _build_status_bar(root_container)
    tooltip_max_pairs = "Maximum number of primer pairs returned after filtering."
    tooltip_target_gene = "Optional Ensembl gene ID used for transcript-specific checks."
    tooltip_spec = "Enable transcriptome specificity filter for off-target exclusion."
    _attach_tooltip(max_pairs_entry, tooltip_max_pairs)
    _attach_tooltip(target_gene_id_entry, tooltip_target_gene)
    _attach_tooltip(spec_top50_chk, tooltip_spec)

    state: dict[str, object] = {
        "rows": [],
        "rows_after_ntthal": None,
        "parsed_primer3": None,
        "checked": set(),
        "spidey_output": "",
        "spidey_meta": {"used": False, "boundaries": []},
        "design_stats": None,
        "last_filter_counts": None,
        "last_runtime_breakdown": None,
        "payload": None,
        "template_len": 0,
        "bounds": [],
        "snp_cdna_positions": [],
        "current_snp_bed": "",
        "selected_row": None,
        "orf": None,
        "ensembl_gene": "",
        "ensembl_gene_id": "",
        "ensembl_species": "Homo_sapiens",
        "ensembl_type": "cdna",
        "map_redraw_job": None,
        "last_map_width": 0,
        "post_filter_busy": False,
        "post_filter_job_id": 0,
        "design_busy": False,
        "run_status_popup": None,
        "run_status_popup_msg_var": None,
        "run_status_popup_active": False,
        "full_status_text": "Ready.",
        "exon_hitboxes": [],
        "exon_hover_items": None,
        "hover_exon_no": None,
        "spec_passed": set(),
        "spec_pass_keys": set(),
        "spec_db_available": True,
        "mfeprimer_available": bool(str(defaults["mfeprimer"]).strip()) and Path(str(defaults["mfeprimer"])).is_file(),
        "result_sort_col": "pd",
        "result_sort_desc": True,
        "genomic_seq_full": "",
    }
    spec_top50_forced_off = {"value": False}
    snp_check_forced_off = {"value": False}
    snp_view_forced_off = {"value": False}

    def _save_runtime_settings() -> None:
        try:
            runtime_settings_path.parent.mkdir(parents=True, exist_ok=True)
            max_bases = _sanitize_max_genomic_view_bases(max_genomic_view_bases_var.get())
            max_genomic_view_bases_var.set(str(max_bases))
            payload = {
                "binary_profile": str(binary_profile_var.get() or "").strip(),
                "max_genomic_view_bases": int(max_bases),
                "mfeprimer_spec_params": str(mfeprimer_spec_params_var.get() or "").strip(),
            }
            runtime_settings_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
        except Exception:
            pass

    spec_sync_guard = {"busy": False}

    def _sync_spec_params_from_preset(*_args: object) -> None:
        if bool(spec_sync_guard.get("busy")):
            return
        try:
            spec_sync_guard["busy"] = True
            raw = spec_param_raw_for_preset(spec_sensitivity_preset_var.get())
            if str(mfeprimer_spec_params_var.get() or "").strip() != raw:
                mfeprimer_spec_params_var.set(raw)
        finally:
            spec_sync_guard["busy"] = False

    def _sync_preset_from_spec_params(*_args: object) -> None:
        if bool(spec_sync_guard.get("busy")):
            return
        try:
            spec_sync_guard["busy"] = True
            preset = preset_from_spec_param_raw(str(mfeprimer_spec_params_var.get() or ""))
            if str(spec_sensitivity_preset_var.get() or "") != preset:
                spec_sensitivity_preset_var.set(preset)
        finally:
            spec_sync_guard["busy"] = False

    spec_sensitivity_preset_var.trace_add("write", _sync_spec_params_from_preset)
    mfeprimer_spec_params_var.trace_add("write", _sync_preset_from_spec_params)
    _sync_preset_from_spec_params()

    def _max_genomic_view_bases() -> int:
        n = _sanitize_max_genomic_view_bases(max_genomic_view_bases_var.get())
        if str(n) != str(max_genomic_view_bases_var.get()).strip():
            max_genomic_view_bases_var.set(str(n))
        return n

    def _set_genomic_sequence_display(seq: str) -> None:
        full = str(seq or "")
        state["genomic_seq_full"] = full
        limit = _max_genomic_view_bases()
        shown = full
        if len(full) > limit:
            shown = (
                full[:limit]
                + f"\n...[truncated: showing first {limit:,} of {len(full):,} bases; adjust in Options]..."
            )
        gen_text.delete("1.0", "end")
        gen_text.insert("1.0", shown)

    def _binary_profile_targets(profile_name: str) -> dict[str, str]:
        default_bin_roots = _tool_bin_roots()
        clang_roots = [root / "clang_profiles" for root in default_bin_roots]
        name = str(profile_name or "").strip()

        def _resolve_exec(base_dirs: list[Path], stem: str, fallback_dirs: list[Path] | None = None) -> str:
            roots = list(base_dirs)
            if fallback_dirs:
                roots.extend(fallback_dirs)
            for root in roots:
                for exec_name in candidate_exec_names(stem):
                    p = root / exec_name
                    if p.exists():
                        return str(p)
            cands = candidate_exec_names(stem)
            fallback_base = (base_dirs or fallback_dirs or [Path(_primerl_path("tools", "bin"))])[0]
            return str(fallback_base / cands[0]) if cands else str(fallback_base / stem)

        if name == "Clang znver2":
            bases = [root / "znver2" for root in clang_roots]
            return {
                "primer3": _resolve_exec(bases, "primer3_core"),
                "ntthal": _resolve_exec(bases, "ntthal"),
                "oligotm": _resolve_exec(bases, "oligotm"),
            }
        if name == "Clang znver4":
            bases = [root / "znver4" for root in clang_roots]
            return {
                "primer3": _resolve_exec(bases, "primer3_core"),
                "ntthal": _resolve_exec(bases, "ntthal"),
                "oligotm": _resolve_exec(bases, "oligotm"),
            }
        if name == "Clang x86-64-v3":
            bases = [root / "x86_64_v3" for root in clang_roots]
            return {
                "primer3": _resolve_exec(bases, "primer3_core"),
                "ntthal": _resolve_exec(bases, "ntthal"),
                "oligotm": _resolve_exec(bases, "oligotm"),
            }
        if name == "Apple Silicon native":
            bases = [root / "apple_silicon" for root in clang_roots]
            return {
                "primer3": _resolve_exec(bases, "primer3_core", fallback_dirs=default_bin_roots),
                "ntthal": _resolve_exec(bases, "ntthal", fallback_dirs=default_bin_roots),
                "oligotm": _resolve_exec(bases, "oligotm", fallback_dirs=default_bin_roots),
            }
        bases = default_bin_roots
        return {
            "primer3": _resolve_exec(bases, "primer3_core"),
            "ntthal": _resolve_exec(bases, "ntthal"),
            "oligotm": _resolve_exec(bases, "oligotm"),
        }

    def _apply_binary_profile(profile_name: str, notify: bool = True, persist: bool = True) -> bool:
        canonical = str(profile_name or "").strip()
        if canonical in {"Original", "Perf Optimized"}:
            canonical = "Upstream original src"
        if canonical == "Clang Primer3":
            canonical = "Clang x86-64-v3"
        if canonical not in BINARY_PROFILE_CHOICES:
            canonical = "Upstream original src"
        targets = _binary_profile_targets(canonical)
        missing: list[str] = []
        if Path(targets["primer3"]).exists():
            primer3_var.set(targets["primer3"])
        else:
            missing.append(targets["primer3"])
        if Path(targets["ntthal"]).exists():
            ntthal_var.set(targets["ntthal"])
        else:
            missing.append(targets["ntthal"])
        if Path(targets["oligotm"]).exists():
            oligotm_var.set(targets["oligotm"])
        else:
            missing.append(targets["oligotm"])
        binary_profile_var.set(canonical)
        if persist:
            _save_runtime_settings()
        if notify:
            if missing:
                _set_status("Binary profile applied with missing files; run clang build script and verify paths.")
            else:
                _set_status(f"Binary profile set to {canonical}.")
        return len(missing) == 0

    def _detect_and_apply_cpu_profile() -> None:
        detected_profile, reason = _detect_best_binary_profile_for_cpu()
        ok = _apply_binary_profile(detected_profile, notify=False, persist=False)
        if not ok:
            _apply_binary_profile("Upstream original src", notify=False, persist=False)
            _set_status(f"CPU auto-detect chose {detected_profile} ({reason}), but binaries were missing; using Upstream original src.")
        else:
            _set_status(f"CPU auto-detect selected {detected_profile} ({reason}).")
        _save_runtime_settings()

    def _sync_gene_display() -> None:
        gene_name_var.set(str(state.get("ensembl_gene") or ""))

    def _active_target_gene_id() -> str:
        manual = _normalize_ensembl_gene_id(target_ensembl_gene_id_var.get())
        if manual:
            return manual
        return _normalize_ensembl_gene_id(str(state.get("ensembl_gene_id") or ""))

    def _active_target_species_slug() -> str:
        gid = _active_target_gene_id()
        inferred = _infer_species_slug_from_gene_id(gid)
        if inferred:
            return inferred
        return str(state.get("ensembl_species") or "").strip().lower()

    def _default_snp_check_enabled_for_species(species_slug: str) -> bool:
        slug = str(species_slug or "").strip().lower()
        if slug == "mus_musculus":
            return False
        if slug in {
            "danio_rerio",
            "rattus_norvegicus",
            "drosophila_melanogaster",
            "caenorhabditis_elegans",
        }:
            return True
        return True

    def _is_human_target_context() -> bool:
        gid = _active_target_gene_id()
        if gid.startswith("ENSG"):
            return True
        species_txt = str(state.get("ensembl_species") or "").strip().lower()
        return species_txt == "homo_sapiens"

    def _update_map_snp_view_controls() -> None:
        is_human_target = _is_human_target_context()
        if is_human_target:
            if bool(show_snp_map_var.get()):
                snp_view_forced_off["value"] = True
                show_snp_map_var.set(0)
            map_snp_view_chk.configure(state="disabled")
            return
        map_snp_view_chk.configure(state="normal")
        if snp_view_forced_off["value"]:
            show_snp_map_var.set(1)
            snp_view_forced_off["value"] = False

    def _update_snp_check_controls() -> None:
        is_human_target = _is_human_target_context()
        if is_human_target:
            if bool(enable_snp_check_var.get()):
                snp_check_forced_off["value"] = True
                enable_snp_check_var.set(0)
            snp_check_chk.configure(state="disabled")
            snp_policy_box.configure(state="disabled")
            return

        spec_enabled = bool(spec_top50_var.get())
        if not spec_enabled:
            if bool(enable_snp_check_var.get()):
                snp_check_forced_off["value"] = True
                enable_snp_check_var.set(0)
            snp_check_chk.configure(state="disabled")
            snp_policy_box.configure(state="disabled")
            return

        if snp_check_forced_off["value"]:
            species_slug = _active_target_species_slug()
            enable_snp_check_var.set(1 if _default_snp_check_enabled_for_species(species_slug) else 0)
            snp_check_forced_off["value"] = False

        snp_check_chk.configure(state="normal")
        if bool(enable_snp_check_var.get()):
            snp_policy_box.configure(state="readonly")
        else:
            snp_policy_box.configure(state="disabled")

    def _update_spec_sensitivity_control() -> None:
        if spec_top50_chk.instate(("disabled",)):
            spec_sensitivity_box.configure(state="disabled")
        else:
            spec_sensitivity_box.configure(state="readonly")

    def _update_spec_toggle_availability(*_args: object) -> None:
        has_target_id = bool(_normalize_ensembl_gene_id(target_ensembl_gene_id_var.get()))
        has_spec_db = bool(state.get("spec_db_available", True))
        has_mfeprimer = bool(state.get("mfeprimer_available", True))
        if has_target_id and has_spec_db and has_mfeprimer:
            spec_top50_chk.configure(state="normal")
            _update_spec_sensitivity_control()
            if spec_top50_forced_off["value"]:
                spec_top50_var.set(1)
                spec_top50_forced_off["value"] = False
            _update_snp_check_controls()
            _update_map_snp_view_controls()
            return

        if bool(spec_top50_var.get()):
            spec_top50_forced_off["value"] = True
        spec_top50_var.set(0)
        spec_top50_chk.configure(state="disabled")
        _update_spec_sensitivity_control()
        _update_snp_check_controls()
        _update_map_snp_view_controls()

    def _infer_species_slug_from_gene_id(gene_id: str) -> str:
        gid = _normalize_ensembl_gene_id(gene_id)
        if not gid:
            return ""
        if gid.startswith("FBGN") or gid.startswith("ENSDMEG"):
            return "drosophila_melanogaster"
        if gid.startswith("ENSDARG"):
            return "danio_rerio"
        if gid.startswith("ENSMUSG"):
            return "mus_musculus"
        if gid.startswith("ENSRNOG"):
            return "rattus_norvegicus"
        if gid.startswith("ENSG"):
            return "homo_sapiens"
        return ""

    def _find_best_ensembl_transcriptome_db(species_slug: str) -> str:
        slug = (species_slug or "").strip().lower()
        if not slug:
            return ""
        db_dir = Path(_primerl_data_path("databases", "ensembl"))
        if not db_dir.exists() or not db_dir.is_dir():
            return ""
        files = [p for p in db_dir.iterdir() if p.is_file() and p.suffix.lower() in {".fa", ".fasta"}]
        if not files:
            return ""

        token = slug
        token_alt = slug.replace("_", ".")
        token_alt2 = slug.replace("_", "")
        matches: list[Path] = []
        for p in files:
            n = p.name.lower()
            if token in n or token_alt in n or token_alt2 in n:
                matches.append(p)
        if not matches:
            return ""

        def _score(p: Path) -> tuple[int, int, int]:
            name = p.name.lower()
            indexed = 1 if Path(str(p) + ".primerqc").exists() else 0
            cdna_all = 1 if ".cdna.all." in name else 0
            return (indexed, cdna_all, len(name))

        matches.sort(key=_score, reverse=True)
        return str(matches[0])

    def _db_matches_species_slug(db_path: str, species_slug: str) -> bool:
        slug = (species_slug or "").strip().lower()
        if not slug:
            return True
        name = Path(str(db_path or "")).name.lower()
        if not name:
            return False
        slug_dot = slug.replace("_", ".")
        slug_compact = slug.replace("_", "")
        return slug in name or slug_dot in name or slug_compact in name

    def _set_spec_db_available(available: bool) -> None:
        state["spec_db_available"] = bool(available)
        _update_spec_toggle_availability()

    def _install_ensembl_db_for_species_async(
        *,
        species_slug: str,
        on_success: Callable[[str], None] | None = None,
        on_cancel_or_fail: Callable[[str], None] | None = None,
        show_cancel_popup: bool = False,
    ) -> None:
        slug = str(species_slug or "").strip().lower()
        if not slug:
            if on_cancel_or_fail is not None:
                on_cancel_or_fail("missing species")
            return
        mfe_path = Path((mfeprimer_var.get() or "").strip())
        if not mfe_path.exists() or mfe_path.is_dir():
            if on_cancel_or_fail is not None:
                on_cancel_or_fail("valid MFEprimer executable not configured")
            return

        cancel_flag = {"value": False}
        cancel_sentinel = "__CANCELLED__"
        progress_win: tk.Toplevel | None = None
        progress_text_var: tk.StringVar | None = None
        slug_label_map = {slug_txt: label for label, slug_txt in ENSEMBL_DB_SPECIES_CHOICES}
        species_label = slug_label_map.get(slug, slug.replace("_", " "))

        def _update_progress(txt: str) -> None:
            if progress_text_var is not None:
                progress_text_var.set(txt)
            _set_status(txt)

        def _finish(ok: bool, msg: str, fasta_path: str = "") -> None:
            if progress_win is not None and progress_win.winfo_exists():
                try:
                    progress_win.destroy()
                except Exception:
                    pass
            if ok:
                if fasta_path:
                    mfeprimer_transcriptome_fasta_var.set(fasta_path)
                _set_spec_db_available(True)
                if on_success is not None:
                    on_success(fasta_path)
                _set_status(msg)
            else:
                _set_spec_db_available(False)
                if on_cancel_or_fail is not None:
                    on_cancel_or_fail(msg)
                if msg != cancel_sentinel:
                    _set_status(f"Ensembl DB install failed: {msg}")

        if show_cancel_popup:
            progress_win = tk.Toplevel(root)
            progress_win.title("Install Transcriptome DB")
            progress_win.transient(root)
            frm = ttk.Frame(progress_win, padding=10)
            frm.pack(fill="both", expand=True)
            progress_text_var = tk.StringVar(value=f"Preparing install for {species_label} ...")
            ttk.Label(frm, textvariable=progress_text_var, wraplength=420, justify="left").grid(row=0, column=0, sticky="w")

            def _cancel_install() -> None:
                cancel_flag["value"] = True
                _update_progress("Cancel requested ...")

            ttk.Button(frm, text="Cancel", command=_cancel_install).grid(row=1, column=0, sticky="e", pady=(8, 0))
            frm.columnconfigure(0, weight=1)
            _center_dialog_on_parent(progress_win, root)
            progress_win.grab_set()
            progress_win.focus_force()

        def _worker() -> None:
            try:
                if cancel_flag["value"]:
                    raise RuntimeError(cancel_sentinel)
                db_dir = Path(_primerl_data_path("databases", "ensembl"))
                db_dir.mkdir(parents=True, exist_ok=True)

                fname, dl_url = _resolve_ensembl_cdna_download(slug)
                gz_path = db_dir / fname
                fa_name = fname[:-3] if fname.lower().endswith(".gz") else fname
                fasta_path = db_dir / fa_name

                root.after(0, lambda: _update_progress(f"Downloading Ensembl transcriptome: {fname} ..."))
                _download_http_file(dl_url, gz_path, timeout_sec=120)
                if cancel_flag["value"]:
                    raise RuntimeError(cancel_sentinel)

                root.after(0, lambda: _update_progress(f"Extracting {fname} ..."))
                _gunzip_file(gz_path, fasta_path)
                try:
                    gz_path.unlink(missing_ok=True)
                except Exception:
                    pass
                if cancel_flag["value"]:
                    raise RuntimeError(cancel_sentinel)

                cpu = max(1, min(16, _available_cpu_threads(reserve=1, max_cap=16)))
                idx_cmd = [str(mfe_path), "index", "-i", str(fasta_path), "-c", str(cpu), "-f"]
                root.after(0, lambda: _update_progress(f"Indexing transcriptome with MFEprimer ({cpu} CPU) ..."))
                flags = 0
                if os.name == "nt":
                    flags |= int(getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000))
                proc = subprocess.Popen(
                    idx_cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    creationflags=flags,
                )
                while proc.poll() is None:
                    if cancel_flag["value"]:
                        try:
                            proc.terminate()
                        except Exception:
                            pass
                        raise RuntimeError(cancel_sentinel)
                    time.sleep(0.2)
                stdout_txt, stderr_txt = proc.communicate()
                if proc.returncode != 0:
                    err_txt = (stderr_txt or stdout_txt or "MFEprimer index failed").strip()
                    raise RuntimeError(err_txt[:500])
                if cancel_flag["value"]:
                    raise RuntimeError(cancel_sentinel)

                idx_path = Path(str(fasta_path) + ".primerqc")
                if not idx_path.exists():
                    raise RuntimeError("Index command finished, but .primerqc file was not created.")

                root.after(
                    0,
                    lambda: _finish(
                        True,
                        f"Installed Ensembl transcriptome DB for {species_label}: {fasta_path.name}",
                        str(fasta_path),
                    ),
                )
            except Exception as exc:
                emsg = str(exc or "install failed")
                root.after(0, lambda e=emsg: _finish(False, e, ""))

        threading.Thread(target=_worker, name="ensembl-db-install", daemon=True).start()

    def _auto_select_transcriptome_db_for_target(
        *,
        gene_id: str,
        species_hint: str = "",
    ) -> str:
        slug = _infer_species_slug_from_gene_id(gene_id)
        if not slug and species_hint:
            slug = (species_hint or "").strip().lower()
        if not slug:
            return ""
        best = _find_best_ensembl_transcriptome_db(slug)
        cur = (mfeprimer_transcriptome_fasta_var.get() or "").strip()
        if not best:
            # Avoid carrying a stale DB from a different species.
            if cur and not _db_matches_species_slug(cur, slug):
                mfeprimer_transcriptome_fasta_var.set("")
                _set_spec_db_available(False)
                return (
                    f"No transcriptome DB found for {slug}; cleared mismatched DB "
                    f"{Path(cur).name}. Install/select a matching DB to use Spec testing."
                )
            _set_spec_db_available(False)
            return f"No transcriptome DB found for {slug}. Install/select a matching DB to use Spec testing."
        if cur and Path(cur) == Path(best):
            _set_spec_db_available(True)
            return ""
        mfeprimer_transcriptome_fasta_var.set(best)
        _set_spec_db_available(True)
        return f"Auto-selected transcriptome DB for {slug}: {Path(best).name}"

    busy_controls: list[tk.Widget] = []

    def _draw_cdna_map() -> None:
        _hide_exon_hover()
        map_canvas.delete("all")
        width = max(220, map_canvas.winfo_width())
        left = 30
        right = width - 30
        y_mid = 58
        y_top = 46
        y_bottom = 70
        length = int(state.get("template_len") or 0)
        rows = list(state.get("rows") or [])
        bounds = [int(b) for b in (state.get("bounds") or [])]
        snp_pos = [int(p) for p in (state.get("snp_cdna_positions") or [])]
        show_snp_map = bool(int(show_snp_map_var.get() or 0)) and (not _is_human_target_context())
        orf = state.get("orf")

        if length <= 0:
            state["exon_hitboxes"] = []
            state["hover_exon_no"] = None
            map_canvas.create_text(width // 2, y_mid, text="Run primer search to draw cDNA/intron-exon map", fill="#555")
            return

        def x_for(pos: int) -> float:
            p = max(0, min(length, pos))
            return left + ((right - left) * (p / max(1, length)))

        # cDNA baseline with ORF/UTR coloring (v036-style cue):
        # coding region dark blue, UTRs yellow.
        if isinstance(orf, tuple) and len(orf) == 2:
            o_start = max(0, min(length, int(orf[0])))
            o_end = max(0, min(length, int(orf[1])))
            if o_start > o_end:
                o_start, o_end = o_end, o_start

            if o_start > 0:
                map_canvas.create_line(left, y_mid, x_for(o_start), y_mid, width=10, fill="#e6c84f")
            map_canvas.create_line(x_for(o_start), y_mid, x_for(o_end), y_mid, width=10, fill="#1f4a7a")
            if o_end < length:
                map_canvas.create_line(x_for(o_end), y_mid, right, y_mid, width=10, fill="#e6c84f")
        else:
            map_canvas.create_line(left, y_mid, right, y_mid, width=10, fill="#1f4a7a")
        map_canvas.create_text(left - 12, y_mid, text="5'", fill="#333")
        map_canvas.create_text(right + 12, y_mid, text="3'", fill="#333")

        # Build exon hover hitboxes in cDNA coordinates.
        exon_bounds = [0] + sorted(set(b for b in bounds if 0 < b < length)) + [length]
        exon_hitboxes: list[tuple[float, float, float, float, int]] = []
        for idx in range(len(exon_bounds) - 1):
            s = exon_bounds[idx]
            e = exon_bounds[idx + 1]
            x0 = float(x_for(s))
            x1 = float(x_for(e))
            left_x = min(x0, x1)
            right_x = max(x0, x1)
            if right_x - left_x < 2.0:
                right_x = left_x + 2.0
            exon_hitboxes.append((left_x, right_x, float(y_top - 6), float(y_bottom + 6), idx + 1))
        state["exon_hitboxes"] = exon_hitboxes
        state["hover_exon_no"] = None

        for b in bounds:
            if 0 < b < length:
                bx = x_for(b)
                map_canvas.create_line(bx, y_top, bx, y_bottom, fill="white", width=3)
                map_canvas.create_line(bx, y_top, bx, y_bottom, fill="#0f2742", width=1)
        if show_snp_map:
            for p in snp_pos:
                if 0 < p <= length:
                    sx = x_for(p)
                    map_canvas.create_line(sx, y_top, sx, y_bottom, fill="#b02121", width=1)
        if show_snp_map and snp_pos:
            map_canvas.create_text(right - 65, y_top - 12, text=f"SNPs: {len(snp_pos)}", fill="#b02121")

        sel = state.get("selected_row")
        if sel is None:
            hint = "Retrieved cDNA map (exon boundaries). Run primer search."
            if show_snp_map:
                hint = "Retrieved cDNA map (exon boundaries + SNP markers). Run primer search."
            if rows:
                hint = "Select a primer pair row to map it"
            map_canvas.create_text(width // 2, 98, text=hint, fill="#666")
            return
        if not isinstance(sel, int) or sel < 0 or sel >= len(rows):
            return

        row = rows[sel]
        f_start = int(row[1])
        f_end = f_start + int(row[2])
        r_end = int(row[8])
        r_start = r_end - int(row[6])
        amp_start = f_start
        amp_end = r_end

        map_canvas.create_rectangle(x_for(amp_start), 82, x_for(amp_end), 94, fill="#8ec5ff", outline="#4a90d9")
        map_canvas.create_rectangle(x_for(f_start), y_top - 8, x_for(f_end), y_top + 2, fill="#2e8b57", outline="#1f5f3b")
        map_canvas.create_rectangle(x_for(r_start), y_bottom - 2, x_for(r_end), y_bottom + 8, fill="#a12e2e", outline="#6e1e1e")
        map_canvas.create_text(
            width // 2,
            108,
            text=f"Pair {sel + 1}: F {f_start}-{f_end}, R {r_start}-{r_end}, Amp {int(row[9])} bp",
            fill="#333",
        )

    def _safe_draw_cdna_map() -> None:
        try:
            _draw_cdna_map()
        except Exception:
            # Map rendering must never block the primary results table flow.
            pass

    def _hide_exon_hover() -> None:
        raw = state.get("exon_hover_items")
        if isinstance(raw, tuple):
            for iid in raw:
                try:
                    map_canvas.delete(iid)
                except Exception:
                    pass
        state["exon_hover_items"] = None
        state["hover_exon_no"] = None

    def _show_exon_hover(x: int, y: int, exon_no: int) -> None:
        _hide_exon_hover()
        label = f"Exon {exon_no}"
        tx = max(8, min(int(map_canvas.winfo_width()) - 90, int(x) + 12))
        ty = max(8, int(y) - 18)
        txt_id = map_canvas.create_text(tx, ty, text=label, anchor="nw", fill="#222")
        bbox = map_canvas.bbox(txt_id)
        if not bbox:
            state["exon_hover_items"] = (txt_id,)
            return
        bg_id = map_canvas.create_rectangle(
            bbox[0] - 4,
            bbox[1] - 2,
            bbox[2] + 4,
            bbox[3] + 2,
            fill="#fffbe6",
            outline="#b8b08a",
        )
        map_canvas.tag_raise(txt_id, bg_id)
        state["exon_hover_items"] = (bg_id, txt_id)

    def _on_map_hover(evt: tk.Event[tk.Misc]) -> None:
        hitboxes = state.get("exon_hitboxes")
        if not isinstance(hitboxes, list) or not hitboxes:
            _hide_exon_hover()
            return
        x = float(evt.x)
        y = float(evt.y)
        exon_no: int | None = None
        for x0, x1, y0, y1, ex_no in hitboxes:
            if x0 <= x <= x1 and y0 <= y <= y1:
                exon_no = int(ex_no)
                break
        if exon_no is None:
            if state.get("hover_exon_no") is not None:
                _hide_exon_hover()
            return
        if int(state.get("hover_exon_no") or -1) == exon_no:
            return
        state["hover_exon_no"] = exon_no
        _show_exon_hover(int(x), int(y), exon_no)

    def _schedule_draw_cdna_map(_evt: object | None = None) -> None:
        try:
            new_w = int(getattr(_evt, "width", map_canvas.winfo_width()))
        except Exception:
            new_w = map_canvas.winfo_width()
        last_w = int(state.get("last_map_width") or 0)
        if last_w and abs(new_w - last_w) < 6:
            return
        state["last_map_width"] = new_w

        pending = state.get("map_redraw_job")
        if pending:
            try:
                root.after_cancel(pending)
            except Exception:
                pass
        rows_n = len(list(state.get("rows") or []))
        debounce_ms = 120 if rows_n > 200 else 60
        state["map_redraw_job"] = root.after(debounce_ms, _safe_draw_cdna_map)

    def _copy_selected() -> None:
        sel = tree.selection()
        if not sel:
            return
        header = "\t".join([c[1] for c in COLUMNS])
        rows = ["\t".join(tree.item(i, "values")) for i in sel]
        root.clipboard_clear()
        root.clipboard_append(header + "\n" + "\n".join(rows))
        _set_status(f"Copied {len(sel)} selected rows.")

    def _selected_result_row_indices() -> list[int]:
        checked = sorted(_checked_ids())
        if checked:
            return checked
        out: list[int] = []
        for iid in tree.selection():
            try:
                out.append(int(iid))
            except ValueError:
                continue
        return sorted(set(out))

    def _export_selected_to_microsynth() -> None:
        idxs = _selected_result_row_indices()
        all_rows = list(state.get("rows") or [])
        if not idxs:
            _set_status("No primer pairs selected. Check rows or select rows first.")
            return
        rows: list[list[object]] = [all_rows[i] for i in idxs if 0 <= i < len(all_rows)]
        if not rows:
            _set_status("Selected rows are not available.")
            return

        load_workbook = _load_openpyxl_workbook_loader()
        if load_workbook is None:
            _set_status("openpyxl is required for Excel export but is not available.")
            return

        tpl_path = _resolve_microsynth_template_path()
        gene_raw = str(gene_name_var.get() or state.get("ensembl_gene") or target_ensembl_gene_id_var.get() or "gene")
        gene_token = _sanitize_oligo_name_token(gene_raw)

        try:
            used_fallback_template = False
            if tpl_path.exists():
                with warnings.catch_warnings():
                    warnings.filterwarnings(
                        "ignore",
                        message="Cannot parse header or footer so it will be ignored",
                        category=UserWarning,
                    )
                    warnings.filterwarnings(
                        "ignore",
                        message="wmf image format is not supported so the image is being dropped",
                        category=UserWarning,
                    )
                    wb = load_workbook(tpl_path)
                if "DNA Order" not in wb.sheetnames:
                    raise RuntimeError("Template sheet 'DNA Order' not found.")
                ws = wb["DNA Order"]
            else:
                # Keep export working in source-only app bundles where template xlsx is intentionally not shipped.
                from openpyxl import Workbook  # type: ignore[import-not-found]

                wb = Workbook()
                ws = wb.active
                ws.title = "DNA Order"
                ws.cell(row=1, column=1, value="Name")
                ws.cell(row=1, column=2, value="Sequence")
                ws.cell(row=1, column=3, value="Length")
                used_fallback_template = True

            row_ptr = 2
            while True:
                a = ws.cell(row=row_ptr, column=1).value
                b = ws.cell(row=row_ptr, column=2).value
                if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
                    break
                row_ptr += 1

            pair_no = 1
            written = 0
            for row in rows:
                f_seq = str(row[0]).strip()
                r_seq = str(row[4]).strip()
                if not f_seq or not r_seq:
                    continue
                f_len = len(re.sub(r"[^A-Za-z]", "", f_seq))
                r_len = len(re.sub(r"[^A-Za-z]", "", r_seq))

                ws.cell(row=row_ptr, column=1, value=f"RT-{gene_token} {pair_no}F")
                ws.cell(row=row_ptr, column=2, value=f_seq)
                ws.cell(row=row_ptr, column=3, value=f_len)
                row_ptr += 1

                ws.cell(row=row_ptr, column=1, value=f"RT-{gene_token} {pair_no}R")
                ws.cell(row=row_ptr, column=2, value=r_seq)
                ws.cell(row=row_ptr, column=3, value=r_len)
                row_ptr += 1

                pair_no += 1
                written += 2

            if written == 0:
                _set_status("No valid primer rows were exported.")
                return

            out_dir = Path(_primerl_data_path("runtime", "exports"))
            out_dir.mkdir(parents=True, exist_ok=True)
            ts = time.strftime("%Y%m%d_%H%M%S")
            out_path = out_dir / f"MicrosynthUploadFormDNA_filled_{gene_token}_{ts}.xlsx"
            wb.save(out_path)
            if used_fallback_template:
                _set_status(
                    f"Exported {written} oligos to {out_path} (built-in fallback sheet; Microsynth template not bundled)."
                )
            else:
                _set_status(f"Exported {written} oligos to {out_path}")
            try:
                open_file(str(out_path))
            except Exception:
                pass
        except Exception as exc:
            _set_status(f"Microsynth export failed: {exc}")

    def _show_filter_pie() -> None:
        raw = state.get("last_filter_counts")
        if not isinstance(raw, dict):
            _set_status("No filter breakdown yet. Run primer search first.")
            return
        fc = dict(raw)
        try:
            parsed = max(0, int(fc.get("parsed", 0)))
            skipped_span = max(0, int(fc.get("skipped_span", 0)))
            skipped_overlap = max(0, int(fc.get("skipped_overlap", 0)))
            skipped_repeat_run = max(0, int(fc.get("skipped_repeat_run", 0)))
            after_filter_options = max(0, int(fc.get("after_filter_options", 0)))
            after_ntthal = max(0, int(fc.get("after_ntthal", 0)))
            after_ntthal_cutoff = max(0, int(fc.get("after_ntthal_cutoff", 0)))
            after_mfeprimer = max(0, int(fc.get("after_mfeprimer", 0)))
            after_mfeprimer_spec = max(0, int(fc.get("after_mfeprimer_spec", 0)))
            final_returned = max(0, int(fc.get("final_returned", 0)))
            spec_target_removed = max(0, int(fc.get("spec_target_removed", 0)))
            spec_snp_removed = max(0, int(fc.get("spec_snp_removed", 0)))
        except Exception:
            _set_status("Filter breakdown is malformed.")
            return

        ntthal_refine_drop = max(0, after_filter_options - after_ntthal)
        ntthal_cutoff_drop = max(0, after_ntthal - after_ntthal_cutoff)
        mfeprimer_drop = max(0, after_ntthal_cutoff - after_mfeprimer)
        low_dg_drop = ntthal_refine_drop + ntthal_cutoff_drop + mfeprimer_drop
        total_removed = max(0, parsed - final_returned)
        spec_total_drop = max(0, after_mfeprimer - after_mfeprimer_spec)
        spec_snp_removed = min(spec_snp_removed, spec_total_drop)
        spec_target_removed = min(spec_target_removed, max(0, spec_total_drop - spec_snp_removed))
        spec_other_removed = max(0, spec_total_drop - spec_snp_removed - spec_target_removed)
        known_non_gc = (
            skipped_span
            + skipped_overlap
            + skipped_repeat_run
            + low_dg_drop
            + spec_target_removed
            + spec_snp_removed
            + spec_other_removed
        )
        gc_drop = max(0, total_removed - known_non_gc)

        parts: list[tuple[str, int, str]] = [
            ("Filtered by span boundary", skipped_span, "#e15759"),
            ("Filtered by overlap boundary", skipped_overlap, "#b07aa1"),
            ("Filtered by GC filter", gc_drop, "#9c755f"),
            ("Filtered by repeat/run filter", skipped_repeat_run, "#f28e2b"),
            ("Filtered for low dG", low_dg_drop, "#4e79a7"),
            ("Specificity check: Not target-specific", spec_target_removed, "#edc948"),
            ("Specificity check: SNP-filtered", spec_snp_removed, "#d62728"),
            ("Specificity check: Other/unclassified", spec_other_removed, "#8f63bf"),
        ]

        win = tk.Toplevel(root)
        win.title("Lost Primer Pie")
        win.transient(root)
        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        canvas = tk.Canvas(frm, width=420, height=320, bg="white")
        canvas.grid(row=0, column=0, rowspan=2, sticky="nsew")
        legend = tk.Text(frm, width=52, height=18, wrap="word")
        legend.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        legend.configure(state="normal")
        legend.delete("1.0", "end")
        legend.tag_configure("hdr", font=("TkDefaultFont", 10, "bold"))

        x0, y0, x1, y1 = 30, 20, 300, 290
        start = 0.0
        pie_total = sum(max(0, int(v)) for _, v, _ in parts)
        nonzero_parts = [(label, value, color) for label, value, color in parts if value > 0]
        cx = (x0 + x1) / 2.0
        cy = (y0 + y1) / 2.0
        rx = max(1.0, (x1 - x0) / 2.0)
        ry = max(1.0, (y1 - y0) / 2.0)
        legend.insert("end", "Lost Primer Pie\n", ("hdr",))
        for label, value, color in parts:
            pct = 0.0 if total_removed <= 0 else ((100.0 * value) / total_removed)
            swatch_tag = f"swatch_{color.replace('#', '')}"
            legend.tag_configure(swatch_tag, background=color)
            legend.insert("end", "  ", (swatch_tag,))
            legend.insert("end", " ")
            legend.insert("end", f"{label}: {value} ({pct:.1f}%)\n")
        drawn_slices = 0
        consumed = 0.0
        for idx, (_label, value, color) in enumerate(nonzero_parts):
            extent = 0.0
            if pie_total > 0:
                if idx == len(nonzero_parts) - 1:
                    extent = max(0.0, 360.0 - consumed)
                else:
                    extent = 360.0 * (value / pie_total)
            if extent > 0.0:
                # Draw wedge as polygon to avoid create_arc rendering glitches on some Tk/Windows builds.
                steps = max(6, int(extent / 3.0))
                pts: list[float] = [cx, cy]
                for step in range(steps + 1):
                    ang = start + (extent * step / steps)
                    rad = math.radians(ang)
                    pts.append(cx + rx * math.cos(rad))
                    pts.append(cy - ry * math.sin(rad))
                canvas.create_polygon(pts, fill=color, outline="white", width=1)
                consumed += extent
                drawn_slices += 1
            start += extent
        if total_removed <= 0 or drawn_slices == 0:
            canvas.create_text(
                (x0 + x1) // 2,
                (y0 + y1) // 2,
                text="No filtered-out primers",
                fill="#555",
            )
        legend.insert(
            "end",
            "\nPercentages are relative to filtered-out primers only (100%).\n",
        )
        legend.insert(
            "end",
            "Specificity check categories: 'Not target-specific' means target-rule failure; "
            "'Other/unclassified' are specificity-stage drops not tagged as target or SNP.\n",
        )
        legend.insert("end", f"\nTotal filtered out: {total_removed}\n")
        legend.insert("end", f"Parsed Primer3 rows: {parsed}\n")
        legend.configure(state="disabled")
        frm.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        ttk.Button(frm, text="Close", command=win.destroy).grid(row=1, column=1, sticky="e", pady=(8, 0))
        _center_dialog_on_parent(win, root)
        win.focus_force()

    def _show_runtime_pie() -> None:
        raw = state.get("last_runtime_breakdown")
        if not isinstance(raw, dict):
            _set_status("No runtime breakdown yet. Run primer search first.")
            return
        rb = dict(raw)
        try:
            t_primer3 = max(0.0, float(rb.get("primer3_s", 0.0)))
            t_spidey = max(0.0, float(rb.get("spidey_s", rb.get("spidey_s", 0.0))))
            t_collect = max(0.0, float(rb.get("collect_s", 0.0)))
            t_ntthal_refine = max(0.0, float(rb.get("ntthal_refine_s", 0.0)))
            t_ntthal_cutoff = max(0.0, float(rb.get("ntthal_cutoff_s", 0.0)))
            t_mfe_dimer = max(0.0, float(rb.get("mfe_dimer_s", 0.0)))
            t_mfe_spec = max(0.0, float(rb.get("mfe_spec_s", 0.0)))
            t_sort = max(0.0, float(rb.get("sort_s", 0.0)))
        except Exception:
            _set_status("Runtime breakdown is malformed.")
            return
        profile_name = str(rb.get("binary_profile") or "Unknown").strip()
        primer3_bin = str(rb.get("primer3_bin") or "").strip()
        ntthal_bin = str(rb.get("ntthal_bin") or "").strip()

        parts: list[tuple[str, float, str]] = [
            ("Primer3", t_primer3, "#4e79a7"),
            ("spidey", t_spidey, "#f28e2b"),
            ("Collect candidates", t_collect, "#e15759"),
            ("ntthal refine", t_ntthal_refine, "#76b7b2"),
            ("ntthal cutoff", t_ntthal_cutoff, "#59a14f"),
            ("MFEprimer dimer", t_mfe_dimer, "#edc948"),
            ("MFEprimer Spec", t_mfe_spec, "#b07aa1"),
            ("Sort/render prep", t_sort, "#9c755f"),
        ]
        total_runtime = sum(v for _, v, _ in parts)

        win = tk.Toplevel(root)
        win.title("Runtime Pie")
        win.transient(root)
        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)

        canvas = tk.Canvas(frm, width=420, height=320, bg="white")
        canvas.grid(row=0, column=0, rowspan=2, sticky="nsew")
        legend = tk.Text(frm, width=52, height=18, wrap="word")
        legend.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        legend.configure(state="normal")
        legend.delete("1.0", "end")
        legend.tag_configure("hdr", font=("TkDefaultFont", 10, "bold"))

        x0, y0, x1, y1 = 30, 20, 300, 290
        start = 0.0
        pie_total = sum(max(0.0, float(v)) for _, v, _ in parts)
        nonzero_parts = [(label, value, color) for label, value, color in parts if value > 0.0]
        cx = (x0 + x1) / 2.0
        cy = (y0 + y1) / 2.0
        rx = max(1.0, (x1 - x0) / 2.0)
        ry = max(1.0, (y1 - y0) / 2.0)
        legend.insert("end", "Runtime Pie\n", ("hdr",))
        legend.insert("end", f"Binary profile: {profile_name}\n")
        if primer3_bin:
            legend.insert("end", f"Primer3 binary: {primer3_bin}\n")
        if ntthal_bin:
            legend.insert("end", f"ntthal binary: {ntthal_bin}\n")
        legend.insert("end", "\n")
        for label, value, color in parts:
            pct = 0.0 if total_runtime <= 0.0 else ((100.0 * value) / total_runtime)
            swatch_tag = f"swatch_{color.replace('#', '')}"
            legend.tag_configure(swatch_tag, background=color)
            legend.insert("end", "  ", (swatch_tag,))
            legend.insert("end", " ")
            legend.insert("end", f"{label}: {value:.2f}s ({pct:.1f}%)\n")
        drawn_slices = 0
        consumed = 0.0
        for idx, (_label, value, color) in enumerate(nonzero_parts):
            extent = 0.0
            if pie_total > 0.0:
                if idx == len(nonzero_parts) - 1:
                    extent = max(0.0, 360.0 - consumed)
                else:
                    extent = 360.0 * (value / pie_total)
            if extent > 0.0:
                # Draw wedge as polygon to avoid create_arc rendering glitches on some Tk/Windows builds.
                steps = max(6, int(extent / 3.0))
                pts: list[float] = [cx, cy]
                for step in range(steps + 1):
                    ang = start + (extent * step / steps)
                    rad = math.radians(ang)
                    pts.append(cx + rx * math.cos(rad))
                    pts.append(cy - ry * math.sin(rad))
                canvas.create_polygon(pts, fill=color, outline="white", width=1)
                consumed += extent
                drawn_slices += 1
            start += extent
        if total_runtime <= 0.0 or drawn_slices == 0:
            canvas.create_text(
                (x0 + x1) // 2,
                (y0 + y1) // 2,
                text="No runtime data",
                fill="#555",
            )
        legend.insert("end", "\nPercentages are relative to total runtime (100%).\n")
        legend.insert("end", f"\nTotal runtime: {total_runtime:.2f}s\n")
        legend.configure(state="disabled")
        frm.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        ttk.Button(frm, text="Close", command=win.destroy).grid(row=1, column=1, sticky="e", pady=(8, 0))
        _center_dialog_on_parent(win, root)
        win.focus_force()

    def _status_for_label(msg: str, max_chars: int = 260) -> str:
        s = str(msg or "")
        if len(s) <= max_chars:
            return s
        clipped = s[: max(0, max_chars - 24)].rstrip()
        return f"{clipped} ... [Copy status]"

    def _ensure_run_status_popup(initial_msg: str) -> None:
        if bool(state.get("run_status_popup_active")):
            _update_run_status_popup(initial_msg)
            return
        win = tk.Toplevel(root)
        win.title("Find Primers Status")
        win.transient(root)
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)
        msg_var = tk.StringVar(value=str(initial_msg or "Working ..."))
        ttk.Label(frm, textvariable=msg_var, justify="left", anchor="w", wraplength=420).pack(fill="x")
        win.update_idletasks()
        _center_dialog_on_parent(win, root)

        def _block_close() -> None:
            return

        win.protocol("WM_DELETE_WINDOW", _block_close)
        state["run_status_popup"] = win
        state["run_status_popup_msg_var"] = msg_var
        state["run_status_popup_active"] = True

    def _update_run_status_popup(msg: str) -> None:
        win = state.get("run_status_popup")
        msg_var = state.get("run_status_popup_msg_var")
        if not bool(state.get("run_status_popup_active")):
            return
        if isinstance(win, tk.Toplevel):
            try:
                if not win.winfo_exists():
                    return
            except Exception:
                return
        if isinstance(msg_var, tk.StringVar):
            msg_var.set(str(msg or "Working ..."))
        try:
            root.update_idletasks()
        except Exception:
            pass

    def _close_run_status_popup() -> None:
        win = state.get("run_status_popup")
        state["run_status_popup_active"] = False
        state["run_status_popup"] = None
        state["run_status_popup_msg_var"] = None
        if isinstance(win, tk.Toplevel):
            try:
                if win.winfo_exists():
                    win.destroy()
            except Exception:
                pass

    def _set_run_status(msg: str) -> None:
        _set_status(msg)
        if bool(state.get("run_status_popup_active")):
            _update_run_status_popup(msg)

    def _path_is_file(path: str) -> bool:
        txt = str(path or "").strip()
        return bool(txt) and Path(txt).is_file()

    def _startup_toolchain_checks() -> int:
        warnings: list[str] = []
        primer3_ok = _path_is_file(primer3_var.get())
        spidey_ok = _path_is_file(spidey_var.get())
        mfe_ok = _path_is_file(mfeprimer_var.get())
        state["mfeprimer_available"] = mfe_ok

        if not primer3_ok:
            msg = (
                "Primer3 executable not found. PrimeRL requires primer3_core in tools/bin "
                "or configured on PATH."
            )
            if getattr(sys, "frozen", False):
                messagebox.showerror("PrimeRL startup error", msg)
                return 2
            warnings.append(msg)

        if not mfe_ok:
            run_mfeprimer_var.set(0)
            if bool(spec_top50_var.get()):
                spec_top50_var.set(0)
            warnings.append(
                "MFEprimer executable not found; MFE dimer and transcriptome specificity checks are disabled."
            )

        if not spidey_ok:
            warnings.append(
                "spidey executable not found; automatic boundary detection without Ensembl context will fail."
            )

        _update_spec_toggle_availability()
        if warnings:
            _set_status(" | ".join(warnings))
        return 0

    def _set_status(msg: str) -> None:
        state["full_status_text"] = str(msg or "")
        status_var.set(_status_for_label(str(msg or "")))
        root.update_idletasks()

    def _set_row_checked(iid: str, checked: bool) -> None:
        vals = list(tree.item(iid, "values"))
        if not vals:
            return
        vals[0] = "[x]" if checked else "[ ]"
        tree.item(iid, values=tuple(vals))

    def _checked_ids() -> set[int]:
        raw = state.get("checked")
        if isinstance(raw, set):
            return set(int(x) for x in raw)
        return set()

    def _set_checked_ids(ids: set[int]) -> None:
        state["checked"] = set(ids)

    def _result_sort_value(row: list[object], col_id: str) -> float:
        idx_map = {"amp": 9, "pd": 10, "pd_full": 13}
        idx = idx_map.get(col_id)
        if idx is None:
            return 0.0
        try:
            return float(row[idx])
        except (ValueError, TypeError, IndexError):
            return float("-inf")

    def _apply_results_sort(rows: list[list[object]]) -> list[list[object]]:
        col_id = str(state.get("result_sort_col") or "pd")
        desc = bool(state.get("result_sort_desc"))
        if col_id not in sortable_result_columns:
            return list(rows)
        return sorted(
            rows,
            key=lambda r: _result_sort_value(r, col_id),
            reverse=desc,
        )

    def _update_result_header_labels() -> None:
        sort_col = str(state.get("result_sort_col") or "")
        desc = bool(state.get("result_sort_desc"))
        for cid, title, _width in COLUMNS:
            if cid == sort_col and cid in sortable_result_columns:
                arrow = "(v)" if desc else "(^)"
                tree.heading(cid, text=f"{title} {arrow}")
            else:
                tree.heading(cid, text=title)

    def _on_result_sort_header(col_id: str) -> None:
        if col_id not in sortable_result_columns:
            return
        cur_col = str(state.get("result_sort_col") or "")
        cur_desc = bool(state.get("result_sort_desc"))
        if col_id == cur_col:
            state["result_sort_desc"] = not cur_desc
        else:
            state["result_sort_col"] = col_id
            state["result_sort_desc"] = True
        _update_result_header_labels()
        current_rows = list(state.get("rows") or [])
        _render_rows(current_rows)

    def _render_rows(rows: list[list[object]]) -> None:
        rows = _apply_results_sort(rows)
        for item in tree.get_children(""):
            tree.delete(item)
        spec_passed_raw = state.get("spec_passed")
        spec_passed = set(int(x) for x in spec_passed_raw) if isinstance(spec_passed_raw, set) else set()
        spec_pass_keys_raw = state.get("spec_pass_keys")
        spec_pass_keys = set(str(x) for x in spec_pass_keys_raw) if isinstance(spec_pass_keys_raw, set) else set()
        for idx, row in enumerate(rows):
            row_key = _row_spec_key(row)
            is_pass = (row_key in spec_pass_keys) if spec_pass_keys else (idx in spec_passed)
            tags = ("spec_pass",) if is_pass else ()
            tree.insert(
                "",
                "end",
                iid=str(idx),
                values=_pairs_to_values(row, checked=False, spec_pass=is_pass),
                tags=tags,
            )

        state["rows"] = rows
        state["checked"] = set()
        state["selected_row"] = 0 if rows else None
        if rows:
            tree.selection_set("0")
            tree.focus("0")
            tree.see("0")
        _update_result_header_labels()
        _safe_draw_cdna_map()

    def _build_payload(
        *,
        rows: list[list[object]],
        final_returned: int,
        stats_dict: dict[str, object],
        spidey_meta: object,
    ) -> dict[str, object]:
        return {
            "stats": {
                "parsed": int(stats_dict.get("parsed", 0)),
                "skipped_repeat_run": int(stats_dict.get("skipped_repeat_run", 0)),
                "skipped_order": int(stats_dict.get("skipped_order", 0)),
                "skipped_span": int(stats_dict.get("skipped_span", 0)),
                "skipped_overlap": int(stats_dict.get("skipped_overlap", 0)),
            },
            "returned_pairs": final_returned,
            "spidey": spidey_meta,
            "spidey": spidey_meta,
            "pairs": rows,
        }

    def _toggle_checked_click(evt: tk.Event[tk.Misc]) -> str | None:
        row_id = tree.identify_row(evt.y)
        col_id = tree.identify_column(evt.x)
        if not row_id or col_id != "#1":
            return None
        try:
            idx = int(row_id)
        except ValueError:
            return None
        checked = _checked_ids()
        if idx in checked:
            checked.remove(idx)
            _set_row_checked(row_id, False)
        else:
            checked.add(idx)
            _set_row_checked(row_id, True)
        _set_checked_ids(checked)
        return "break"

    def _run_spec_on_checked() -> None:
        try:
            sel = tree.selection()
            if len(sel) != 1:
                _set_status("Select exactly one primer pair row to open Primer-BLAST.")
                return
            all_rows = list(state.get("rows") or [])
            try:
                idx = int(sel[0])
            except ValueError:
                _set_status("Invalid selected row.")
                return
            if idx < 0 or idx >= len(all_rows):
                _set_status("Selected row is not available.")
                return

            row = all_rows[idx]
            fseq = str(row[0]).strip()
            rseq = str(row[4]).strip()
            template = _seq_from_text(mrna_text.get("1.0", "end"))
            if not template:
                _set_status("mRNA sequence is empty; cannot prepare Primer-BLAST template.")
                return
            ensembl_species = str(state.get("ensembl_species") or "Homo_sapiens")
            organism = ensembl_species.replace("_", " ").strip()

            base = "https://www.ncbi.nlm.nih.gov/tools/primer-blast/index.cgi"
            # Best-effort prefill keys; Primer-BLAST may ignore unsupported URL params.
            # URL length limits vary by browser/proxy/server. For very long cDNA, avoid
            # putting the template into the URL to prevent 414 Request-URI Too Long.
            max_cdna_in_url = 2500
            if len(template) <= max_cdna_in_url:
                q = urlencode(
                    {
                        "INPUT_SEQUENCE": template,
                        "PRIMER_LEFT_INPUT": fseq,
                        "PRIMER_RIGHT_INPUT": rseq,
                        "ORGANISM": organism,
                    }
                )
                url = f"{base}?{q}"
                used_template_prefill = True
            else:
                q = urlencode(
                    {
                        "PRIMER_LEFT_INPUT": fseq,
                        "PRIMER_RIGHT_INPUT": rseq,
                        "ORGANISM": organism,
                    }
                )
                url = f"{base}?{q}"
                used_template_prefill = False

            payload = (
                f"Primer-BLAST prefill payload\n"
                f"Organism: {organism}\n"
                f"Forward primer: {fseq}\n"
                f"Reverse primer: {rseq}\n"
                f"PCR template (cDNA):\n{template}\n"
            )
            root.clipboard_clear()
            root.clipboard_append(payload)

            webbrowser.open(url, new=2)
            if used_template_prefill:
                _set_status(
                    "Opened Primer-BLAST with template prefill. Payload copied to clipboard."
                )
            else:
                _set_status(
                    "Opened Primer-BLAST without template URL prefill (long cDNA). Payload copied to clipboard."
                )

        except Exception as exc:
            messagebox.showerror("Open Primer-BLAST", str(exc))
            _set_status(f"Open Primer-BLAST error: {exc}")

    def _on_result_select(_evt: object | None = None) -> None:
        sel = tree.selection()
        if not sel:
            state["selected_row"] = None
        else:
            try:
                state["selected_row"] = int(sel[0])
            except ValueError:
                state["selected_row"] = None
        _safe_draw_cdna_map()

    def _rebuild_cached_candidates(show_dialog: bool = False) -> bool:
        try:
            t_start = time.perf_counter()
            parsed = state.get("parsed_primer3")
            if not isinstance(parsed, dict):
                _set_status("No cached run yet. Click Find primers first.")
                return False

            primer3_path = _ensure_tool_path(primer3_var, "primer3_core", "Primer3")

            bounds = [int(b) for b in (state.get("bounds") or [])]
            template_len = int(state.get("template_len") or 0)
            if template_len <= 0:
                raise ValueError("Cached template is missing. Click Find primers first.")

            _set_status("Applying filter options on cached Primer3 output ...")
            t_collect_0 = time.perf_counter()
            filt = QpcrFilterSettings(
                exclude_rr_q=bool(exclude_rr_var.get()),
                run=int(run_var.get()),
                repeat=int(repeat_var.get()),
                ie_span=bool(ie_span_var.get()),
                ie_overlap=bool(ie_overlap_var.get()),
                exclude_ie=int(exclude_ie_var.get()),
                intron_exon_bounds=tuple(bounds),
            )
            pairs, stats = collect_qpcr_pairs_from_primer3(
                parsed=parsed,
                template_len=template_len,
                settings=filt,
                primer_dimer_fn=_pd_stub,
            )

            rows = [p.to_legacy_row() for p in pairs]
            exon_limit_removed = 0
            if ie_limit_var.get() and template_len > 0:
                lo, hi, ex5, ex3, _ = _selected_exon_window(
                    bounds=bounds,
                    template_len=template_len,
                    ex5_txt=ie_5p_var.get(),
                    ex3_txt=ie_3p_var.get(),
                )
                rows, exon_limit_removed = _filter_rows_to_cdna_window(rows, lo=lo, hi=hi)
            after_filter_options = len(rows)
            t_collect = time.perf_counter() - t_collect_0
            _set_status("Refining thermodynamics with ntthal ...")
            t_nt_0 = time.perf_counter()
            rows = _refine_rows_with_ntthal(rows, primer3_path=primer3_path, ntthal_path=ntthal_var.get())
            t_nt = time.perf_counter() - t_nt_0
            t_total = time.perf_counter() - t_start
            state["rows_after_ntthal"] = [list(r) for r in rows]
            state["design_stats"] = {
                "parsed": stats.parsed,
                "skipped_repeat_run": stats.skipped_repeat_run,
                "skipped_order": stats.skipped_order,
                "skipped_span": stats.skipped_span,
                "skipped_overlap": stats.skipped_overlap,
                "skipped_exon_limit": exon_limit_removed,
                "after_filter_options": after_filter_options,
                "timing_collect_s": t_collect,
                "timing_ntthal_refine_s": t_nt,
                "timing_cached_total_s": t_total,
            }
            return True
        except Exception as exc:
            if show_dialog:
                messagebox.showerror("Apply filters failed", str(exc))
            _set_status(f"Error: {exc}")
            return False

    def _set_busy_controls(enabled: bool) -> None:
        state_txt = "normal" if enabled else "disabled"
        for w in busy_controls:
            try:
                w.configure(state=state_txt)
            except Exception:
                continue
        try:
            if enabled:
                busy_progress.stop()
            else:
                busy_progress.start(10)
        except Exception:
            pass

    def _purge_postfilter_snapshots() -> None:
        try:
            logs_dir = Path(_primerl_data_path("runtime", "logs"))
            if not logs_dir.exists():
                return
            for snap in logs_dir.glob("postfilter_snapshot_*.json"):
                snap.unlink(missing_ok=True)
            (logs_dir / "last_postfilter_snapshot.json").unlink(missing_ok=True)
        except Exception:
            pass


    def _write_run_snapshot(
        *,
        rows: list[list[object]],
        payload: dict[str, object],
        status_msg: str,
    ) -> None:
        try:
            logs_dir = Path(_primerl_data_path("runtime", "logs"))
            logs_dir.mkdir(parents=True, exist_ok=True)
            ts = time.strftime("%Y%m%d_%H%M%S")
            snapshot = {
                "timestamp_local": time.strftime("%Y-%m-%d %H:%M:%S"),
                "status": str(status_msg or ""),
                "target_gene_id": _active_target_gene_id(),
                "target_gene_symbol": str(state.get("ensembl_gene") or ""),
                "ensembl_species": str(state.get("ensembl_species") or ""),
                "transcriptome_db": str(mfeprimer_transcriptome_fasta_var.get() or ""),
                "mfeprimer_path": str(mfeprimer_var.get() or ""),
                "spec_top50_enabled": bool(spec_top50_var.get()),
                "max_pairs": int(max_pairs_var.get()),
                "min_amp": int(min_amp_var.get()),
                "max_amp": int(max_amp_var.get()),
                "mfe_dg_cutoff": float(mfe_dg_cutoff_var.get()),
                "ntthal_ext_cutoff": float(ntthal_ext_cutoff_var.get()),
                "snp_check_enabled": bool(enable_snp_check_var.get()),
                "snp_non3p_policy": str(snp_non3p_policy_var.get() or ""),
                "snp_3p_window": int(snp_3p_window_var.get()),
                "auto_spec_applied": bool(payload.get("auto_spec_applied", False)),
                "filter_counts": dict(payload.get("filter_counts", {})) if isinstance(payload.get("filter_counts"), dict) else {},
                "stats": dict(payload.get("stats", {})) if isinstance(payload.get("stats"), dict) else {},
                "final_returned": int(payload.get("final_returned", len(rows))),
            }
            latest_path = logs_dir / "last_postfilter_snapshot.json"
            txt = json.dumps(snapshot, ensure_ascii=True, indent=2)
            latest_path.write_text(txt, encoding="utf-8")
        except Exception:
            # Snapshotting is best-effort and must not break the GUI workflow.
            pass

    def _finish_post_filter_worker(
        *,
        job_id: int,
        result: dict[str, object] | None,
        err: Exception | None,
        show_dialog: bool,
    ) -> bool:
        if int(state.get("post_filter_job_id") or 0) != job_id:
            return False

        state["post_filter_busy"] = False
        state["design_busy"] = False
        _set_busy_controls(True)

        if err is not None:
            _close_run_status_popup()
            if show_dialog:
                messagebox.showerror("Apply filters failed", str(err))
            _set_status(f"Error: {err}")
            return False

        payload = result if isinstance(result, dict) else {}
        rows = [list(r) for r in payload.get("rows", [])]
        final_returned = int(payload.get("final_returned", len(rows)))
        stats_dict = dict(payload.get("stats", {}))
        spidey_meta = payload.get("spidey", {"used": False, "boundaries": []})
        filter_counts = payload.get("filter_counts")
        if isinstance(filter_counts, dict):
            state["last_filter_counts"] = dict(filter_counts)
        runtime_breakdown = payload.get("runtime_breakdown")
        if isinstance(runtime_breakdown, dict):
            state["last_runtime_breakdown"] = dict(runtime_breakdown)
        msg = str(payload.get("status", "Done."))
        auto_spec_applied = bool(payload.get("auto_spec_applied", False))
        spec_pass_keys_raw = payload.get("spec_pass_keys")
        if isinstance(spec_pass_keys_raw, list):
            state["spec_pass_keys"] = set(str(x) for x in spec_pass_keys_raw if isinstance(x, str))
        else:
            state["spec_pass_keys"] = set()
        state["spec_passed"] = set()
        _render_rows(rows)
        state["payload"] = _build_payload(
            rows=rows,
            final_returned=final_returned,
            stats_dict=stats_dict,
            spidey_meta=spidey_meta,
        )
        _write_run_snapshot(rows=rows, payload=payload, status_msg=msg)

        _close_run_status_popup()
        _set_status(msg)
        if show_dialog and bool(payload.get("spec_no_target_specificity", False)):
            tested_n = max(0, int(payload.get("spec_tested_count", 0)))
            messagebox.showwarning("Specificity check", "No Pair passed the specificity test! Close paralog? Try focusing on UTRs")
        return True

    def _apply_post_filters_from_cached(show_dialog: bool = False) -> bool:
        if bool(state.get("post_filter_busy")):
            _set_status("A post-filter run is already in progress.")
            return False

        try:
            cached = state.get("rows_after_ntthal")
            if not isinstance(cached, list):
                _set_status("No cached run yet. Click Find primers first.")
                return False

            _purge_postfilter_snapshots()
            rows = [list(r) for r in cached]
            after_ntthal = len(rows)
            run_ntthal_cutoff = bool(run_ntthal_cutoff_var.get())
            ntthal_ext_cutoff = float(ntthal_ext_cutoff_var.get())
            run_mfe = bool(run_mfeprimer_var.get())
            mfe_path = mfeprimer_var.get()
            mfe_cutoff = float(mfe_dg_cutoff_var.get())
            run_spec = bool(spec_top50_var.get())
            spec_guard_note = ""
            run_snp_check = bool(enable_snp_check_var.get()) and (not _is_human_target_context())
            target_gene_id = _active_target_gene_id()
            target_gene_symbol = str(state.get("ensembl_gene") or "").strip()
            auto_db_msg = _auto_select_transcriptome_db_for_target(
                gene_id=target_gene_id,
                species_hint=str(state.get("ensembl_species") or ""),
            )
            spec_db = mfeprimer_transcriptome_fasta_var.get()
            if run_spec and not target_gene_id and not target_gene_symbol:
                run_spec = False
                spec_top50_var.set(0)
                spec_guard_note = (
                    "Specificity testing was disabled because target gene information is missing. "
                    "Retrieve from Ensembl or enter Target Ensembl gene ID to enable it."
                )
            if run_spec and target_gene_id:
                expected_slug = _infer_species_slug_from_gene_id(target_gene_id)
                db_name = Path(str(spec_db or "")).name.lower()
                if expected_slug:
                    slug_dot = expected_slug.replace("_", ".")
                    slug_compact = expected_slug.replace("_", "")
                    species_match = (
                        expected_slug in db_name
                        or slug_dot in db_name
                        or slug_compact in db_name
                    )
                    if not species_match:
                        run_spec = False
                        spec_top50_var.set(0)
                        spec_guard_note = (
                            "Specificity testing was disabled because transcriptome DB species mismatched the "
                            f"target ({expected_slug} vs {Path(str(spec_db or '')).name})."
                        )
            snp_bed = str(state.get("current_snp_bed") or "").strip() if run_snp_check else ""
            snp_non3p_policy = str(snp_non3p_policy_var.get() or "soft").strip().lower()
            if snp_non3p_policy not in {"soft", "hard"}:
                snp_non3p_policy = "soft"
            snp_3p_window = int(snp_3p_window_var.get())
            if snp_3p_window < 1:
                raise ValueError("3' SNP window must be >= 1.")
            target_snp_cdna_positions = [int(p) for p in (state.get("snp_cdna_positions") or [])] if run_snp_check else []
            spec_selection_mode = str(spec_selection_mode_var.get() or "strict_pass").strip().lower()
            if spec_selection_mode not in {"strict_pass", "score_top_pct"}:
                spec_selection_mode = "strict_pass"
            spec_remove_pct = float(spec_remove_pct_var.get())
            if spec_remove_pct < 0 or spec_remove_pct > 100:
                raise ValueError("Spec remove % must be between 0 and 100.")
            spec_param_warning = ""

            def _capture_spec_param_warning(msg: str) -> None:
                nonlocal spec_param_warning
                spec_param_warning = msg

            spec_extra_args = resolve_spec_param_tokens(
                str(mfeprimer_spec_params_var.get() or ""),
                on_error=_capture_spec_param_warning,
            )
            if spec_param_warning and show_dialog:
                messagebox.showwarning("Specificity check parameters", spec_param_warning)
            spec_min_amp = int(SPEC_OFFTARGET_MIN_AMP_SIZE_BP)
            spec_max_amp = int(SPEC_OFFTARGET_MAX_AMP_SIZE_BP)
            max_pairs = int(max_pairs_var.get())
            stats_raw = state.get("design_stats")
            stats_dict = dict(stats_raw) if isinstance(stats_raw, dict) else {}
            spidey_meta = state.get("spidey_meta") or {"used": False, "boundaries": []}
            primer3_path_for_run = str(primer3_var.get() or "").strip()
            ntthal_path_for_run = str(ntthal_var.get() or "").strip()
            binary_profile_for_run = _infer_binary_profile_from_paths(primer3_path_for_run, ntthal_path_for_run)
        except Exception as exc:
            if show_dialog:
                messagebox.showerror("Apply filters failed", str(exc))
            _set_status(f"Error: {exc}")
            return False

        job_id = int(state.get("post_filter_job_id") or 0) + 1
        state["post_filter_job_id"] = job_id
        state["post_filter_busy"] = True
        _set_busy_controls(False)
        _set_status("Applying post-filters in background ...")

        def _set_status_async(msg: str) -> None:
            root.after(
                0,
                lambda m=msg, j=job_id: (
                    _set_run_status(m) if int(state.get("post_filter_job_id") or 0) == j else None
                ),
            )

        def _worker() -> None:
            try:
                t_post_0 = time.perf_counter()
                work_rows = [list(r) for r in rows]

                ntthal_note = ""
                t_nt_cutoff = 0.0
                if run_ntthal_cutoff:
                    _set_status_async("Filtering in process (apply cutoffs).")
                    t0 = time.perf_counter()
                    work_rows, ntthal_note = _filter_rows_by_ntthal_ext_cutoff(
                        work_rows,
                        ext_dg_cutoff=ntthal_ext_cutoff,
                    )
                    t_nt_cutoff = time.perf_counter() - t0
                after_ntthal_cutoff = len(work_rows)

                mfe_note = ""
                t_mfe_dimer = 0.0
                if run_mfe:
                    _set_status_async("Filtering in process (apply cutoffs).")
                    t0 = time.perf_counter()
                    work_rows, mfe_note = _filter_rows_with_mfeprimer(
                        work_rows,
                        mfeprimer_path=mfe_path,
                        dg_cutoff=mfe_cutoff,
                    )
                    t_mfe_dimer = time.perf_counter() - t0
                after_mfeprimer = len(work_rows)

                mfe_spec_note = ""
                t_mfe_spec = 0.0
                spec_metrics: dict[str, int] = {"target_removed": 0, "snp_removed": 0}
                spec_pass_keys: set[str] = set()
                spec_tested_count = 0
                spec_no_target_specificity = False
                if run_snp_check:
                    _set_status_async("Filtering in process (apply cutoffs).")
                    pre_snp_len = len(work_rows)
                    work_rows, snp_removed_local, snp_pairs_3p, snp_pairs_non3p = _filter_rows_by_target_transcript_snps(
                        work_rows,
                        snp_cdna_positions=target_snp_cdna_positions,
                        window3=snp_3p_window,
                        non3p_hard=(snp_non3p_policy == "hard"),
                    )
                    if snp_removed_local > 0:
                        spec_metrics["snp_removed"] = int(spec_metrics.get("snp_removed", 0)) + int(snp_removed_local)
                        mfe_spec_note += (
                            f" Target-transcript SNP filter removed {snp_removed_local} pair(s) "
                            f"(3' hits={snp_pairs_3p}, non3' hits={snp_pairs_non3p}, "
                            f"window3p={snp_3p_window}, policy_non3p={snp_non3p_policy})."
                        )
                    elif pre_snp_len > 0 and target_snp_cdna_positions:
                        mfe_spec_note += (
                            f" Target-transcript SNP filter removed 0 pair(s) "
                            f"(mapped_snps={len(target_snp_cdna_positions)}, "
                            f"window3p={snp_3p_window}, policy_non3p={snp_non3p_policy})."
                        )
                if run_spec:
                    spec_cap = 50
                    ranked_rows = _sort_rows_by_ext_dimer_desc(work_rows)
                    spec_rows = ranked_rows[:spec_cap]
                    untested_rows = ranked_rows[spec_cap:]
                    spec_tested_count = len(spec_rows)
                    _set_status_async(
                        f"Specificity check against transcriptome database (top {len(spec_rows)}) ..."
                    )
                    t0 = time.perf_counter()
                    spec_rows, mfe_spec_note = _filter_rows_with_mfeprimer_spec(
                        spec_rows,
                        mfeprimer_path=mfe_path,
                        db_fasta_path=spec_db,
                        target_gene_id=target_gene_id,
                        target_gene_symbol=target_gene_symbol,
                        snp_check_enabled=run_snp_check,
                        snp_bed_path=snp_bed,
                        snp_non3p_policy=snp_non3p_policy,
                        snp_3p_window=snp_3p_window,
                        max_amplicons=1,
                        min_amp_size=spec_min_amp,
                        max_amp_size=spec_max_amp,
                        spec_selection_mode=spec_selection_mode,
                        spec_remove_pct=spec_remove_pct,
                        metrics_out=spec_metrics,
                        spec_extra_args=spec_extra_args,
                    )
                    t_mfe_spec = time.perf_counter() - t0
                    if spec_param_warning:
                        mfe_spec_note = (mfe_spec_note + " " + spec_param_warning).strip()
                    spec_pass_keys = set(_row_spec_key(r) for r in spec_rows)
                    if spec_tested_count > 0 and len(spec_rows) == 0 and "target-gene-ID-only rule" in mfe_spec_note:
                        spec_no_target_specificity = True
                    if len(ranked_rows) > spec_cap:
                        mfe_spec_note += (
                            f" Spec run limited to top {spec_cap} pairs by Ext. dimer dG; "
                            f"{len(untested_rows)} additional pair(s) were not spec-tested and were kept unmarked."
                        )
                    work_rows = spec_rows + untested_rows
                elif run_snp_check:
                    # User enabled SNP-only filtering while spec is off.
                    if not target_snp_cdna_positions:
                        mfe_spec_note += " SNP checking enabled, but no mapped SNPs were available."
                after_mfeprimer_spec = len(work_rows)

                _set_status_async("Sorting and rendering primer table ...")
                t_sort_0 = time.perf_counter()
                work_rows = _sort_rows_by_ext_dimer_desc(work_rows)
                if max_pairs >= 0:
                    work_rows = work_rows[:max_pairs]
                t_sort = time.perf_counter() - t_sort_0
                final_returned = len(work_rows)
                t_post = time.perf_counter() - t_post_0

                msg = (
                    f"Done. Parsed={int(stats_dict.get('parsed', 0))}, "
                    f"after filter options={int(stats_dict.get('after_filter_options', after_ntthal))}, "
                    f"after ntthal QC={after_ntthal}, after ntthal cutoff={after_ntthal_cutoff}, "
                    f"after MFEprimer QC={after_mfeprimer}, "
                    f"after MFEprimer transcriptome spec={after_mfeprimer_spec}, "
                    f"returned={final_returned}, "
                    f"skipped repeat/run={int(stats_dict.get('skipped_repeat_run', 0))}, "
                    f"skipped span={int(stats_dict.get('skipped_span', 0))}, "
                    f"overlap={int(stats_dict.get('skipped_overlap', 0))}"
                )
                t_pre = (
                    f"timing_pre_s(primer3={float(stats_dict.get('timing_primer3_s', 0.0)):.2f}, "
                    f"spidey={float(stats_dict.get('timing_spidey_s', 0.0)):.2f}, "
                    f"collect={float(stats_dict.get('timing_collect_s', 0.0)):.2f}, "
                    f"ntthal_refine={float(stats_dict.get('timing_ntthal_refine_s', 0.0)):.2f})"
                )
                t_post_txt = (
                    f"timing_post_s(ntthal_cutoff={t_nt_cutoff:.2f}, mfe_dimer={t_mfe_dimer:.2f}, "
                    f"mfe_spec={t_mfe_spec:.2f}, sort={t_sort:.2f}, total_post={t_post:.2f})"
                )
                msg += f" | {t_pre} | {t_post_txt}"
                if auto_db_msg:
                    msg += f" | {auto_db_msg}"
                if spec_guard_note:
                    msg += f" | {spec_guard_note}"
                if ntthal_note:
                    msg += f" | {ntthal_note}"
                if mfe_note:
                    msg += f" | {mfe_note}"
                if mfe_spec_note:
                    msg += f" | {mfe_spec_note}"

                root.after(
                    0,
                    lambda: _finish_post_filter_worker(
                        job_id=job_id,
                        result={
                            "rows": work_rows,
                            "final_returned": final_returned,
                            "auto_spec_applied": bool(run_spec),
                            "stats": stats_dict,
                            "spidey": spidey_meta,
                            "filter_counts": {
                                "parsed": int(stats_dict.get("parsed", 0)),
                                "skipped_repeat_run": int(stats_dict.get("skipped_repeat_run", 0)),
                                "skipped_order": int(stats_dict.get("skipped_order", 0)),
                                "skipped_span": int(stats_dict.get("skipped_span", 0)),
                                "skipped_overlap": int(stats_dict.get("skipped_overlap", 0)),
                                "after_filter_options": int(stats_dict.get("after_filter_options", after_ntthal)),
                                "after_ntthal": after_ntthal,
                                "after_ntthal_cutoff": after_ntthal_cutoff,
                                "after_mfeprimer": after_mfeprimer,
                                "after_mfeprimer_spec": after_mfeprimer_spec,
                                "spec_target_removed": int(spec_metrics.get("target_removed", 0)),
                                "spec_snp_removed": int(spec_metrics.get("snp_removed", 0)),
                                "final_returned": final_returned,
                            },
                            "runtime_breakdown": {
                                "primer3_s": float(stats_dict.get("timing_primer3_s", 0.0)),
                                "spidey_s": float(stats_dict.get("timing_spidey_s", 0.0)),
                                "spidey_s": float(stats_dict.get("timing_spidey_s", 0.0)),
                                "collect_s": float(stats_dict.get("timing_collect_s", 0.0)),
                                "ntthal_refine_s": float(stats_dict.get("timing_ntthal_refine_s", 0.0)),
                                "ntthal_cutoff_s": t_nt_cutoff,
                                "mfe_dimer_s": t_mfe_dimer,
                                "mfe_spec_s": t_mfe_spec,
                                "sort_s": t_sort,
                                "binary_profile": binary_profile_for_run,
                                "primer3_bin": Path(primer3_path_for_run).name if primer3_path_for_run else "",
                                "ntthal_bin": Path(ntthal_path_for_run).name if ntthal_path_for_run else "",
                            },
                            "status": msg,
                            "spec_pass_keys": sorted(spec_pass_keys),
                            "spec_tested_count": int(spec_tested_count),
                            "spec_no_target_specificity": bool(spec_no_target_specificity),
                        },
                        err=None,
                        show_dialog=show_dialog,
                    ),
                )
            except Exception as exc:
                root.after(
                    0,
                    lambda: _finish_post_filter_worker(
                        job_id=job_id,
                        result=None,
                        err=exc,
                        show_dialog=show_dialog,
                    ),
                )

        threading.Thread(target=_worker, name="post-filter-worker", daemon=True).start()
        return True

    def _apply_filter_settings() -> None:
        if not _rebuild_cached_candidates(show_dialog=True):
            return
        _apply_post_filters_from_cached(show_dialog=True)

    def _run_design() -> None:
        if bool(state.get("design_busy")) or bool(state.get("post_filter_busy")):
            _set_status("A primer run is already in progress.")
            return
        try:
            _ensure_run_status_popup("Preparing input sequences ...")
            _set_run_status("Preparing input sequences ...")
            template = _seq_from_text(mrna_text.get("1.0", "end"))
            if not template:
                raise ValueError("mRNA sequence is empty.")
            genomic_full = str(state.get("genomic_seq_full") or "")
            genomic = clean_sequence(genomic_full) if genomic_full else _seq_from_text(gen_text.get("1.0", "end"))
            if not genomic:
                raise ValueError("Genomic sequence is empty.")

            primer3_path = _ensure_tool_path(primer3_var, "primer3_core", "Primer3")
            sp_path = spidey_var.get().strip()
            ensembl_bounds = sorted(
                set(
                    int(b)
                    for b in (state.get("bounds") or [])
                    if int(b) > 0
                )
            )
            has_ensembl_context = bool(str(state.get("ensembl_gene_id") or "").strip())
            use_ensembl_bounds = bool(ensembl_bounds) and has_ensembl_context
            if not use_ensembl_bounds:
                sp_path = _ensure_tool_path(spidey_var, "spidey", "spidey")
            ntthal_path = ntthal_var.get()
            spidey_large = bool(spidey_large_var.get())
            ie_limit = bool(ie_limit_var.get())
            ex5_txt = ie_5p_var.get()
            ex3_txt = ie_3p_var.get()

            run_settings = Primer3RunSettings(
                min_tm_q=float(min_tm_var.get()),
                max_tm_q=float(max_tm_var.get()),
                max_diff_q=float(max_diff_var.get()),
                pri_win_min_q=int(min_len_var.get()),
                pri_win_max_q=int(max_len_var.get()),
                min_ampsize_q=int(min_amp_var.get()),
                max_ampsize_q=int(max_amp_var.get()),
                exclude_gc=bool(exclude_gc_var.get()),
                exclude_clamp=bool(gc_clamp_var.get()),
                num_return=12000,
            )
            filt = QpcrFilterSettings(
                exclude_rr_q=bool(exclude_rr_var.get()),
                run=int(run_var.get()),
                repeat=int(repeat_var.get()),
                ie_span=bool(ie_span_var.get()),
                ie_overlap=bool(ie_overlap_var.get()),
                exclude_ie=int(exclude_ie_var.get()),
                intron_exon_bounds=(),
            )
        except Exception as exc:
            _close_run_status_popup()
            messagebox.showerror("Run failed", str(exc))
            _set_status(f"Error: {exc}")
            return

        state["design_busy"] = True
        _set_busy_controls(False)

        def _set_run_status_async(msg: str) -> None:
            root.after(0, lambda m=msg: _set_run_status(m))

        def _finish_design_worker(result: dict[str, object] | None, err: Exception | None) -> None:
            if err is not None:
                state["design_busy"] = False
                _set_busy_controls(True)
                _close_run_status_popup()
                messagebox.showerror("Run failed", str(err))
                _set_status(f"Error: {err}")
                return

            payload = result if isinstance(result, dict) else {}
            rows = [list(r) for r in payload.get("rows", [])]
            parsed = dict(payload.get("parsed", {}))
            stats = dict(payload.get("stats", {}))
            spidey_meta = payload.get("spidey_meta", {"used": False, "boundaries": []})
            bounds = [int(b) for b in payload.get("bounds", [])]
            sp_txt = str(payload.get("spidey_output", ""))
            template_local = str(payload.get("template", ""))

            state["rows_after_ntthal"] = rows
            state["parsed_primer3"] = parsed
            state["design_stats"] = stats
            state["spidey_meta"] = spidey_meta
            state["spidey_output"] = sp_txt
            state["template_len"] = len(template_local)
            state["bounds"] = bounds
            state["orf"] = _find_longest_orf(template_local)
            if not _apply_post_filters_from_cached(show_dialog=True):
                state["design_busy"] = False
                _set_busy_controls(True)
                _close_run_status_popup()

        def _worker() -> None:
            try:
                t_run_0 = time.perf_counter()
                _set_run_status_async("Primer calculation in progress ...")
                t_p3_0 = time.perf_counter()
                ok, p3_out, err = run_primer3_qpcr_output(
                    template_seq=template,
                    primer3_path=primer3_path,
                    settings=run_settings,
                )
                t_p3 = time.perf_counter() - t_p3_0
                if not ok:
                    raise RuntimeError(err or "Primer3 run failed")

                t_sp = 0.0
                if use_ensembl_bounds:
                    _set_run_status_async("Using Ensembl exon boundaries ...")
                    sp_txt = ""
                    bounds = list(ensembl_bounds)
                    spidey_meta = {
                        "used": False,
                        "source": "ensembl_retrieve",
                        "boundaries": bounds,
                    }
                else:
                    _set_run_status_async("Running spidey alignment ...")
                    t_sp_0 = time.perf_counter()
                    ok_sp, sp_txt, sp_err = _run_spidey_alignment(
                        spidey_path=sp_path,
                        genomic_seq=genomic,
                        mrna_seq=template,
                        print_alignment=1,
                        large_intron=spidey_large,
                    )
                    if not ok_sp:
                        raise RuntimeError(sp_err or "spidey run failed")
                    st = analyze_spidey_output(sp_txt)
                    bounds = extract_intron_exon_bounds(sp_txt)
                    spidey_meta = {
                        "used": True,
                        "spidey_signature": st.has_signature,
                        "full_identity_100": st.full_identity,
                        "full_coverage_100": st.full_coverage,
                        "boundaries": bounds,
                    }
                    t_sp = time.perf_counter() - t_sp_0

                _set_run_status_async("Filtering in progress (rule-based options) ...")
                t_collect_0 = time.perf_counter()
                parsed = parse_primer3_kv_output(p3_out)
                filt_local = QpcrFilterSettings(
                    exclude_rr_q=filt.exclude_rr_q,
                    run=filt.run,
                    repeat=filt.repeat,
                    ie_span=filt.ie_span,
                    ie_overlap=filt.ie_overlap,
                    exclude_ie=filt.exclude_ie,
                    intron_exon_bounds=tuple(bounds),
                )
                pairs, stats_obj = collect_qpcr_pairs_from_primer3(
                    parsed=parsed,
                    template_len=len(template),
                    settings=filt_local,
                    primer_dimer_fn=_pd_stub,
                )
                rows = [p.to_legacy_row() for p in pairs]
                exon_limit_removed = 0
                if ie_limit and len(template) > 0:
                    lo, hi, _ex5, _ex3, _ = _selected_exon_window(
                        bounds=bounds,
                        template_len=len(template),
                        ex5_txt=ex5_txt,
                        ex3_txt=ex3_txt,
                    )
                    rows, exon_limit_removed = _filter_rows_to_cdna_window(rows, lo=lo, hi=hi)
                after_filter_options = len(rows)
                t_collect = time.perf_counter() - t_collect_0

                _set_run_status_async("Filtering in progress (thermodynamics) ...")
                t_nt_0 = time.perf_counter()
                rows = _refine_rows_with_ntthal(rows, primer3_path=primer3_path, ntthal_path=ntthal_path)
                t_nt = time.perf_counter() - t_nt_0
                t_total = time.perf_counter() - t_run_0

                stats_dict = {
                    "parsed": stats_obj.parsed,
                    "skipped_repeat_run": stats_obj.skipped_repeat_run,
                    "skipped_order": stats_obj.skipped_order,
                    "skipped_span": stats_obj.skipped_span,
                    "skipped_overlap": stats_obj.skipped_overlap,
                    "skipped_exon_limit": exon_limit_removed,
                    "after_filter_options": after_filter_options,
                    "timing_primer3_s": t_p3,
                    "timing_spidey_s": t_sp,
                    "timing_collect_s": t_collect,
                    "timing_ntthal_refine_s": t_nt,
                    "timing_design_total_s": t_total,
                }

                root.after(
                    0,
                    lambda: _finish_design_worker(
                        {
                            "rows": rows,
                            "parsed": parsed,
                            "stats": stats_dict,
                            "spidey_meta": spidey_meta,
                            "bounds": list(bounds),
                            "spidey_output": sp_txt,
                            "template": template,
                        },
                        None,
                    ),
                )
            except Exception as exc:
                root.after(0, lambda: _finish_design_worker(None, exc))

        threading.Thread(target=_worker, name="design-worker", daemon=True).start()

    def _retrieve_ensembl() -> None:
        dialog = tk.Toplevel(root)
        dialog.title("Retrieve gene from Ensembl")
        dialog.transient(root)

        gene_var = tk.StringVar(value=str(state.get("ensembl_gene") or ""))
        species_var = tk.StringVar(value=str(state.get("ensembl_species") or "Homo_sapiens"))
        etype_var = tk.StringVar(value=str(state.get("ensembl_type") or "cdna"))

        frm = ttk.Frame(dialog, padding=10)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Gene name").grid(row=0, column=0, sticky="w", pady=3)
        ttk.Entry(frm, textvariable=gene_var, width=25).grid(row=0, column=1, sticky="w", pady=3)
        ttk.Label(frm, text="Organism").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Combobox(frm, textvariable=species_var, values=ENSEMBL_SPECIES, width=25).grid(row=1, column=1, sticky="w", pady=3)
        ttk.Label(frm, text="Retrieve").grid(row=2, column=0, sticky="w", pady=3)
        ttk.Combobox(frm, textvariable=etype_var, values=ENSEMBL_TYPES, width=15, state="readonly").grid(row=2, column=1, sticky="w", pady=3)
        ttk.Label(frm, text="qPCR mode loads genomic + cDNA", foreground="#666").grid(row=3, column=0, columnspan=2, sticky="w", pady=(2, 6))

        result: dict[str, bool] = {"ok": False}

        def _ok() -> None:
            result["ok"] = True
            dialog.destroy()

        def _cancel() -> None:
            dialog.destroy()

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=2, sticky="w")
        ttk.Button(btns, text="OK", command=_ok).pack(side="left")
        ttk.Button(btns, text="Cancel", command=_cancel).pack(side="left", padx=6)

        _center_dialog_on_parent(dialog, root)
        dialog.grab_set()
        dialog.focus_force()
        dialog.wait_window()
        if not result["ok"]:
            return

        gene = gene_var.get().strip()
        species = species_var.get().strip()
        ensembl_type = etype_var.get().strip()
        if not gene or not species or not ensembl_type:
            messagebox.showerror("Ensembl", "Please fill gene, organism and retrieval type.")
            return

        state["ensembl_gene"] = gene
        state["ensembl_species"] = species
        state["ensembl_type"] = ensembl_type
        _sync_gene_display()
        _set_status("Retrieving sequence from Ensembl ...")
        root.update_idletasks()

        def _ui_after(func: Callable[[], None]) -> None:
            try:
                root.after(0, func)
            except Exception:
                pass

        def _set_retrieve_status_async(msg: str) -> None:
            _ui_after(lambda m=str(msg): _set_run_status(m))

        def _end_retrieve_busy() -> None:
            state["design_busy"] = False
            _set_busy_controls(True)
            _close_run_status_popup()

        def _fail_retrieve(exc: Exception) -> None:
            _end_retrieve_busy()
            messagebox.showerror("Ensembl", str(exc))
            _set_status(f"Ensembl error: {exc}")

        def _on_stage2_done(result: dict[str, object] | None, err: Exception | None) -> None:
            if err is not None:
                _fail_retrieve(err)
                return
            payload = dict(result or {})
            try:
                g_seq = str(payload.get("g_seq") or "")
                c_seq = str(payload.get("c_seq") or "")
                transcript_id = str(payload.get("transcript_id") or "")
                gene_id = str(payload.get("gene_id") or "")
                tx_bounds = [int(b) for b in (payload.get("tx_bounds") or [])]
                tx_snp_pos = [int(p) for p in (payload.get("tx_snp_pos") or [])]
                snp_bed_path = str(payload.get("snp_bed_path") or "")
                snp_count = int(payload.get("snp_count") or 0)
                snp_note = str(payload.get("snp_note") or "")
                auto_db_msg = str(payload.get("auto_db_msg") or "")

                if not g_seq or not c_seq:
                    raise RuntimeError(f"Unable to retrieve genomic/cDNA sequence for {transcript_id}")

                _set_genomic_sequence_display(g_seq)
                mrna_text.delete("1.0", "end")
                mrna_text.insert("1.0", c_seq)
                state["rows_after_ntthal"] = None
                state["parsed_primer3"] = None
                state["design_stats"] = None
                state["spidey_meta"] = {"used": False, "boundaries": []}
                state["spidey_output"] = ""
                state["template_len"] = len(c_seq)
                state["bounds"] = list(tx_bounds)
                state["snp_cdna_positions"] = list(tx_snp_pos)
                state["orf"] = _find_longest_orf(c_seq)
                state["spec_passed"] = set()
                _render_rows([])

                msg = f"{gene} ({species}) retrieved successfully: {transcript_id} [{gene_id}]"
                if snp_bed_path:
                    msg += f" | SNP BED temp loaded ({snp_count})"
                else:
                    msg += " | SNP BED unavailable; SNP check will be skipped."
                if auto_db_msg:
                    msg += f" | {auto_db_msg}"
                if snp_note and snp_note != "ok":
                    msg += f" ({snp_note})"
                _set_status(msg)
            except Exception as exc:
                messagebox.showerror("Ensembl", str(exc))
                _set_status(f"Ensembl error: {exc}")
            finally:
                _end_retrieve_busy()

        def _start_stage2(
            *,
            lookup_payload: dict[str, object],
            transcript_id: str,
            gene_id: str,
            auto_db_msg: str,
            snp_bed_path: str,
            snp_count: int,
            snp_note: str,
        ) -> None:
            _set_run_status("Downloading genomic and cDNA sequence from Ensembl ...")

            def _worker_stage2() -> None:
                try:
                    _set_retrieve_status_async("Mapping transcript bounds and SNP positions ...")
                    tx_bounds, tx_snp_pos = _cdna_bounds_and_snp_positions_from_transcript(
                        lookup_payload=lookup_payload,
                        transcript_id=transcript_id,
                        snp_bed_path=snp_bed_path,
                    )

                    _set_retrieve_status_async("Downloading genomic and cDNA sequence from Ensembl ...")
                    g_url = build_sequence_id_url(transcript_id, "genomic")
                    c_url = build_sequence_id_url(transcript_id, "cdna")
                    g_data = fetch_json_with_transport(g_url, _http_transport)
                    c_data = fetch_json_with_transport(c_url, _http_transport)
                    if isinstance(g_data, EnsemblError) or isinstance(c_data, EnsemblError):
                        msg = (g_data.message if isinstance(g_data, EnsemblError) else "") + " " + (
                            c_data.message if isinstance(c_data, EnsemblError) else ""
                        )
                        raise RuntimeError(msg.strip())
                    g_seq = str(dict(g_data).get("seq") or "")
                    c_seq = str(dict(c_data).get("seq") or "")
                    if not g_seq or not c_seq:
                        raise RuntimeError(f"Unable to retrieve genomic/cDNA sequence for {transcript_id}")

                    _ui_after(
                        lambda: _on_stage2_done(
                            {
                                "g_seq": g_seq,
                                "c_seq": c_seq,
                                "transcript_id": transcript_id,
                                "gene_id": gene_id,
                                "tx_bounds": list(tx_bounds),
                                "tx_snp_pos": list(tx_snp_pos),
                                "snp_bed_path": snp_bed_path,
                                "snp_count": int(snp_count),
                                "snp_note": snp_note,
                                "auto_db_msg": auto_db_msg,
                            },
                            None,
                        )
                    )
                except Exception as exc:
                    _ui_after(lambda e=exc: _on_stage2_done(None, e))

            threading.Thread(target=_worker_stage2, name="ensembl-retrieve-stage2", daemon=True).start()

        def _on_stage1_done(result: dict[str, object] | None, err: Exception | None) -> None:
            if err is not None:
                _fail_retrieve(err)
                return

            payload = dict(result or {})
            try:
                lookup_payload = dict(payload.get("lookup_payload") or {})
                gene_id = str(payload.get("gene_id") or "")
                snp_bed_path = str(payload.get("snp_bed_path") or "")
                snp_count = int(payload.get("snp_count") or 0)
                snp_note = str(payload.get("snp_note") or "")
                choices = list(payload.get("choices") or [])

                if not gene_id:
                    raise RuntimeError("Ensembl response missing stable gene ID.")
                if not choices:
                    raise RuntimeError(f"No transcripts returned for {gene}")

                # Temporary SNP BED is overwritten for each retrieved gene.
                state["current_snp_bed"] = snp_bed_path
                state["ensembl_gene_id"] = gene_id
                target_ensembl_gene_id_var.set(gene_id)

                auto_db_msg = _auto_select_transcriptome_db_for_target(
                    gene_id=gene_id,
                    species_hint=species,
                )
                target_slug = _infer_species_slug_from_gene_id(gene_id) or str(species or "").strip().lower()
                target_db_path = _find_best_ensembl_transcriptome_db(target_slug) if target_slug else ""
                if not target_db_path and target_slug:
                    _set_spec_db_available(False)
                    prompt_msg = (
                        f"No matching transcriptome DB for {target_slug} is installed.\n\n"
                        "NOTE: This is only necessary once and important for full functionality!\n"
                        "Dependent on your hardware this might take between 3 to 15 min.\n\n"
                        "Download and index now for specificity analysis?"
                    )
                    install_now = bool(messagebox.askyesno("Missing transcriptome DB", prompt_msg))
                    if install_now:
                        _set_status(f"Installing {target_slug} transcriptome DB for specificity analysis ...")

                        def _on_install_ok(fasta_path: str) -> None:
                            _set_spec_db_available(True)
                            _set_status(
                                f"Installed transcriptome DB for {target_slug}: "
                                f"{Path(str(fasta_path)).name if fasta_path else ''}. Spec testing re-enabled."
                            )

                        def _on_install_cancel_or_fail(err_txt: str) -> None:
                            run_spec = bool(spec_top50_var.get())
                            spec_top50_var.set(0)
                            _set_spec_db_available(False)
                            if str(err_txt) == "__CANCELLED__":
                                _set_status("Transcriptome DB install canceled. Spec testing disabled.")
                            else:
                                _set_status(f"Transcriptome DB install failed ({err_txt}). Spec testing disabled.")
                            if run_spec:
                                spec_top50_forced_off["value"] = True

                        _install_ensembl_db_for_species_async(
                            species_slug=target_slug,
                            on_success=_on_install_ok,
                            on_cancel_or_fail=_on_install_cancel_or_fail,
                            show_cancel_popup=True,
                        )
                    else:
                        run_spec = bool(spec_top50_var.get())
                        spec_top50_var.set(0)
                        _set_spec_db_available(False)
                        if run_spec:
                            spec_top50_forced_off["value"] = True
                        _set_status("No matching transcriptome DB installed. Spec testing disabled.")
                else:
                    _set_spec_db_available(True)

                preferred = choose_preferred_transcript(choices)
                if preferred is None:
                    raise RuntimeError(f"No valid transcript IDs returned for {gene}")
                chosen_label = preferred.display_label
                if len(choices) > 1:
                    pick = _choose_transcript_dialog(root, choices, preferred.display_label)
                    if not pick:
                        _end_retrieve_busy()
                        _set_status("Ensembl transcript selection canceled.")
                        return
                    chosen_label = pick
                by_label = {c.display_label: c.transcript_id for c in choices}
                transcript_id = by_label.get(chosen_label, preferred.transcript_id)
                _start_stage2(
                    lookup_payload=lookup_payload,
                    transcript_id=str(transcript_id),
                    gene_id=gene_id,
                    auto_db_msg=auto_db_msg,
                    snp_bed_path=snp_bed_path,
                    snp_count=snp_count,
                    snp_note=snp_note,
                )
            except Exception as exc:
                _fail_retrieve(exc)

        def _worker_stage1() -> None:
            try:
                _set_retrieve_status_async("Querying Ensembl gene lookup ...")
                lookup_url = build_lookup_symbol_url(species, gene)
                lookup = fetch_json_with_transport(lookup_url, _http_transport)
                if isinstance(lookup, EnsemblNoGeneFound):
                    raise RuntimeError("No gene found under this name.")
                if isinstance(lookup, EnsemblError):
                    raise RuntimeError(lookup.message)
                lookup_payload = dict(lookup)
                gene_id = _normalize_ensembl_gene_id(str(lookup_payload.get("id") or ""))
                if not gene_id:
                    raise RuntimeError("Ensembl response missing stable gene ID.")

                _set_retrieve_status_async("Fetching SNP annotation ...")
                snp_bed_path, snp_count, snp_note = _fetch_ensembl_variation_bed_temp(
                    gene_id=gene_id,
                    transport=_http_transport,
                )
                choices = extract_transcript_choices(lookup_payload)
                if not choices:
                    raise RuntimeError(f"No transcripts returned for {gene}")

                _ui_after(
                    lambda: _on_stage1_done(
                        {
                            "lookup_payload": lookup_payload,
                            "gene_id": gene_id,
                            "snp_bed_path": snp_bed_path,
                            "snp_count": int(snp_count),
                            "snp_note": snp_note,
                            "choices": choices,
                        },
                        None,
                    )
                )
            except Exception as exc:
                _ui_after(lambda e=exc: _on_stage1_done(None, e))

        state["design_busy"] = True
        _set_busy_controls(False)
        _ensure_run_status_popup("Retrieving sequence from Ensembl ...")
        threading.Thread(target=_worker_stage1, name="ensembl-retrieve-stage1", daemon=True).start()

    def _open_local_doc_page(title: str, filename: str) -> None:
        candidates = [
            Path(_primerl_path("docs", filename)),
            Path(__file__).resolve().parents[2] / "docs" / filename,
            Path.cwd() / "docs" / filename,
        ]
        page_path = next((p for p in candidates if p.exists() and p.is_file()), None)
        if page_path is None:
            messagebox.showerror(title, f"{filename} not found under docs/.")
            _set_status(f"{title} open failed: docs/{filename} was not found.")
            return
        try:
            webbrowser.open(page_path.resolve().as_uri(), new=2)
            _set_status(f"Opened {title}: {page_path}")
        except Exception as exc:
            messagebox.showerror(title, str(exc))
            _set_status(f"{title} open failed: {exc}")

    def _open_preferences() -> None:
        def _open_local_html(title: str, filename: str) -> None:
            _open_local_doc_page(title, filename)

        def _open_how_to() -> None:
            _open_local_html("How-To", "HOW_TO.html")

        def _open_about() -> None:
            _open_local_html("About", "ABOUT.html")

        def _open_existing_primers_tester() -> None:
            t_win = tk.Toplevel(win)
            t_win.title("Test Existing Primers")
            t_win.transient(win)
            t_frm = ttk.Frame(t_win, padding=10)
            t_frm.pack(fill="both", expand=True)

            species_labels = [label for label, _slug in ENSEMBL_DB_SPECIES_CHOICES]
            species_slug_by_label = {label: slug for label, slug in ENSEMBL_DB_SPECIES_CHOICES}
            species_label_by_slug = {slug: label for label, slug in ENSEMBL_DB_SPECIES_CHOICES}
            current_slug = str(state.get("ensembl_species") or "homo_sapiens").strip().lower()
            default_species_label = species_label_by_slug.get(current_slug, species_labels[0] if species_labels else "")

            fp_var = tk.StringVar(value="")
            rp_var = tk.StringVar(value="")
            gene_var = tk.StringVar(value=str(state.get("ensembl_gene") or ""))
            species_var = tk.StringVar(value=default_species_label)
            db_hint_var = tk.StringVar(value="")
            checks_summary_var = tk.StringVar(value="Checks passed: -")
            busy = {"value": False}

            def _safe_int(txt: str, default: int) -> int:
                try:
                    return int(str(txt).strip())
                except Exception:
                    return int(default)

            def _safe_float(txt: str, default: float) -> float:
                try:
                    return float(str(txt).strip())
                except Exception:
                    return float(default)

            def _simple_tm(seq: str) -> float:
                s = str(seq or "").upper()
                at = s.count("A") + s.count("T")
                gc = s.count("G") + s.count("C")
                return float((2 * at) + (4 * gc))

            def _gc_pct(seq: str) -> float:
                s = str(seq or "").upper()
                if not s:
                    return 0.0
                gc = s.count("G") + s.count("C")
                return (100.0 * gc) / len(s)

            def _has_excluded_repeats_or_runs_local(seq: str, run_n: int, repeat_n: int) -> bool:
                run_n = max(1, int(run_n))
                repeat_real = max(0, int(repeat_n) - 1)
                run_pat = re.compile(rf"(C{{{run_n},}}|A{{{run_n},}}|G{{{run_n},}}|T{{{run_n},}})", re.IGNORECASE)
                if run_pat.search(seq):
                    return True
                rep_pat = re.compile(rf"(.{{2,}})\1{{{repeat_real},}}")
                return bool(rep_pat.search(seq))

            def _reverse_complement(seq: str) -> str:
                tr = str.maketrans("ACGTacgt", "TGCAtgca")
                return str(seq or "").translate(tr)[::-1]

            def _find_all_subseq(haystack: str, needle: str, max_hits: int = 64) -> list[int]:
                src = str(haystack or "")
                pat = str(needle or "")
                if not src or not pat:
                    return []
                hits: list[int] = []
                start = 0
                while len(hits) < max_hits:
                    pos = src.find(pat, start)
                    if pos < 0:
                        break
                    hits.append(pos)
                    start = pos + 1
                return hits

            def _resolve_tester_ensembl_context(
                species_slug_txt: str,
                target_gene_txt: str,
            ) -> tuple[dict[str, object] | None, str]:
                gene_txt = str(target_gene_txt or "").strip()
                if not gene_txt:
                    return None, "missing gene"

                gene_id = _normalize_ensembl_gene_id(gene_txt)
                if gene_id:
                    lookup_url = f"https://rest.ensembl.org/lookup/id/{gene_id}?expand=1"
                else:
                    species_norm = str(species_slug_txt or "").strip().lower()
                    if not species_norm:
                        return None, "missing species for Ensembl symbol lookup"
                    lookup_url = build_lookup_symbol_url(species_norm, gene_txt)

                lookup = fetch_json_with_transport(lookup_url, _http_transport)
                if isinstance(lookup, EnsemblNoGeneFound):
                    return None, "No Ensembl gene found for provided symbol"
                if isinstance(lookup, EnsemblError):
                    return None, str(lookup.message or "Ensembl lookup failed")
                if not isinstance(lookup, dict):
                    return None, "Unexpected Ensembl lookup payload"

                lookup_payload = dict(lookup)
                resolved_gene_id = _normalize_ensembl_gene_id(str(lookup_payload.get("id") or gene_id))
                choices = extract_transcript_choices(lookup_payload)
                preferred = choose_preferred_transcript(choices)
                if preferred is None:
                    return None, "No transcript choices available from Ensembl"

                transcript_id = str(preferred.transcript_id or "").strip()
                if not transcript_id:
                    return None, "Ensembl transcript ID missing"

                tx_bounds, _ = _cdna_bounds_and_snp_positions_from_transcript(
                    lookup_payload=lookup_payload,
                    transcript_id=transcript_id,
                    snp_bed_path="",
                )

                c_url = build_sequence_id_url(transcript_id, "cdna")
                c_data = fetch_json_with_transport(c_url, _http_transport)
                if isinstance(c_data, EnsemblError):
                    return None, str(c_data.message or "Unable to fetch Ensembl cDNA")
                if not isinstance(c_data, dict):
                    return None, "Unexpected Ensembl cDNA payload"

                cdna_seq = clean_sequence(str(c_data.get("seq") or "")).upper()
                if not cdna_seq:
                    return None, f"Unable to retrieve cDNA sequence for transcript {transcript_id}"

                return {
                    "gene_id": resolved_gene_id,
                    "transcript_id": transcript_id,
                    "cdna_seq": cdna_seq,
                    "bounds": list(tx_bounds),
                }, ""

            def _select_cdna_binding_pair(
                *,
                cdna_seq: str,
                f_seq: str,
                r_seq: str,
                min_amp: int,
                max_amp: int,
            ) -> tuple[dict[str, int] | None, str]:
                f_hits = _find_all_subseq(cdna_seq, f_seq)
                r_hits = _find_all_subseq(cdna_seq, _reverse_complement(r_seq))
                if not f_hits or not r_hits:
                    return None, f"Primer mapping not found on cDNA (forward_hits={len(f_hits)}, reverse_hits={len(r_hits)})"

                target_amp = (int(min_amp) + int(max_amp)) / 2.0
                candidates: list[tuple[int, float, int, int, int]] = []
                len_f = len(f_seq)
                len_r = len(r_seq)
                for pos_f in f_hits:
                    for r_start in r_hits:
                        if r_start <= (pos_f + len_f):
                            continue
                        realpos_r = r_start + len_r
                        amp_size = realpos_r - pos_f
                        in_range = (int(min_amp) <= amp_size <= int(max_amp))
                        candidates.append((0 if in_range else 1, abs(amp_size - target_amp), amp_size, pos_f, realpos_r))

                if not candidates:
                    return None, "Mapped primer hits were incompatible in orientation/order"

                candidates.sort(key=lambda x: (x[0], x[1], x[2]))
                _range_flag, _dist, amp_size, pos_f, realpos_r = candidates[0]
                return {
                    "pos_f": int(pos_f),
                    "realpos_r": int(realpos_r),
                    "amp_size": int(amp_size),
                    "forward_hits": int(len(f_hits)),
                    "reverse_hits": int(len(r_hits)),
                }, ""

            def _selected_species_slug() -> str:
                return str(species_slug_by_label.get(species_var.get().strip(), "")).strip().lower()

            def _selected_species_db() -> str:
                slug = _selected_species_slug()
                return _find_best_ensembl_transcriptome_db(slug) if slug else ""

            def _refresh_db_hint(*_args: object) -> None:
                db_path = _selected_species_db()
                if db_path and Path(db_path).exists() and Path(str(db_path) + ".primerqc").exists():
                    db_hint_var.set(f"Spec DB: {Path(db_path).name}")
                elif db_path and Path(db_path).exists():
                    db_hint_var.set(f"Spec DB not indexed: {Path(db_path).name}.primerqc missing")
                else:
                    db_hint_var.set("Spec DB not found for selected species (install in Options).")

            def _render_test_pie(counts: dict[str, object]) -> None:
                pie_canvas.delete("all")
                pie_legend.configure(state="normal")
                pie_legend.delete("1.0", "end")
                pie_legend.tag_configure("hdr", font=("TkDefaultFont", 10, "bold"))
                checks_total = max(0, int(counts.get("checks_total", 0)))
                checks_passed = max(0, int(counts.get("checks_passed", 0)))
                checks_failed = max(0, int(counts.get("checks_failed", 0)))
                if checks_passed + checks_failed > checks_total:
                    checks_total = checks_passed + checks_failed
                parts: list[tuple[str, int, str]] = [
                    ("Passed checks", checks_passed, "#59a14f"),
                    ("Failed checks", checks_failed, "#e15759"),
                ]

                x0, y0, x1, y1 = 18, 14, 290, 250
                start = 0.0
                pie_total = sum(max(0, int(v)) for _, v, _ in parts)
                nonzero_parts = [(label, value, color) for label, value, color in parts if value > 0]
                cx = (x0 + x1) / 2.0
                cy = (y0 + y1) / 2.0
                rx = max(1.0, (x1 - x0) / 2.0)
                ry = max(1.0, (y1 - y0) / 2.0)

                pie_legend.insert("end", "Primer Test Outcome Pie\n", ("hdr",))

                drawn_slices = 0
                consumed = 0.0
                for idx, (_label, value, color) in enumerate(nonzero_parts):
                    extent = 0.0
                    if pie_total > 0:
                        if idx == len(nonzero_parts) - 1:
                            extent = max(0.0, 360.0 - consumed)
                        else:
                            extent = 360.0 * (value / pie_total)
                    if extent > 0.0:
                        steps = max(6, int(extent / 3.0))
                        pts: list[float] = [cx, cy]
                        for step in range(steps + 1):
                            ang = start + (extent * step / steps)
                            rad = math.radians(ang)
                            pts.append(cx + rx * math.cos(rad))
                            pts.append(cy - ry * math.sin(rad))
                        pie_canvas.create_polygon(pts, fill=color, outline="white", width=1)
                        consumed += extent
                        drawn_slices += 1
                    start += extent

                if checks_total <= 0 or drawn_slices == 0:
                    pie_canvas.create_text(
                        int((x0 + x1) / 2),
                        int((y0 + y1) / 2),
                        text="No checks run yet",
                        fill="#555555",
                    )
                statuses_raw = counts.get("check_statuses")
                statuses = statuses_raw if isinstance(statuses_raw, list) else []
                if statuses:
                    for item in statuses:
                        if not isinstance(item, dict):
                            continue
                        nm = str(item.get("name", "Check"))
                        ok = bool(item.get("passed", False))
                        detail = str(item.get("detail", "") or "").strip()
                        status_txt = "PASS" if ok else "FAIL"
                        if (nm == "Primer rules") and (not ok) and detail:
                            pie_legend.insert("end", f"- {nm}: {status_txt} ({detail})\n")
                        else:
                            pie_legend.insert("end", f"- {nm}: {status_txt}\n")
                else:
                    pie_legend.insert("end", "- No checks run yet\n")
                pie_legend.configure(state="disabled")

            def _evaluate_existing_primer_pair(
                f_seq: str,
                r_seq: str,
                species_slug: str,
                target_gene_raw: str,
            ) -> dict[str, object]:
                statuses: list[dict[str, object]] = []
                notes: list[str] = []

                def _add_status(name: str, passed: bool, detail: str) -> None:
                    statuses.append(
                        {
                            "name": str(name),
                            "passed": bool(passed),
                            "detail": str(detail or ""),
                        }
                    )
                    if detail:
                        notes.append(f"{name}: {detail}")

                min_len_n = max(1, _safe_int(min_len_var.get(), 20))
                max_len_n = max(min_len_n, _safe_int(max_len_var.get(), 24))
                min_tm_n = _safe_float(min_tm_var.get(), 58.0)
                max_tm_n = _safe_float(max_tm_var.get(), 62.0)
                if max_tm_n < min_tm_n:
                    max_tm_n = min_tm_n
                max_delta_tm_n = max(0.0, _safe_float(max_diff_var.get(), 2.0))
                run_n = max(1, _safe_int(run_var.get(), 4))
                repeat_n = max(2, _safe_int(repeat_var.get(), 4))
                min_amp_n = max(1, _safe_int(min_amp_var.get(), 100))
                max_amp_n = max(min_amp_n, _safe_int(max_amp_var.get(), 300))
                target_gene_txt = str(target_gene_raw or "").strip()
                target_gene_id = _normalize_ensembl_gene_id(target_gene_txt)
                target_gene_symbol = "" if target_gene_id else target_gene_txt
                context_needed = bool(ie_span_var.get()) or bool(ie_overlap_var.get()) or bool(ie_limit_var.get())
                tester_ensembl_ctx: dict[str, object] | None = None
                tester_ensembl_err = ""
                if context_needed and target_gene_txt:
                    tester_ensembl_ctx, tester_ensembl_err = _resolve_tester_ensembl_context(species_slug, target_gene_txt)
                    if tester_ensembl_ctx is not None:
                        resolved_gene_id = _normalize_ensembl_gene_id(str(tester_ensembl_ctx.get("gene_id") or ""))
                        if resolved_gene_id:
                            target_gene_id = resolved_gene_id
                            target_gene_symbol = ""

                len_f = len(f_seq)
                len_r = len(r_seq)
                tm_f = _simple_tm(f_seq)
                tm_r = _simple_tm(r_seq)
                gc_f = _gc_pct(f_seq)
                gc_r = _gc_pct(r_seq)

                primer_rule_issues: list[str] = []
                if len_f < min_len_n or len_f > max_len_n or len_r < min_len_n or len_r > max_len_n:
                    primer_rule_issues.append(
                        f"Length required {min_len_n}-{max_len_n} bp, got F={len_f}, R={len_r}"
                    )
                if tm_f < min_tm_n or tm_f > max_tm_n or tm_r < min_tm_n or tm_r > max_tm_n:
                    primer_rule_issues.append(
                        f"Tm required {min_tm_n:.1f}-{max_tm_n:.1f}, got F={tm_f:.1f}, R={tm_r:.1f}"
                    )
                if abs(tm_f - tm_r) > max_delta_tm_n:
                    primer_rule_issues.append(
                        f"Delta Tm max {max_delta_tm_n:.1f}, got {abs(tm_f - tm_r):.1f}"
                    )
                if bool(exclude_gc_var.get()) and (gc_f < 40.0 or gc_f > 60.0 or gc_r < 40.0 or gc_r > 60.0):
                    primer_rule_issues.append(
                        f"GC required 40-60%, got F={gc_f:.1f}%, R={gc_r:.1f}%"
                    )
                if bool(gc_clamp_var.get()) and ((f_seq[-1] not in "GC") or (r_seq[-1] not in "GC")):
                    primer_rule_issues.append("GC clamp failed (3' base must be G/C for both primers)")
                _add_status(
                    "Primer rules",
                    passed=(len(primer_rule_issues) == 0),
                    detail=("OK" if not primer_rule_issues else "; ".join(primer_rule_issues)),
                )

                if bool(exclude_rr_var.get()):
                    rr_fail = _has_excluded_repeats_or_runs_local(f_seq, run_n, repeat_n) or _has_excluded_repeats_or_runs_local(
                        r_seq, run_n, repeat_n
                    )
                    _add_status("Repeat/run filter", passed=(not rr_fail), detail=("OK" if not rr_fail else "Repeat/run pattern detected"))
                else:
                    _add_status("Repeat/run filter", passed=True, detail="Disabled")

                if not context_needed:
                    _add_status("Exon/intron context checks", passed=True, detail="Not required")
                elif not target_gene_txt:
                    _add_status(
                        "Exon/intron context checks",
                        passed=True,
                        detail="Disabled (no Ensembl gene provided)",
                    )
                elif tester_ensembl_ctx is None:
                    _add_status(
                        "Exon/intron context checks",
                        passed=False,
                        detail=f"Ensembl context unavailable: {tester_ensembl_err or 'lookup failed'}",
                    )
                else:
                    tx_id = str(tester_ensembl_ctx.get("transcript_id") or "")
                    cdna_seq = clean_sequence(str(tester_ensembl_ctx.get("cdna_seq") or "")).upper()
                    bounds_raw = tester_ensembl_ctx.get("bounds")
                    bounds: list[int] = []
                    if isinstance(bounds_raw, list):
                        for b in bounds_raw:
                            try:
                                bi = int(b)
                            except Exception:
                                continue
                            if bi > 0:
                                bounds.append(bi)
                    bounds = sorted(set(bounds))

                    pair_map, pair_map_err = _select_cdna_binding_pair(
                        cdna_seq=cdna_seq,
                        f_seq=f_seq,
                        r_seq=r_seq,
                        min_amp=min_amp_n,
                        max_amp=max_amp_n,
                    )
                    if pair_map is None:
                        _add_status(
                            "Exon/intron context checks",
                            passed=False,
                            detail=(f"Primer mapping failed on {tx_id}: {pair_map_err}" if tx_id else pair_map_err),
                        )
                    else:
                        pos_f = int(pair_map.get("pos_f", 0))
                        realpos_r = int(pair_map.get("realpos_r", 0))
                        amp_size_ctx = int(pair_map.get("amp_size", 0))
                        len_f_ctx = len(f_seq)
                        len_r_ctx = len(r_seq)

                        context_ok = True
                        detail_parts: list[str] = []
                        if tx_id:
                            detail_parts.append(f"tx={tx_id}")
                        detail_parts.append(f"amp={amp_size_ctx}bp")
                        detail_parts.append(f"boundaries={len(bounds)}")

                        if bool(ie_span_var.get()):
                            span_ok = bool(bounds) and any(pos_f < b < realpos_r for b in bounds)
                            context_ok = context_ok and span_ok
                            detail_parts.append(f"span={'OK' if span_ok else 'FAIL'}")

                        if bool(ie_overlap_var.get()):
                            exclude_ie_n = max(0, _safe_int(exclude_ie_var.get(), 0))
                            if exclude_ie_n <= 0:
                                detail_parts.append("overlap=SKIP(radius=0)")
                            else:
                                overlap_ok = False
                                for b in bounds:
                                    if pos_f < (b - exclude_ie_n) and (b + exclude_ie_n) < (pos_f + len_f_ctx):
                                        overlap_ok = True
                                    if (realpos_r - len_r_ctx) < (b - exclude_ie_n) and (b + exclude_ie_n) < realpos_r:
                                        overlap_ok = True
                                context_ok = context_ok and overlap_ok
                                detail_parts.append(f"overlap={'OK' if overlap_ok else 'FAIL'}")

                        if bool(ie_limit_var.get()):
                            lo, hi, ex5, ex3, exons = _selected_exon_window(
                                bounds=bounds,
                                template_len=len(cdna_seq),
                                ex5_txt=str(ie_5p_var.get()),
                                ex3_txt=str(ie_3p_var.get()),
                            )
                            f_end = pos_f + len_f_ctx
                            r_start = realpos_r - len_r_ctx
                            limit_ok = pos_f >= lo and f_end <= hi and r_start >= lo and realpos_r <= hi
                            context_ok = context_ok and limit_ok
                            detail_parts.append(
                                f"limit={'OK' if limit_ok else 'FAIL'}[{ex5}-{ex3}/{exons}]"
                            )

                        detail_parts.append(
                            f"hits(F={int(pair_map.get('forward_hits', 0))},R={int(pair_map.get('reverse_hits', 0))})"
                        )
                        _add_status(
                            "Exon/intron context checks",
                            passed=context_ok,
                            detail="; ".join(part for part in detail_parts if part),
                        )

                amp_size = max(min_amp_n, min(max_amp_n, max(len_f + len_r + 1, 140)))
                realpos_r = max(amp_size, len_f + len_r + 1)
                seed_rows: list[list[object]] = [
                    [
                        f_seq,
                        0,
                        len_f,
                        f"{tm_f:.2f}",
                        r_seq,
                        0,
                        len_r,
                        f"{tm_r:.2f}",
                        realpos_r,
                        amp_size,
                        "0.00",
                        0,
                        0,
                        "0.00",
                    ]
                ]
                refined_seed_rows = [list(r) for r in seed_rows]
                try:
                    refined_seed_rows = _refine_rows_with_ntthal(
                        [list(r) for r in seed_rows],
                        primer3_path=primer3_var.get(),
                        ntthal_path=ntthal_var.get(),
                    )
                except Exception as exc:
                    notes.append(f"ntthal refine error: {exc}")

                if bool(run_ntthal_cutoff_var.get()):
                    try:
                        rows_nt = [list(r) for r in refined_seed_rows]
                        rows_nt, nt_note = _filter_rows_by_ntthal_ext_cutoff(
                            rows_nt,
                            ext_dg_cutoff=_safe_float(ntthal_ext_cutoff_var.get(), -3.5),
                        )
                        nt_ok = len(rows_nt) > 0
                        _add_status(
                            "Thermodynamics",
                            passed=nt_ok,
                            detail=(nt_note or ("OK" if nt_ok else "Failed")),
                        )
                    except Exception as exc:
                        _add_status("Thermodynamics", passed=False, detail=f"Error: {exc}")
                else:
                    _add_status("Thermodynamics", passed=True, detail="Disabled")

                if bool(run_mfeprimer_var.get()):
                    try:
                        rows_mfe = [list(r) for r in refined_seed_rows]
                        rows_mfe, mfe_note = _filter_rows_with_mfeprimer(
                            rows_mfe,
                            mfeprimer_path=str(mfeprimer_var.get() or ""),
                            dg_cutoff=_safe_float(mfe_dg_cutoff_var.get(), -2.0),
                        )
                        mfe_ok = len(rows_mfe) > 0
                        _add_status(
                            "MFEprimer dimer",
                            passed=mfe_ok,
                            detail=(mfe_note or ("OK" if mfe_ok else "Failed")),
                        )
                    except Exception as exc:
                        _add_status("MFEprimer dimer", passed=False, detail=f"Error: {exc}")
                else:
                    _add_status("MFEprimer dimer", passed=True, detail="Disabled")

                try:
                    db_path = _find_best_ensembl_transcriptome_db(species_slug)
                    db = Path(db_path) if db_path else None
                    idx = Path(str(db) + ".primerqc") if db is not None else None
                    if db is None or not db.exists() or idx is None or not idx.exists():
                        _add_status(
                            "Transcriptome specificity",
                            passed=False,
                            detail="Transcriptome DB missing or not indexed for selected species",
                        )
                    else:
                        spec_warn = ""

                        def _on_spec_param_error(msg: str) -> None:
                            nonlocal spec_warn
                            spec_warn = str(msg or "")

                        spec_extra = resolve_spec_param_tokens(
                            str(mfeprimer_spec_params_var.get() or ""),
                            on_error=_on_spec_param_error,
                        )
                        spec_metrics: dict[str, int] = {"target_removed": 0, "snp_removed": 0}
                        spec_rows = [list(r) for r in refined_seed_rows]
                        spec_rows, spec_note = _filter_rows_with_mfeprimer_spec(
                            spec_rows,
                            mfeprimer_path=str(mfeprimer_var.get() or ""),
                            db_fasta_path=str(db),
                            target_gene_id=target_gene_id,
                            target_gene_symbol=target_gene_symbol,
                            snp_check_enabled=False,
                            max_amplicons=1,
                            min_amp_size=int(SPEC_OFFTARGET_MIN_AMP_SIZE_BP),
                            max_amp_size=int(SPEC_OFFTARGET_MAX_AMP_SIZE_BP),
                            spec_selection_mode="strict_pass",
                            spec_remove_pct=10.0,
                            metrics_out=spec_metrics,
                            spec_extra_args=spec_extra,
                        )
                        spec_ok = len(spec_rows) > 0
                        spec_detail_parts = [str(spec_note or "").strip()]
                        if not target_gene_id and not target_gene_symbol:
                            spec_detail_parts.append("No target gene provided: count-only mode")
                        if spec_warn:
                            spec_detail_parts.append(spec_warn)
                        spec_detail = " | ".join(p for p in spec_detail_parts if p) or ("OK" if spec_ok else "Failed")
                        _add_status("Transcriptome specificity", passed=spec_ok, detail=spec_detail)
                except Exception as exc:
                    _add_status("Transcriptome specificity", passed=False, detail=f"Error: {exc}")

                checks_total = len(statuses)
                checks_passed = sum(1 for s in statuses if bool(s.get("passed", False)))
                checks_failed = max(0, checks_total - checks_passed)
                overall_pass = checks_failed == 0
                summary = f"Checks passed: {checks_passed}/{checks_total}."
                notes_txt = " | ".join(n for n in notes if n)
                return {
                    "passed": bool(overall_pass),
                    "counts": {
                        "checks_total": checks_total,
                        "checks_passed": checks_passed,
                        "checks_failed": checks_failed,
                        "check_statuses": statuses,
                    },
                    "notes": (summary + (" | " + notes_txt if notes_txt else "")),
                }

            def _set_busy(flag: bool) -> None:
                busy["value"] = bool(flag)
                state_txt = "disabled" if flag else "normal"
                try:
                    run_btn.configure(state=state_txt)
                    fp_entry.configure(state=state_txt)
                    rp_entry.configure(state=state_txt)
                    gene_entry.configure(state=state_txt)
                    species_box.configure(state=("disabled" if flag else "readonly"))
                except Exception:
                    pass

            def _finish_test(payload: dict[str, object] | None, err: Exception | None) -> None:
                _set_busy(False)
                if err is not None:
                    checks_summary_var.set("Checks passed: 0/1")
                    _render_test_pie(
                        {
                            "checks_total": 1,
                            "checks_passed": 0,
                            "checks_failed": 1,
                            "check_statuses": [{"name": "Internal run", "passed": False, "detail": str(err)}],
                        }
                    )
                    return
                data = payload if isinstance(payload, dict) else {}
                counts = data.get("counts")
                if isinstance(counts, dict):
                    passed_n = max(0, int(counts.get("checks_passed", 0)))
                    total_n = max(0, int(counts.get("checks_total", 0)))
                    checks_summary_var.set(f"Checks passed: {passed_n}/{total_n}")
                    _render_test_pie(counts)
                else:
                    checks_summary_var.set("Checks passed: -")

            def _run_test() -> None:
                if busy["value"]:
                    return
                f_seq = clean_sequence(fp_var.get()).upper()
                r_seq = clean_sequence(rp_var.get()).upper()
                if not f_seq or not r_seq:
                    messagebox.showerror("Test existing primers", "Please enter both primer sequences.")
                    return

                species_slug = _selected_species_slug()
                if not species_slug:
                    messagebox.showerror("Test existing primers", "Please select a species.")
                    return

                _set_busy(True)
                checks_summary_var.set("Checks passed: ...")

                def _worker() -> None:
                    try:
                        payload = _evaluate_existing_primer_pair(
                            f_seq=f_seq,
                            r_seq=r_seq,
                            species_slug=species_slug,
                            target_gene_raw=gene_var.get(),
                        )
                        root.after(0, lambda: _finish_test(payload, None))
                    except Exception as exc:
                        root.after(0, lambda: _finish_test(None, exc))

                threading.Thread(target=_worker, daemon=True).start()

            row = 0
            ttk.Label(t_frm, text="Forward primer").grid(row=row, column=0, sticky="w", pady=(0, 4))
            fp_entry = ttk.Entry(t_frm, textvariable=fp_var, width=44)
            fp_entry.grid(row=row, column=1, sticky="ew", padx=(8, 0), pady=(0, 4))
            row += 1
            ttk.Label(t_frm, text="Reverse primer").grid(row=row, column=0, sticky="w", pady=4)
            rp_entry = ttk.Entry(t_frm, textvariable=rp_var, width=44)
            rp_entry.grid(row=row, column=1, sticky="ew", padx=(8, 0), pady=4)
            row += 1
            ttk.Label(t_frm, text="Target gene (symbol or Ensembl ID)").grid(row=row, column=0, sticky="w", pady=4)
            gene_entry = ttk.Entry(t_frm, textvariable=gene_var, width=44)
            gene_entry.grid(row=row, column=1, sticky="ew", padx=(8, 0), pady=4)
            row += 1
            ttk.Label(t_frm, text="Species (spec DB)").grid(row=row, column=0, sticky="w", pady=4)
            species_box = ttk.Combobox(
                t_frm,
                textvariable=species_var,
                values=species_labels,
                state="readonly",
                width=34,
            )
            species_box.grid(row=row, column=1, sticky="w", padx=(8, 0), pady=4)
            species_box.bind("<<ComboboxSelected>>", _refresh_db_hint)
            row += 1
            ttk.Label(t_frm, textvariable=db_hint_var, foreground="#444444").grid(
                row=row,
                column=0,
                columnspan=2,
                sticky="w",
                pady=(0, 6),
            )
            row += 1

            run_btn = ttk.Button(t_frm, text="Run Primer Test", command=_run_test, bootstyle="secondary")
            run_btn.grid(row=row, column=0, sticky="w", pady=(0, 8))
            checks_lbl = tk.Label(
                t_frm,
                textvariable=checks_summary_var,
                anchor="w",
                font=("TkDefaultFont", 10, "bold"),
            )
            checks_lbl.grid(row=row, column=1, sticky="w", padx=(8, 0), pady=(0, 8))
            row += 1

            pie_canvas = tk.Canvas(t_frm, width=320, height=270, bg="white")
            pie_canvas.grid(row=row, column=0, sticky="nsew")
            pie_legend = tk.Text(t_frm, width=52, height=15, wrap="word")
            pie_legend.grid(row=row, column=1, sticky="nsew", padx=(10, 0))
            row += 1

            ttk.Button(t_frm, text="Close", command=t_win.destroy).grid(row=row, column=1, sticky="e", pady=(8, 0))
            t_frm.columnconfigure(1, weight=1)
            t_frm.rowconfigure(row - 1, weight=1)
            _refresh_db_hint()
            _render_test_pie({"checks_total": 0, "checks_passed": 0, "checks_failed": 0, "check_statuses": []})
            _center_dialog_on_parent(t_win, win)
            t_win.grab_set()
            t_win.focus_force()

        win = tk.Toplevel(root)
        win.title("Options")
        win.transient(root)
        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)
        pref_style = ttk.Style(win)
        pref_style.configure("HowTo.TButton", font=("Segoe UI", 11, "bold"), padding=(18, 10))
        pref_style.configure("About.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 8))

        profile_row = 0
        ttk.Label(frm, text="Binary profile").grid(row=profile_row, column=0, sticky="w", pady=(8, 3))
        binary_profile_box = ttk.Combobox(
            frm,
            textvariable=binary_profile_var,
            values=list(BINARY_PROFILE_CHOICES),
            width=20,
            state="readonly",
        )
        binary_profile_box.grid(row=profile_row, column=1, sticky="w", padx=6, pady=(8, 3))
        ttk.Button(
            frm,
            text="Apply",
            command=lambda: _apply_binary_profile(binary_profile_var.get(), notify=True, persist=True),
        ).grid(row=profile_row, column=2, sticky="w", pady=(8, 3))
        ttk.Button(
            frm,
            text="[HOW-TO] Guide",
            command=_open_how_to,
            style="HowTo.TButton",
            width=17,
        ).grid(row=profile_row, column=4, sticky="w", padx=(8, 0), pady=(6, 3))
        ttk.Button(
            frm,
            text="Test existing primers",
            command=_open_existing_primers_tester,
            bootstyle="secondary",
            width=19,
        ).grid(row=profile_row + 1, column=4, columnspan=2, sticky="w", padx=(8, 0), pady=(2, 6))
        ttk.Button(
            frm,
            text="About",
            command=_open_about,
            style="About.TButton",
            width=10,
        ).grid(row=profile_row, column=5, sticky="w", padx=(6, 0), pady=(8, 3))
        binary_profile_box.bind(
            "<<ComboboxSelected>>",
            lambda _evt: _apply_binary_profile(binary_profile_var.get(), notify=True, persist=True),
        )

        rr = profile_row + 2
        ttk.Label(frm, text="Max genomic view bases").grid(row=rr, column=0, sticky="w", pady=2)
        ttk.Entry(frm, textvariable=max_genomic_view_bases_var, width=12).grid(row=rr, column=1, sticky="w", padx=6, pady=2)

        def _apply_max_genomic_view_bases() -> None:
            n = _max_genomic_view_bases()
            _save_runtime_settings()
            full = str(state.get("genomic_seq_full") or "")
            if full:
                _set_genomic_sequence_display(full)
            _set_status(f"Max genomic view bases set to {n:,}.")

        ttk.Button(frm, text="Apply", command=_apply_max_genomic_view_bases).grid(row=rr, column=2, sticky="w", pady=2)
        rr += 1
        ttk.Separator(frm, orient="horizontal").grid(row=rr, column=0, columnspan=6, sticky="ew", pady=(8, 8))
        rr += 1

        species_values = [label for label, _slug in ENSEMBL_DB_SPECIES_CHOICES]
        species_slug_by_label = {label: slug for label, slug in ENSEMBL_DB_SPECIES_CHOICES}
        ensembl_db_species_var = tk.StringVar(value=species_values[0])
        install_status_var = tk.StringVar(value="")

        ttk.Label(frm, text="Install Ensembl DB").grid(row=rr, column=0, sticky="w", pady=(0, 2))
        species_box = ttk.Combobox(
            frm,
            textvariable=ensembl_db_species_var,
            values=species_values,
            width=35,
            state="readonly",
        )
        species_box.grid(row=rr, column=1, sticky="w", padx=6, pady=(0, 2))

        install_btn = ttk.Button(frm, text="Download + Index")
        install_btn.grid(row=rr, column=2, sticky="w", pady=(0, 2))
        rr += 1

        ttk.Label(frm, textvariable=install_status_var, foreground="#444").grid(row=rr, column=1, columnspan=2, sticky="w", padx=6, pady=(0, 6))
        rr += 1

        def _finish_install(success: bool, msg: str, fasta_path: str = "") -> None:
            if win.winfo_exists():
                try:
                    species_box.configure(state="readonly")
                    install_btn.configure(state="normal")
                except Exception:
                    pass
            if success:
                if fasta_path:
                    mfeprimer_transcriptome_fasta_var.set(fasta_path)
                install_status_var.set("Installed and indexed.")
                _set_status(msg)
            else:
                install_status_var.set(f"Failed: {msg}")
                _set_status(f"Ensembl DB install failed: {msg}")
                messagebox.showerror("Ensembl DB install", msg)

        def _install_ensembl_db() -> None:
            species_label = ensembl_db_species_var.get().strip()
            species_slug = species_slug_by_label.get(species_label, "")
            if not species_slug:
                messagebox.showerror("Ensembl DB install", "Please select a supported species.")
                return

            install_status_var.set("Working... this can take several minutes.")
            species_box.configure(state="disabled")
            install_btn.configure(state="disabled")
            _set_status(f"Preparing Ensembl {species_label} transcriptome install ...")
            _install_ensembl_db_for_species_async(
                species_slug=species_slug,
                on_success=lambda p: _finish_install(
                    True,
                    f"Installed Ensembl transcriptome DB: {Path(str(p)).name if p else ''}",
                    p,
                ),
                on_cancel_or_fail=lambda e: _finish_install(False, str(e), ""),
                show_cancel_popup=False,
            )

        install_btn.configure(command=_install_ensembl_db)

        ttk.Checkbutton(frm, text="Large intron (Spidey -X legacy)", variable=spidey_large_var).grid(row=rr, column=0, columnspan=2, sticky="w", pady=2)
        rr += 1
        mfe_cut_row = ttk.Frame(frm)
        mfe_cut_row.grid(row=rr, column=0, columnspan=3, sticky="w", pady=2)
        ttk.Checkbutton(mfe_cut_row, text="MFEprimer dimer cutoff", variable=run_mfeprimer_var).pack(side="left")
        ttk.Label(mfe_cut_row, text="dg cutoff").pack(side="left", padx=(8, 2))
        ttk.Entry(mfe_cut_row, textvariable=mfe_dg_cutoff_var, width=10).pack(side="left")
        rr += 1
        nt_cut_row = ttk.Frame(frm)
        nt_cut_row.grid(row=rr, column=0, columnspan=3, sticky="w", pady=2)
        ttk.Checkbutton(nt_cut_row, text="ntthal dimer cutoff", variable=run_ntthal_cutoff_var).pack(side="left")
        ttk.Label(nt_cut_row, text="Ext dG cutoff").pack(side="left", padx=(8, 2))
        ttk.Entry(nt_cut_row, textvariable=ntthal_ext_cutoff_var, width=10).pack(side="left")
        rr += 1

        ttk.Separator(frm, orient="horizontal").grid(row=rr, column=0, columnspan=6, sticky="ew", pady=(8, 8))
        rr += 1
        spec_param_group = ttk.LabelFrame(frm, text="Specificity check parameters")
        spec_param_group.grid(row=rr, column=0, columnspan=6, sticky="ew", pady=(2, 2))
        ttk.Entry(spec_param_group, textvariable=mfeprimer_spec_params_var, width=40).grid(
            row=0, column=0, sticky="w", padx=6, pady=(6, 2)
        )
        ttk.Button(spec_param_group, text="Apply", command=_save_runtime_settings).grid(
            row=0, column=1, sticky="w", padx=(6, 6), pady=(6, 2)
        )
        spec_param_group.columnconfigure(0, weight=1)
        rr += 1
        ttk.Label(frm, text="Non-3' SNP policy").grid(row=rr, column=0, sticky="w", pady=2)
        ttk.Combobox(
            frm,
            textvariable=snp_non3p_policy_var,
            values=("soft", "hard"),
            width=10,
            state="readonly",
        ).grid(row=rr, column=1, sticky="w", padx=6, pady=2)
        rr += 1
        ttk.Label(frm, text="3' SNP window").grid(row=rr, column=0, sticky="w", pady=2)
        ttk.Entry(frm, textvariable=snp_3p_window_var, width=10).grid(row=rr, column=1, sticky="w", padx=6, pady=2)

        def _close_preferences() -> None:
            _save_runtime_settings()
            win.destroy()

        btns = ttk.Frame(frm)
        btns.grid(row=rr + 1, column=0, columnspan=3, sticky="e", pady=(8, 0))
        ttk.Button(btns, text="Close", command=_close_preferences).pack(side="right")
        frm.columnconfigure(1, weight=1)
        win.protocol("WM_DELETE_WINDOW", _close_preferences)
        _center_dialog_on_parent(win, root)
        win.grab_set()
        win.focus_force()

    def _build_action_buttons() -> None:
        btns = ttk.Frame(lf_ctrl)
        btns.grid(row=0, column=0, sticky=NSEW)
        lf_ctrl.rowconfigure(0, weight=1)
        for row_idx in range(4):
            btns.rowconfigure(row_idx, weight=1)
        retrieve_btn = ttk.Button(btns, text="Choose Target Gene", command=_retrieve_ensembl, style="Primary.TButton", bootstyle="success")
        retrieve_btn.grid(row=0, column=0, columnspan=2, padx=0, pady=(0, PAD_S // 2), sticky="ew")
        find_btn = ttk.Button(btns, text="Find Primers", command=_run_design, style="Primary.TButton", bootstyle="success")
        find_btn.grid(row=1, column=0, columnspan=2, padx=0, pady=(PAD_S // 2, PAD_S // 2), sticky="ew")
        prefs_btn = ttk.Button(btns, text="Options", command=_open_preferences, bootstyle="secondary")
        prefs_btn.grid(row=2, column=0, padx=(0, PAD_S // 2), pady=(PAD_S // 2, PAD_S // 2), sticky="ew")
        runtime_pie_btn = ttk.Button(btns, text="Runtime Pie", command=_show_runtime_pie, bootstyle="secondary")
        runtime_pie_btn.grid(row=2, column=1, padx=(PAD_S // 2, 0), pady=(PAD_S // 2, PAD_S // 2), sticky="ew")
        primer_blast_btn = ttk.Button(
            btns,
            text="Blast Selected Primer Pair",
            command=_run_spec_on_checked,
            bootstyle="secondary",
        )
        primer_blast_btn.grid(row=3, column=0, columnspan=2, padx=0, pady=(PAD_S // 2, 0), sticky="ew")
        btns.columnconfigure(0, weight=1)
        btns.columnconfigure(1, weight=1)

        filter_btn_row = ttk.Frame(lf_opt)
        filter_btn_row.grid(row=action_row, column=0, columnspan=2, sticky="ew", padx=PAD_S, pady=(PAD_S, 0))
        filter_btn_row.columnconfigure(0, weight=1, uniform="filter_actions")
        filter_btn_row.columnconfigure(1, weight=1, uniform="filter_actions")
        apply_btn = ttk.Button(
            filter_btn_row,
            text="Refresh Filters",
            command=_apply_filter_settings,
            bootstyle="secondary",
        )
        apply_btn.grid(row=0, column=0, padx=(0, 4), pady=0, sticky="ew")
        pie_btn = ttk.Button(filter_btn_row, text="Filter Pie", command=_show_filter_pie, bootstyle="secondary")
        pie_btn.grid(row=0, column=1, padx=(4, 0), pady=0, sticky="ew")
        busy_controls.extend([retrieve_btn, find_btn, apply_btn, pie_btn, runtime_pie_btn, primer_blast_btn, copy_sel_btn, export_ms_btn])

    def _wire_events() -> None:
        tree.bind("<<TreeviewSelect>>", _on_result_select)
        tree.bind("<Button-1>", _toggle_checked_click, add="+")
        map_canvas.bind("<Configure>", _schedule_draw_cdna_map)
        map_canvas.bind("<Motion>", _on_map_hover)
        map_canvas.bind("<Leave>", lambda _evt: _hide_exon_hover())
        gen_text.bind(
            "<KeyRelease>",
            lambda _evt: state.__setitem__("genomic_seq_full", ""),
            add="+",
        )

    def _load_default_sequences() -> None:
        if defaults["genomic"]:
            try:
                _set_genomic_sequence_display(Path(defaults["genomic"]).read_text(encoding="utf-8", errors="replace"))
            except OSError:
                pass
        if defaults["mrna"]:
            try:
                mrna_text.insert("1.0", Path(defaults["mrna"]).read_text(encoding="utf-8", errors="replace"))
            except OSError:
                pass

    def _autosize_window() -> None:
        # Autosize on startup using desktop work area (excludes taskbar where supported)
        # and leave extra vertical headroom so bottom controls stay visible on laptops.
        root.update_idletasks()
        work_w = max(800, int(root.winfo_vrootwidth() or root.winfo_screenwidth()))
        work_h = max(600, int(root.winfo_vrootheight() or root.winfo_screenheight()))
        work_x = int(root.winfo_vrootx() or 0)
        work_y = int(root.winfo_vrooty() or 0)

        req_w = max(980, root.winfo_reqwidth())
        req_h = max(680, root.winfo_reqheight())

        max_w = max(900, work_w - 24)
        max_h = max(520, work_h - 56)

        win_w = min(req_w, int(work_w * 0.96), max_w)
        win_h = min(req_h, int(work_h * 0.88), max_h)

        pos_x = work_x + max(0, (work_w - win_w) // 2)
        pos_y = work_y + max(0, (work_h - win_h) // 2)
        root.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")

    _build_action_buttons()
    _wire_events()
    _detect_and_apply_cpu_profile()
    _sync_gene_display()
    target_ensembl_gene_id_var.trace_add("write", _update_spec_toggle_availability)
    _update_spec_toggle_availability()
    _load_default_sequences()
    startup_code = _startup_toolchain_checks()
    if startup_code != 0:
        root.destroy()
        return startup_code
    _autosize_window()

    _safe_draw_cdna_map()

    root.mainloop()
    return 0
